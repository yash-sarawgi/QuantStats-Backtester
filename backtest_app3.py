#!/usr/bin/env python3
"""
QuantStats Strategy Backtesting Report Generator  v3.0
+ TradingView XLSX Full Tearsheet Analyzer

FEATURES:
  - Run backtest on Alpaca API or CSV data → full QuantStats HTML tearsheet
  - Configurable benchmark ticker (default SPY) fetched via yfinance
  - TradingView .xlsx export → converts trade P&L → daily returns →
    full QuantStats tearsheet identical to the backtest module
  - All timezone / dtype issues resolved
"""

# ─── AUTO-INSTALL ────────────────────────────────────────────────────────────
import subprocess, sys

REQUIRED = {
    "quantstats": "quantstats",
    "pandas":     "pandas",
    "numpy":      "numpy",
    "pytz":       "pytz",
    "openpyxl":   "openpyxl",
    "yfinance":   "yfinance",
}

def _install_all():
    for mod, pkg in REQUIRED.items():
        try:
            __import__(mod)
        except ImportError:
            print(f"Installing {pkg}…")
            subprocess.check_call(
                [sys.executable, "-m", "pip", "install", pkg, "-q",
                 "--break-system-packages"],
                stderr=subprocess.DEVNULL)

_install_all()

# ─── IMPORTS ────────────────────────────────────────────────────────────────
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading, traceback, os, tempfile, webbrowser, re
import pandas as pd
import numpy as np
import pytz
import quantstats as qs
import yfinance as yf
from datetime import datetime, timedelta
import openpyxl

# ─── THEME ──────────────────────────────────────────────────────────────────
BG     = "#0d1117"
BG2    = "#161b22"
BG3    = "#21262d"
BORDER = "#30363d"
ACCENT = "#58a6ff"
GREEN  = "#3fb950"
WARN   = "#d29922"
RED    = "#f85149"
PURPLE = "#bc8cff"
ORANGE = "#ffa657"
FG     = "#e6edf3"
FG2    = "#8b949e"

_P   = sys.platform
MONO = ("Cascadia Code", 10) if _P == "win32" else ("Menlo", 10)
UI   = ("Segoe UI",      10) if _P == "win32" else ("SF Pro Display", 10)
HEAD = ("Segoe UI Semibold", 11) if _P == "win32" else ("SF Pro Display", 11)

TF_MAP = {
    "1 Min": "1Min", "5 Min": "5Min", "15 Min": "15Min",
    "30 Min": "30Min", "1 Hour": "1Hour", "4 Hour": "4Hour", "1 Day": "1Day",
}

DEFAULT_STRATEGY = '''\
# ══════════════════════════════════════════════════════════════════════════
# STRATEGY CODE  —  Required Function Signature:
#
#   def strategy(data: pd.DataFrame) -> pd.Series
#
# INPUT  `data`  (index = UTC-aware DatetimeIndex):
#   open, high, low, close  – OHLC float
#   volume                  – float
#   vwap                    – float
#
# OUTPUT  pd.Series of SIGNALS aligned to data.index:
#   +1 = Long  |  0 = Flat  |  -1 = Short
#
# Signals are shifted 1 bar to avoid look-ahead bias
# (signal at bar N fills at bar N+1 open price).
# ══════════════════════════════════════════════════════════════════════════

import pandas as pd
import numpy as np

def strategy(data: pd.DataFrame) -> pd.Series:
    """Golden Cross: SMA-20 above SMA-50 → long, else flat."""
    close = data["close"]
    fast  = close.rolling(20).mean()
    slow  = close.rolling(50).mean()
    sig   = pd.Series(0, index=data.index, dtype=float)
    sig[fast > slow] = 1
    return sig
'''

# ════════════════════════════════════════════════════════════════════════════
# HELPERS
# ════════════════════════════════════════════════════════════════════════════

def _strip_tz(s: pd.Series) -> pd.Series:
    """
    Convert UTC-aware DatetimeIndex → tz-naive.
    Fixes: 'Cannot compare dtypes datetime64[us, UTC] and datetime64[us]'
    that QuantStats raises when the index still carries timezone info.
    """
    if hasattr(s.index, "tz") and s.index.tz is not None:
        s = s.copy()
        s.index = s.index.tz_convert("UTC").tz_localize(None)
    return s


def _safe_qs(fn, *a, **kw):
    try:    return fn(*a, **kw)
    except: return float("nan")


def fetch_benchmark(ticker: str, start: str, end: str) -> pd.Series:
    """
    Download daily close returns for the benchmark via yfinance.
    Returns a tz-naive daily return Series.
    Raises a descriptive error if the ticker is invalid / no data.
    """
    # Pad end by 5 days so we cover weekends/holidays
    end_pad = (datetime.strptime(end, "%Y-%m-%d") + timedelta(days=5)).strftime("%Y-%m-%d")
    raw = yf.download(ticker, start=start, end=end_pad,
                      auto_adjust=True, progress=False)
    if raw.empty:
        raise ValueError(
            f"Benchmark '{ticker}': no data returned from yfinance.\n"
            f"Check the ticker symbol and date range.")
    close = raw["Close"]
    if isinstance(close, pd.DataFrame):
        close = close.iloc[:, 0]
    close.index = pd.to_datetime(close.index).tz_localize(None)
    rets = close.pct_change().dropna()
    rets.name = ticker.upper()
    return rets

# ════════════════════════════════════════════════════════════════════════════
# DATA LAYER
# ════════════════════════════════════════════════════════════════════════════

def detect_and_load_csv(filepath: str) -> pd.DataFrame:
    df = pd.read_csv(filepath)
    df.columns = [c.strip().lower() for c in df.columns]
    ts_candidates = ["timestamp","datetime","date","time","t",
                     "date_time","bar_time","open_time"]
    ts_col = next((c for c in ts_candidates if c in df.columns), None)
    if ts_col is None:
        for c in df.columns:
            if re.search(r"\d{4}-\d{2}-\d{2}", str(df[c].iloc[0])):
                ts_col = c; break
    if ts_col is None:
        raise ValueError("Cannot detect timestamp column. "
                         "Expected: 'timestamp', 'datetime', or 'date'.")
    df["timestamp"] = pd.to_datetime(df[ts_col], utc=True, errors="coerce")
    df = df.dropna(subset=["timestamp"]).set_index("timestamp").sort_index()
    aliases = {
        "open":   ["open","o","open_price"],
        "high":   ["high","h","high_price"],
        "low":    ["low","l","low_price"],
        "close":  ["close","c","last","adj_close","close_price"],
        "volume": ["volume","vol","v","qty","quantity"],
        "vwap":   ["vwap","vw","weighted_avg"],
    }
    for canon, alts in aliases.items():
        if canon not in df.columns:
            for a in alts:
                if a in df.columns:
                    df.rename(columns={a: canon}, inplace=True); break
    miss = [r for r in ["open","high","low","close"] if r not in df.columns]
    if miss:
        raise ValueError(f"CSV missing required columns: {miss}. "
                         f"Found: {list(df.columns)}")
    for col in ["volume","vwap"]:
        if col not in df.columns:
            df[col] = df["close"] if col == "vwap" else 0.0
    for col in ["open","high","low","close","volume","vwap"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    return df[["open","high","low","close","volume","vwap"]].dropna(
        subset=["open","high","low","close"])


def fetch_alpaca_data(ticker, start, end, tf_label, api_key, secret_key):
    try:
        from alpaca.data.historical import StockHistoricalDataClient
        from alpaca.data.requests   import StockBarsRequest
        from alpaca.data.timeframe  import TimeFrame, TimeFrameUnit
    except ImportError:
        raise RuntimeError("alpaca-py not installed. Run: pip install alpaca-py")
    if not api_key or not secret_key:
        raise ValueError("Alpaca API Key and Secret Key are required.")
    tf_alpaca = {
        "1Min":  TimeFrame(1, TimeFrameUnit.Minute),
        "5Min":  TimeFrame(5, TimeFrameUnit.Minute),
        "15Min": TimeFrame(15,TimeFrameUnit.Minute),
        "30Min": TimeFrame(30,TimeFrameUnit.Minute),
        "1Hour": TimeFrame(1, TimeFrameUnit.Hour),
        "4Hour": TimeFrame(4, TimeFrameUnit.Hour),
        "1Day":  TimeFrame(1, TimeFrameUnit.Day),
    }
    tf_str = TF_MAP.get(tf_label, "1Day")
    tz_ny  = pytz.timezone("America/New_York")
    s_dt   = datetime.strptime(start, "%Y-%m-%d").replace(tzinfo=tz_ny)
    e_dt   = datetime.strptime(end,   "%Y-%m-%d").replace(
                 hour=23, minute=59, second=59, tzinfo=tz_ny)
    client = StockHistoricalDataClient(api_key=api_key, secret_key=secret_key)
    req    = StockBarsRequest(symbol_or_symbols=ticker.upper(),
                              timeframe=tf_alpaca[tf_str],
                              start=s_dt, end=e_dt, adjustment="all")
    df = client.get_stock_bars(req).df
    if df.empty:
        raise ValueError(f"No data returned for {ticker} in [{start} → {end}].")
    if isinstance(df.index, pd.MultiIndex):
        df = df.reset_index(level=0, drop=True)
    df.index = pd.to_datetime(df.index, utc=True)
    df.index.name = "timestamp"
    df.columns = [c.lower() for c in df.columns]
    if "volume" not in df.columns: df["volume"] = 0.0
    if "vwap"   not in df.columns: df["vwap"]   = df["close"]
    return df[["open","high","low","close","volume","vwap"]].sort_index()


# ════════════════════════════════════════════════════════════════════════════
# BACKTEST ENGINE
# ════════════════════════════════════════════════════════════════════════════

def run_backtest(data: pd.DataFrame, code: str, name: str,
                 benchmark_ticker: str = "SPY") -> dict:
    """
    Execute strategy code → daily returns → QuantStats HTML tearsheet.
    Benchmark is fetched from yfinance (default SPY).
    Returns dict with summary stats, html_path, and the daily returns series.
    """
    # ── Execute strategy code ────────────────────────────────────────────
    ns = {"pd": pd, "np": np, "__builtins__": __builtins__}
    try:
        exec(compile(code, "<strategy>", "exec"), ns)
    except Exception as e:
        raise RuntimeError(f"Strategy code compilation error:\n{e}")
    if "strategy" not in ns:
        raise ValueError("Strategy code must define `strategy(data)` function.")
    try:
        sigs = ns["strategy"](data.copy())
    except Exception:
        raise RuntimeError(f"Strategy execution error:\n{traceback.format_exc()}")
    if not isinstance(sigs, pd.Series):
        raise ValueError("`strategy()` must return a pandas Series of signals.")

    sigs  = sigs.reindex(data.index).fillna(0).shift(1).fillna(0)
    pret  = data["close"].pct_change().fillna(0)
    port  = (sigs * pret).rename(name)
    port.index = pd.to_datetime(port.index, utc=True)

    daily = port.resample("D").apply(lambda x: (1+x).prod()-1).dropna()
    daily = daily[daily.abs() < 1]
    if daily.empty:
        raise ValueError("Backtest produced no returns. "
                         "Check date range and strategy logic.")

    # Date range for benchmark
    d_start = daily.index[0].strftime("%Y-%m-%d")
    d_end   = daily.index[-1].strftime("%Y-%m-%d")

    # ── Fetch benchmark ──────────────────────────────────────────────────
    bench_name = (benchmark_ticker.strip().upper() or "SPY")
    try:
        bench_raw = fetch_benchmark(bench_name, d_start, d_end)
        bench     = bench_raw.reindex(
            daily.index.tz_localize(None), method="ffill").fillna(0)
        bench.name = bench_name
    except Exception as e:
        # Fallback: buy-and-hold if benchmark fetch fails
        bench = (pret.resample("D")
                     .apply(lambda x: (1+x).prod()-1)
                     .dropna()
                     .reindex(daily.index.tz_localize(None))
                     .fillna(0))
        bench.name = f"Buy & Hold ({name})"
        print(f"[WARN] Benchmark fetch failed ({e}), using buy-and-hold.")

    # ── Strip timezone for QuantStats ────────────────────────────────────
    daily_qs = _strip_tz(daily)
    bench_qs = bench  # already tz-naive from yfinance

    # Align index
    idx = daily_qs.index.intersection(bench_qs.index)
    if len(idx) == 0:
        # Indexes don't overlap (e.g. intraday strategy with daily benchmark)
        # Just pass bench without alignment
        bench_qs = bench_qs.reindex(daily_qs.index, method="ffill").fillna(0)
    else:
        daily_qs = daily_qs.reindex(idx)
        bench_qs = bench_qs.reindex(idx).fillna(0)

    # ── Generate HTML tearsheet ──────────────────────────────────────────
    html = os.path.join(tempfile.gettempdir(),
        f"bt_{re.sub(r'[^a-zA-Z0-9]','_',name)}"
        f"_{datetime.now():%Y%m%d_%H%M%S}.html")

    qs.extend_pandas()
    qs.reports.html(
        returns=daily_qs,
        benchmark=bench_qs,
        title=f"{name}  vs  {bench_name}",
        output=html,
        download_filename=html,
    )

    summary = {
        "Total Return":      f"{_safe_qs(qs.stats.comp, daily_qs)*100:.2f}%",
        "CAGR":              f"{_safe_qs(qs.stats.cagr, daily_qs)*100:.2f}%",
        "Sharpe Ratio":      f"{_safe_qs(qs.stats.sharpe, daily_qs):.3f}",
        "Sortino Ratio":     f"{_safe_qs(qs.stats.sortino, daily_qs):.3f}",
        "Max Drawdown":      f"{_safe_qs(qs.stats.max_drawdown, daily_qs)*100:.2f}%",
        "Volatility (Ann)":  f"{_safe_qs(qs.stats.volatility, daily_qs)*100:.2f}%",
        "Win Rate":          f"{_safe_qs(qs.stats.win_rate, daily_qs)*100:.2f}%",
        "Calmar Ratio":      f"{_safe_qs(qs.stats.calmar, daily_qs):.3f}",
        "Skew":              f"{_safe_qs(qs.stats.skew, daily_qs):.3f}",
        "Kurtosis":          f"{_safe_qs(qs.stats.kurtosis, daily_qs):.3f}",
        "VaR (95%)":         f"{_safe_qs(qs.stats.value_at_risk, daily_qs)*100:.2f}%",
        "CVaR (95%)":        f"{_safe_qs(qs.stats.cvar, daily_qs)*100:.2f}%",
        f"Benchmark ({bench_name}) Return":
            f"{_safe_qs(qs.stats.comp, bench_qs)*100:.2f}%",
        "Bars in Market":    f"{int((sigs.abs()>0).sum()):,}",
    }
    return dict(summary=summary, html_path=html,
                daily_rets=daily, bench_rets=bench_qs,
                bench_name=bench_name, data_rows=len(data))


# ════════════════════════════════════════════════════════════════════════════
# TRADINGVIEW XLSX LOADER + TEARSHEET
# ════════════════════════════════════════════════════════════════════════════

def load_tv_xlsx(filepath: str) -> dict:
    """
    Parse TradingView strategy export (.xlsx).
    Sheets: List of trades / Performance / Trades analysis /
            Risk-adjusted performance / Properties
    """
    wb   = openpyxl.load_workbook(filepath, data_only=True)
    smap = {s.lower(): s for s in wb.sheetnames}

    def _df(key):
        name = next((v for k,v in smap.items() if key in k), None)
        if not name: return pd.DataFrame()
        ws   = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows: return pd.DataFrame()
        hdrs = [str(h) if h is not None else f"c{i}"
                for i,h in enumerate(rows[0])]
        return pd.DataFrame(rows[1:], columns=hdrs)

    def _kv(key):
        name = next((v for k,v in smap.items() if key in k), None)
        if not name: return {}
        ws  = wb[name]
        out = {}
        for row in ws.iter_rows(values_only=True):
            if row[0] is not None:
                out[str(row[0])] = row[1:]
        return out

    lot = _df("list of trade")
    lot.columns = [c.strip() for c in lot.columns]
    col_map = {
        "Trade #":"trade_num","Type":"type","Date and time":"datetime",
        "Signal":"signal","Price USD":"price","Position size (qty)":"qty",
        "Position size (value)":"value","Net P&L USD":"pnl_usd",
        "Net P&L %":"pnl_pct","Favorable excursion USD":"mfe_usd",
        "Favorable excursion %":"mfe_pct","Adverse excursion USD":"mae_usd",
        "Adverse excursion %":"mae_pct","Cumulative P&L USD":"cum_pnl_usd",
        "Cumulative P&L %":"cum_pnl_pct",
    }
    lot.rename(columns={k:v for k,v in col_map.items() if k in lot.columns},
               inplace=True)
    lot["datetime"] = pd.to_datetime(lot["datetime"], errors="coerce")
    for col in ["price","qty","value","pnl_usd","pnl_pct","mfe_usd","mfe_pct",
                "mae_usd","mae_pct","cum_pnl_usd","cum_pnl_pct"]:
        if col in lot.columns:
            lot[col] = pd.to_numeric(lot[col], errors="coerce")

    ent = lot[lot["type"].str.lower().str.contains("entry", na=False)].copy()
    ext = lot[lot["type"].str.lower().str.contains("exit",  na=False)].copy()
    ent = ent.sort_values("trade_num").drop_duplicates("trade_num", keep="last")
    ext = ext.sort_values("trade_num").drop_duplicates("trade_num", keep="last")
    trades = ent.merge(
        ext[["trade_num","datetime","price","pnl_usd","pnl_pct",
             "mfe_usd","mfe_pct","mae_usd","mae_pct","cum_pnl_usd","signal"]],
        on="trade_num", suffixes=("_entry","_exit"), how="inner")
    trades["duration"]  = trades["datetime_exit"] - trades["datetime_entry"]
    trades["direction"] = trades["type"].str.extract(
        r"(long|short)", flags=re.I)[0].str.lower()
    trades.rename(columns={"pnl_usd_exit":"pnl_usd",
                            "pnl_pct_exit":"pnl_pct"}, inplace=True)
    trades = trades.reset_index(drop=True)

    pnl = trades["pnl_usd"].dropna()
    w   = pnl[pnl > 0]; l = pnl[pnl < 0]; e = pnl[pnl == 0]

    trades["exit_date"] = pd.to_datetime(
        trades["datetime_exit"], errors="coerce").dt.normalize()
    daily_pnl = trades.groupby("exit_date")["pnl_usd"].sum()
    equity    = daily_pnl.cumsum()
    drawdown  = equity - equity.cummax()

    props = _kv("properties")
    # Initial capital from Properties sheet (key = "Initial capital")
    init_cap = 100_000.0
    try:
        cap_raw = props.get("Initial capital", (None,))[0]
        if cap_raw is not None:
            init_cap = float(str(cap_raw).replace(",",""))
    except Exception:
        pass

    derived = dict(
        total_trades  = len(trades),
        winners       = len(w),
        losers        = len(l),
        evens         = len(e),
        win_rate      = len(w) / max(len(trades)-len(e), 1) * 100,
        avg_win       = float(w.mean())  if len(w) else 0.0,
        avg_loss      = float(l.mean())  if len(l) else 0.0,
        total_pnl     = float(pnl.sum()),
        gross_profit  = float(w.sum())   if len(w) else 0.0,
        gross_loss    = float(l.sum())   if len(l) else 0.0,
        profit_factor = abs(w.sum()/l.sum()) if l.sum() != 0 else float("inf"),
        largest_win   = float(w.max())   if len(w) else 0.0,
        largest_loss  = float(l.min())   if len(l) else 0.0,
        max_drawdown  = float(drawdown.min()),
        avg_duration  = trades["duration"].mean(),
        equity        = equity,
        drawdown      = drawdown,
        daily_pnl     = daily_pnl,
        init_cap      = init_cap,
    )
    return dict(trades=trades, lot_raw=lot,
                perf=_kv("performance"), tana=_kv("trades analysis"),
                risk=_kv("risk-adjusted"), props=props, derived=derived)


def build_tv_tearsheet(data: dict, name: str,
                        benchmark_ticker: str = "SPY") -> str:
    """
    Convert TradingView trade P&L → daily return series →
    full QuantStats HTML tearsheet with real benchmark.
    Returns path to the generated HTML file.
    """
    d         = data["derived"]
    daily_pnl = d["daily_pnl"].copy()
    init_cap  = d["init_cap"]

    # ── Build daily returns from P&L ─────────────────────────────────────
    # equity[t] = init_cap + cumulative P&L up to day t
    # return[t] = P&L[t] / equity[t-1]
    cum_pnl   = daily_pnl.cumsum()
    equity    = init_cap + cum_pnl
    equity_lag = equity.shift(1).fillna(init_cap)
    daily_ret  = (daily_pnl / equity_lag).rename(name)

    # Ensure DatetimeIndex, tz-naive
    daily_ret.index = pd.to_datetime(daily_ret.index).tz_localize(None)
    daily_ret = daily_ret[daily_ret.abs() < 1].dropna()

    if daily_ret.empty:
        raise ValueError(
            "Could not compute daily returns from trade data. "
            "Check that the 'List of trades' sheet has P&L values.")

    d_start = daily_ret.index[0].strftime("%Y-%m-%d")
    d_end   = daily_ret.index[-1].strftime("%Y-%m-%d")

    # ── Fetch benchmark ──────────────────────────────────────────────────
    bench_name = (benchmark_ticker.strip().upper() or "SPY")
    try:
        bench_raw = fetch_benchmark(bench_name, d_start, d_end)
        bench     = bench_raw.reindex(daily_ret.index, method="ffill").fillna(0)
        bench.name = bench_name
    except Exception as e:
        bench      = pd.Series(0.0, index=daily_ret.index, name="Flat Bench")
        print(f"[WARN] Benchmark fetch failed: {e}. Using flat benchmark.")

    # ── Generate tearsheet ───────────────────────────────────────────────
    html = os.path.join(tempfile.gettempdir(),
        f"tv_{re.sub(r'[^a-zA-Z0-9]','_',name)}"
        f"_{datetime.now():%Y%m%d_%H%M%S}.html")

    qs.extend_pandas()
    qs.reports.html(
        returns=daily_ret,
        benchmark=bench,
        title=f"{name}  vs  {bench_name}  [TradingView]",
        output=html,
        download_filename=html,
    )

    # Quick summary stats
    summary = {
        "Total Return":      f"{_safe_qs(qs.stats.comp, daily_ret)*100:.2f}%",
        "CAGR":              f"{_safe_qs(qs.stats.cagr, daily_ret)*100:.2f}%",
        "Sharpe Ratio":      f"{_safe_qs(qs.stats.sharpe, daily_ret):.3f}",
        "Sortino Ratio":     f"{_safe_qs(qs.stats.sortino, daily_ret):.3f}",
        "Max Drawdown":      f"{_safe_qs(qs.stats.max_drawdown, daily_ret)*100:.2f}%",
        "Volatility (Ann)":  f"{_safe_qs(qs.stats.volatility, daily_ret)*100:.2f}%",
        "Win Rate":          f"{_safe_qs(qs.stats.win_rate, daily_ret)*100:.2f}%",
        "Calmar Ratio":      f"{_safe_qs(qs.stats.calmar, daily_ret):.3f}",
        "Skew":              f"{_safe_qs(qs.stats.skew, daily_ret):.3f}",
        "Kurtosis":          f"{_safe_qs(qs.stats.kurtosis, daily_ret):.3f}",
        "VaR (95%)":         f"{_safe_qs(qs.stats.value_at_risk, daily_ret)*100:.2f}%",
        "CVaR (95%)":        f"{_safe_qs(qs.stats.cvar, daily_ret)*100:.2f}%",
        f"Benchmark ({bench_name})":
            f"{_safe_qs(qs.stats.comp, bench)*100:.2f}%",
        "Net P&L $":         f"${d['total_pnl']:,.2f}",
        "Total Trades":      str(d["total_trades"]),
        "Win Rate (Trades)": f"{d['win_rate']:.1f}%",
        "Profit Factor":
            f"{d['profit_factor']:.3f}" if d["profit_factor"] != float("inf") else "∞",
        "Max Drawdown $":    f"${d['max_drawdown']:,.2f}",
        "Largest Win":       f"${d['largest_win']:,.2f}",
        "Largest Loss":      f"${abs(d['largest_loss']):,.2f}",
    }
    return html, summary, daily_ret, bench


# ════════════════════════════════════════════════════════════════════════════
# GUI
# ════════════════════════════════════════════════════════════════════════════

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("QuantStats Backtest + TradingView Analyzer  v3.0")
        self.configure(bg=BG)
        self.geometry("1420x920")
        self.minsize(1150, 720)

        # ── StringVars ────────────────────────────────────────────────────
        self._csv_path   = tk.StringVar()
        self._ticker     = tk.StringVar(value="AAPL")
        self._benchmark  = tk.StringVar(value="SPY")
        self._start      = tk.StringVar(value="2022-01-01")
        self._end        = tk.StringVar(value=datetime.today().strftime("%Y-%m-%d"))
        self._tf         = tk.StringVar(value="1 Day")
        self._src        = tk.StringVar(value="alpaca")
        self._api_key    = tk.StringVar()
        self._sec_key    = tk.StringVar()
        self._strat_name = tk.StringVar(value="My Strategy")
        self._status     = tk.StringVar(value="Ready.")
        self._bt_html    = None    # backtest tearsheet path
        self._tv_html    = None    # TV tearsheet path
        self._tv_path    = tk.StringVar()
        self._tv_name    = tk.StringVar(value="TV Strategy")
        self._tv_bench   = tk.StringVar(value="SPY")

        self._sty()
        self._build()

    # ── Styles ────────────────────────────────────────────────────────────
    def _sty(self):
        s = ttk.Style(self)
        s.theme_use("clam")
        base = dict(background=BG3, foreground=FG, fieldbackground=BG3,
                    insertbackground=FG, selectbackground=ACCENT,
                    selectforeground=BG, bordercolor=BORDER,
                    lightcolor=BORDER, darkcolor=BORDER, relief="flat", padding=5)
        s.configure("TFrame",        background=BG)
        s.configure("TLabel",        background=BG,  foreground=FG, font=UI)
        s.configure("TEntry",        **base)
        s.configure("TCombobox",     **base)
        s.map("TCombobox",
              fieldbackground=[("readonly",BG3)],
              background=[("readonly",BG3)])
        s.configure("TButton",       background=BG3, foreground=FG,
                    bordercolor=BORDER, font=UI, padding=(10,5), relief="flat")
        s.map("TButton",   background=[("active",BORDER)])
        s.configure("Run.TButton",   background=ACCENT, foreground=BG,
                    font=(UI[0],11,"bold"), padding=(16,8))
        s.map("Run.TButton",  background=[("active","#79c0ff")])
        s.configure("Open.TButton",  background=GREEN, foreground=BG,
                    font=(UI[0],10,"bold"), padding=(12,6))
        s.map("Open.TButton", background=[("active","#56d364")])
        s.configure("TV.TButton",    background=PURPLE, foreground=BG,
                    font=(UI[0],10,"bold"), padding=(12,6))
        s.map("TV.TButton",   background=[("active","#d2a8ff")])
        s.configure("TVTs.TButton",  background="#6e40c9", foreground=FG,
                    font=(UI[0],10,"bold"), padding=(12,6))
        s.map("TVTs.TButton", background=[("active","#8957e5")])
        s.configure("TRadiobutton",  background=BG2, foreground=FG,
                    indicatorbackground=BG3, selectcolor=ACCENT)
        s.configure("TNotebook",     background=BG, bordercolor=BORDER, tabmargins=0)
        s.configure("TNotebook.Tab", background=BG3, foreground=FG2,
                    padding=(14,7), bordercolor=BORDER)
        s.map("TNotebook.Tab",
              background=[("selected",BG2)], foreground=[("selected",FG)])
        s.configure("TProgressbar",  troughcolor=BG3, background=ACCENT,
                    bordercolor=BORDER, thickness=4)
        s.configure("Treeview",
                    background=BG3, foreground=FG, fieldbackground=BG3,
                    bordercolor=BORDER, relief="flat", rowheight=22,
                    font=("Cascadia Code",9) if _P=="win32" else ("Menlo",9))
        s.configure("Treeview.Heading",
                    background=BG2, foreground=FG2,
                    font=(UI[0],9,"bold"), relief="flat")
        s.map("Treeview",
              background=[("selected",ACCENT)],
              foreground=[("selected",BG)])

    # ── Layout ────────────────────────────────────────────────────────────
    def _build(self):
        hdr = tk.Frame(self, bg=BG2, height=54)
        hdr.pack(fill="x"); hdr.pack_propagate(False)
        tk.Label(hdr,
                 text="⚡  QuantStats Backtest  +  TradingView Analyzer  v3.0",
                 bg=BG2, fg=FG,
                 font=(HEAD[0],14,"bold")).pack(side="left", padx=20, pady=10)
        tk.Label(hdr, text="benchmark via yfinance  ·  full tearsheet",
                 bg=BG2, fg=FG2, font=("",9)).pack(side="right", padx=20)
        ttk.Separator(self).pack(fill="x")

        main = tk.Frame(self, bg=BG); main.pack(fill="both", expand=True)

        left = tk.Frame(main, bg=BG, width=400)
        left.pack(side="left", fill="y"); left.pack_propagate(False)
        self._left_panel(left)

        tk.Frame(main, bg=BORDER, width=1).pack(side="left", fill="y")

        right = tk.Frame(main, bg=BG)
        right.pack(side="left", fill="both", expand=True)
        self._right_panel(right)

        sb = tk.Frame(self, bg=BG3, height=26)
        sb.pack(fill="x", side="bottom"); sb.pack_propagate(False)
        self._prog = ttk.Progressbar(sb, mode="indeterminate", length=140)
        self._prog.pack(side="right", padx=10, pady=5)
        tk.Label(sb, textvariable=self._status,
                 bg=BG3, fg=FG2, font=("",9), anchor="w"
                 ).pack(side="left", padx=10, pady=4)

    # ── Left panel ────────────────────────────────────────────────────────
    def _left_panel(self, parent):
        cvs = tk.Canvas(parent, bg=BG, bd=0, highlightthickness=0)
        vsb = ttk.Scrollbar(parent, orient="vertical", command=cvs.yview)
        inn = tk.Frame(cvs, bg=BG)
        inn.bind("<Configure>",
                 lambda e: cvs.configure(scrollregion=cvs.bbox("all")))
        cvs.create_window((0,0), window=inn, anchor="nw")
        cvs.configure(yscrollcommand=vsb.set)
        cvs.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        M = 14

        def sec(t, col=ACCENT):
            f = tk.Frame(inn, bg=BG2)
            f.pack(fill="x", padx=M, pady=(10,0))
            tk.Label(f, text=t, bg=BG2, fg=col,
                     font=(HEAD[0],10,"bold")).pack(anchor="w", padx=10, pady=(8,4))
            return f

        def erow(p, lbl, var, show=None, w=14):
            r = tk.Frame(p, bg=BG2); r.pack(fill="x", padx=10, pady=3)
            tk.Label(r, text=lbl, bg=BG2, fg=FG2, font=("",9),
                     width=w, anchor="w").pack(side="left")
            ttk.Entry(r, textvariable=var, font=MONO, show=show
                      ).pack(side="left", fill="x", expand=True)

        # ── Strategy ──────────────────────────────────────────────────────
        f0 = sec("◈  STRATEGY")
        erow(f0, "Name", self._strat_name)

        # ── Benchmark ─────────────────────────────────────────────────────
        f_bench = sec("◈  BENCHMARK")
        erow(f_bench, "Ticker (SPY)", self._benchmark)
        tk.Label(f_bench,
                 text="  Fetched via yfinance — any valid ticker\n"
                      "  e.g. SPY, QQQ, BTC-USD, ^GSPC",
                 bg=BG2, fg=FG2, font=("",8),
                 justify="left").pack(anchor="w", padx=10, pady=(0,8))

        # ── Data source ────────────────────────────────────────────────────
        f1 = sec("◈  DATA SOURCE")
        rb = tk.Frame(f1, bg=BG2); rb.pack(fill="x", padx=10, pady=4)
        for txt, val in [("Alpaca API","alpaca"),("CSV File","csv")]:
            ttk.Radiobutton(rb, text=txt, variable=self._src, value=val,
                            command=self._toggle_src).pack(side="left", padx=(0,14))

        self._alpaca_f = tk.Frame(f1, bg=BG2); self._alpaca_f.pack(fill="x")
        erow(self._alpaca_f, "API Key",    self._api_key)
        erow(self._alpaca_f, "Secret Key", self._sec_key, show="•")
        tk.Label(self._alpaca_f, text="  alpaca.markets → Paper Trading → API Keys",
                 bg=BG2, fg=FG2, font=("",8)).pack(anchor="w", padx=10, pady=(0,6))

        self._csv_f = tk.Frame(f1, bg=BG2)
        cr = tk.Frame(self._csv_f, bg=BG2); cr.pack(fill="x", padx=10, pady=4)
        tk.Label(cr, text="CSV File", bg=BG2, fg=FG2, font=("",9),
                 width=14, anchor="w").pack(side="left")
        ttk.Entry(cr, textvariable=self._csv_path,
                  font=("",9)).pack(side="left", fill="x", expand=True, padx=(0,4))
        ttk.Button(cr, text="Browse",
                   command=lambda: self._browse(self._csv_path)).pack(side="left")
        tk.Label(self._csv_f,
                 text="  Columns: timestamp, open, high, low, close, volume",
                 bg=BG2, fg=FG2, font=("",8)).pack(anchor="w", padx=10, pady=(0,6))

        # ── Ticker / Timeframe ─────────────────────────────────────────────
        f2 = sec("◈  TICKER & TIMEFRAME")
        erow(f2, "Ticker", self._ticker)
        r = tk.Frame(f2, bg=BG2); r.pack(fill="x", padx=10, pady=3)
        tk.Label(r, text="Timeframe", bg=BG2, fg=FG2, font=("",9),
                 width=14, anchor="w").pack(side="left")
        ttk.Combobox(r, textvariable=self._tf, state="readonly",
                     values=list(TF_MAP.keys()), font=UI
                     ).pack(side="left", fill="x", expand=True)
        erow(f2, "Start Date", self._start)
        erow(f2, "End Date",   self._end)
        tk.Label(f2, text="  Format: YYYY-MM-DD", bg=BG2, fg=FG2,
                 font=("",8)).pack(anchor="w", padx=10, pady=(0,8))

        # ── Backtest run ───────────────────────────────────────────────────
        bf = tk.Frame(inn, bg=BG); bf.pack(fill="x", padx=M, pady=(12,0))
        ttk.Button(bf, text="▶  Run Backtest", style="Run.TButton",
                   command=self._run_bt).pack(fill="x")
        self._bt_open_btn = ttk.Button(
            bf, text="🌐  Open QuantStats Tearsheet",
            style="Open.TButton", command=self._open_bt_report, state="disabled")
        self._bt_open_btn.pack(fill="x", pady=(7,0))

        # ─────────────────────────────────────────────────────────────────
        ttk.Separator(inn).pack(fill="x", padx=M, pady=(16,0))

        # ── TradingView Analyzer ───────────────────────────────────────────
        f3 = sec("◈  TRADINGVIEW ANALYZER", col=PURPLE)
        tk.Label(f3,
                 text="  Export from TradingView:\n"
                      "  Strategy Tester → Export → Excel",
                 bg=BG2, fg=FG2, font=("",8), justify="left"
                 ).pack(anchor="w", padx=10, pady=(0,4))
        tvr = tk.Frame(f3, bg=BG2); tvr.pack(fill="x", padx=10, pady=(0,4))
        ttk.Entry(tvr, textvariable=self._tv_path,
                  font=("",9)).pack(side="left", fill="x", expand=True, padx=(0,4))
        ttk.Button(tvr, text="Browse",
                   command=lambda: self._browse(
                       self._tv_path,
                       [("Excel XLSX","*.xlsx"),("All","*.*")]
                   )).pack(side="left")
        erow(f3, "Strategy Name", self._tv_name)
        erow(f3, "Benchmark",     self._tv_bench)

        tv_bf = tk.Frame(inn, bg=BG); tv_bf.pack(fill="x", padx=M, pady=(8,0))
        ttk.Button(tv_bf,
                   text="📊  Analyze & Build Tearsheet",
                   style="TV.TButton",
                   command=self._run_tv).pack(fill="x")
        self._tv_open_btn = ttk.Button(
            tv_bf,
            text="🌐  Open TV Tearsheet",
            style="TVTs.TButton",
            command=self._open_tv_report, state="disabled")
        self._tv_open_btn.pack(fill="x", pady=(7,0))

        tk.Frame(inn, bg=BG, height=16).pack()
        self._toggle_src()

    # ── Right panel ────────────────────────────────────────────────────────
    def _right_panel(self, parent):
        self._nb = ttk.Notebook(parent)
        self._nb.pack(fill="both", expand=True)

        # Editor
        et = tk.Frame(self._nb, bg=BG2)
        self._nb.add(et, text="  Strategy Editor  ")
        tb = tk.Frame(et, bg=BG3, height=34); tb.pack(fill="x"); tb.pack_propagate(False)
        tk.Label(tb, text="strategy.py", bg=BG3, fg=FG2, font=MONO
                 ).pack(side="left", padx=12, pady=7)
        ttk.Button(tb, text="Reset", command=self._reset_ed
                   ).pack(side="right", padx=8, pady=5)
        self._ed = scrolledtext.ScrolledText(
            et, bg="#0d1117", fg="#e6edf3", font=MONO,
            insertbackground=FG, selectbackground=ACCENT, selectforeground=BG,
            relief="flat", bd=0, wrap="none", undo=True, tabs="1c")
        self._ed.pack(fill="both", expand=True)
        self._ed.insert("1.0", DEFAULT_STRATEGY)
        self._highlight()

        # Backtest results
        self._bt_tab = tk.Frame(self._nb, bg=BG2)
        self._nb.add(self._bt_tab, text="  Backtest Results  ")
        tk.Label(self._bt_tab,
                 text="Configure strategy and data, then click  ▶ Run Backtest",
                 bg=BG2, fg=FG2, font=(UI[0],12)).pack(expand=True)

        # TV Analysis
        self._tv_tab = tk.Frame(self._nb, bg=BG2)
        self._nb.add(self._tv_tab, text="  📊 TV Trade Analysis  ")
        tk.Label(self._tv_tab,
                 text="Browse a TradingView .xlsx export, then click  📊 Analyze",
                 bg=BG2, fg=FG2, font=(UI[0],12)).pack(expand=True)

        # Log
        log_t = tk.Frame(self._nb, bg=BG2)
        self._nb.add(log_t, text="  Log  ")
        self._log = scrolledtext.ScrolledText(
            log_t, bg="#0d1117", fg="#8b949e",
            font=("Cascadia Code",9) if _P=="win32" else ("Menlo",9),
            relief="flat", bd=0, state="disabled", wrap="word")
        self._log.pack(fill="both", expand=True)
        for tag, col in [("ok",GREEN),("err",RED),("warn",WARN),
                         ("info",ACCENT),("plain",FG2)]:
            self._log.tag_config(tag, foreground=col)

    # ── Helpers ────────────────────────────────────────────────────────────
    def _toggle_src(self):
        if self._src.get() == "alpaca":
            self._alpaca_f.pack(fill="x"); self._csv_f.pack_forget()
        else:
            self._csv_f.pack(fill="x"); self._alpaca_f.pack_forget()

    def _browse(self, var, ft=None):
        p = filedialog.askopenfilename(
            filetypes=ft or [("CSV","*.csv"),("All","*.*")])
        if p: var.set(p)

    def _reset_ed(self):
        self._ed.delete("1.0","end")
        self._ed.insert("1.0", DEFAULT_STRATEGY)

    def _highlight(self):
        self._ed.tag_config("kw",      foreground="#ff7b72")
        self._ed.tag_config("str",     foreground="#a5d6ff")
        self._ed.tag_config("comment", foreground="#8b949e")
        content = self._ed.get("1.0","end")
        for m in re.finditer(
                r"\b(def|class|import|from|return|if|else|elif|for|while|"
                r"in|not|and|or|True|False|None|try|except|raise|with|as|"
                r"lambda|pass|break|continue)\b", content):
            self._ed.tag_add("kw", f"1.0+{m.start()}c", f"1.0+{m.end()}c")
        for m in re.finditer(
                r'(""".*?"""|\'\'\'.*?\'\'\'|"[^"]*"|\'[^\']*\')',
                content, re.DOTALL):
            self._ed.tag_add("str", f"1.0+{m.start()}c", f"1.0+{m.end()}c")
        for m in re.finditer(r"#[^\n]*", content):
            self._ed.tag_add("comment", f"1.0+{m.start()}c", f"1.0+{m.end()}c")

    def _log_w(self, msg, tag="plain"):
        self._log.configure(state="normal")
        self._log.insert("end", f"[{datetime.now():%H:%M:%S}] {msg}\n", tag)
        self._log.see("end")
        self._log.configure(state="disabled")

    # ── Backtest ───────────────────────────────────────────────────────────
    def _run_bt(self):
        self._prog.start(12)
        self._status.set("Running backtest…")
        self._log_w("═"*60, "info")
        self._log_w(f"Backtest: {self._strat_name.get()}", "info")
        threading.Thread(target=self._bt_task, daemon=True).start()

    def _bt_task(self):
        try:
            self._log_w("Loading market data…", "plain")
            data = self._load_data()
            self._log_w(
                f"Data loaded: {len(data):,} bars  "
                f"{data.index[0].date()} → {data.index[-1].date()}", "ok")

            bench = self._benchmark.get().strip() or "SPY"
            self._log_w(f"Benchmark: {bench} (fetching via yfinance)…", "plain")

            result = run_backtest(data, self._ed.get("1.0","end"),
                                  self._strat_name.get(), bench)
            self._bt_html = result["html_path"]
            self.after(0, lambda: self._show_bt(result))
            self._log_w("Backtest complete! Full tearsheet ready.", "ok")
            for k, v in result["summary"].items():
                self._log_w(f"  {k:<28}{v}", "plain")
        except Exception as e:
            self._log_w(f"ERROR: {e}", "err")
            self._log_w(traceback.format_exc(), "err")
            self.after(0, lambda: messagebox.showerror("Backtest Error", str(e)))
            self.after(0, lambda: self._status.set(f"Error: {e}"))
        finally:
            self.after(0, self._prog.stop)

    def _load_data(self):
        if self._src.get() == "alpaca":
            return fetch_alpaca_data(
                self._ticker.get().strip().upper(),
                self._start.get().strip(), self._end.get().strip(),
                self._tf.get(),
                self._api_key.get().strip(), self._sec_key.get().strip())
        path = self._csv_path.get().strip()
        if not path or not os.path.exists(path):
            raise ValueError("Please select a valid CSV file.")
        df = detect_and_load_csv(path)
        try:
            s = pd.Timestamp(self._start.get().strip(), tz="UTC")
            e = pd.Timestamp(self._end.get().strip(), tz="UTC") + pd.Timedelta(days=1)
            df = df[(df.index >= s) & (df.index <= e)]
        except Exception: pass
        if df.empty: raise ValueError("CSV data empty after date filtering.")
        return df

    def _show_bt(self, result):
        for w in self._bt_tab.winfo_children(): w.destroy()

        # Header
        hdr = tk.Frame(self._bt_tab, bg=BG2)
        hdr.pack(fill="x", padx=20, pady=(16,0))
        tk.Label(hdr, text=f"✓  {self._strat_name.get()}  vs  {result['bench_name']}",
                 bg=BG2, fg=GREEN,
                 font=(HEAD[0],14,"bold")).pack(side="left")
        ttk.Button(hdr, text="🌐 Open Full Tearsheet",
                   style="Open.TButton",
                   command=self._open_bt_report).pack(side="right")
        ttk.Separator(self._bt_tab).pack(fill="x", padx=20, pady=10)

        # Stats grid
        grid = tk.Frame(self._bt_tab, bg=BG2)
        grid.pack(fill="x", padx=20)
        items = list(result["summary"].items())
        COLS  = 4
        for i,(k,v) in enumerate(items):
            r_, c_ = divmod(i, COLS)
            cell   = tk.Frame(grid, bg=BG3)
            cell.grid(row=r_, column=c_, padx=4, pady=4, sticky="ew")
            grid.columnconfigure(c_, weight=1)
            col = (GREEN if ("%" in v and not v.startswith("-"))
                   else RED if v.startswith("-") else FG)
            if "Drawdown" in k: col = RED if v.startswith("-") else WARN
            if "Benchmark" in k: col = ACCENT
            tk.Label(cell, text=k, bg=BG3, fg=FG2,
                     font=("",8)).pack(anchor="w", padx=8, pady=(7,1))
            tk.Label(cell, text=v, bg=BG3, fg=col,
                     font=(HEAD[0],13,"bold")).pack(anchor="w", padx=8, pady=(0,7))

        self._canvas_equity(self._bt_tab, result["daily_rets"], ACCENT,
                            height=120, padx=20)
        self._bt_open_btn.configure(state="normal")
        self._nb.select(1)
        self._status.set("Backtest complete — tearsheet ready. Click 🌐 to open.")

    def _open_bt_report(self):
        if self._bt_html and os.path.exists(self._bt_html):
            webbrowser.open(f"file://{self._bt_html}")
        else:
            messagebox.showwarning("No Report","Run a backtest first.")

    # ── TradingView ────────────────────────────────────────────────────────
    def _run_tv(self):
        path = self._tv_path.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showerror("File Error",
                                 "Please browse and select a TradingView .xlsx file.")
            return
        self._prog.start(12)
        self._status.set("Analyzing TradingView trades…")
        self._log_w("═"*60, "info")
        self._log_w(f"TV Analyzer: {os.path.basename(path)}", "info")
        threading.Thread(target=self._tv_task, args=(path,), daemon=True).start()

    def _tv_task(self, path):
        try:
            # 1. Load data
            data = load_tv_xlsx(path)
            d    = data["derived"]
            self._log_w(
                f"Trades loaded: {d['total_trades']}  |  "
                f"Win rate: {d['win_rate']:.1f}%  |  "
                f"Net P&L: ${d['total_pnl']:,.2f}", "ok")

            # 2. Generate full QuantStats tearsheet
            bench   = self._tv_bench.get().strip() or "SPY"
            name    = self._tv_name.get().strip() or "TV Strategy"
            self._log_w(
                f"Building tearsheet vs {bench} (yfinance)…", "plain")
            html, summary, daily_ret, bench_ret = build_tv_tearsheet(
                data, name, bench)
            self._tv_html = html
            self._log_w("Tearsheet generated!", "ok")
            for k,v in summary.items():
                self._log_w(f"  {k:<28}{v}", "plain")

            self.after(0, lambda: self._build_tv_ui(data, summary, daily_ret, bench_ret))
        except Exception as e:
            self._log_w(f"TV ERROR: {e}", "err")
            self._log_w(traceback.format_exc(), "err")
            self.after(0, lambda: messagebox.showerror("TV Analyzer Error", str(e)))
        finally:
            self.after(0, self._prog.stop)
            self.after(0, lambda: self._status.set("TV analysis complete."))

    def _open_tv_report(self):
        if self._tv_html and os.path.exists(self._tv_html):
            webbrowser.open(f"file://{self._tv_html}")
        else:
            messagebox.showwarning("No Report",
                                   "Run the TV analysis first.")

    # ── Build TV analysis tab ──────────────────────────────────────────────
    def _build_tv_ui(self, data, qs_summary, daily_ret, bench_ret):
        for w in self._tv_tab.winfo_children(): w.destroy()

        nb2 = ttk.Notebook(self._tv_tab)
        nb2.pack(fill="both", expand=True)

        dash  = tk.Frame(nb2, bg=BG2); nb2.add(dash,  text="  Dashboard  ")
        tbl   = tk.Frame(nb2, bg=BG2); nb2.add(tbl,   text="  Trade Table  ")
        eq    = tk.Frame(nb2, bg=BG2); nb2.add(eq,    text="  Equity & Drawdown  ")
        dist  = tk.Frame(nb2, bg=BG2); nb2.add(dist,  text="  Distribution  ")
        prop  = tk.Frame(nb2, bg=BG2); nb2.add(prop,  text="  All Stats  ")

        self._tv_dashboard(dash, data, qs_summary, daily_ret, bench_ret)
        self._tv_table(tbl, data["trades"])
        self._tv_equity_tab(eq, data["derived"], daily_ret, bench_ret)
        self._tv_dist(dist, data["trades"])
        self._tv_all_stats(prop, data)

        self._tv_open_btn.configure(state="normal")
        self._nb.select(2)

    # ── TV Dashboard ──────────────────────────────────────────────────────
    def _tv_dashboard(self, parent, data, qs_sum, daily_ret, bench_ret):
        d     = data["derived"]
        props = data["props"]
        risk  = data["risk"]

        cvs = tk.Canvas(parent, bg=BG2, bd=0, highlightthickness=0)
        vsb = ttk.Scrollbar(parent, orient="vertical", command=cvs.yview)
        inn = tk.Frame(cvs, bg=BG2)
        inn.bind("<Configure>",
                 lambda e: cvs.configure(scrollregion=cvs.bbox("all")))
        cvs.create_window((0,0), window=inn, anchor="nw")
        cvs.configure(yscrollcommand=vsb.set)
        cvs.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        sym  = props.get("Symbol",("—",))[0]
        tf   = props.get("Timeframe",("—",))[0]
        rng  = props.get("Trading range",("—",))[0]
        bench_name = bench_ret.name if hasattr(bench_ret, "name") else "Benchmark"

        tk.Label(inn, text="📊  TradingView Strategy Tearsheet",
                 bg=BG2, fg=PURPLE,
                 font=(HEAD[0],15,"bold")).pack(anchor="w", padx=20, pady=(16,4))
        tk.Label(inn,
                 text=f"  {sym}  ·  {tf}  ·  {rng}",
                 bg=BG2, fg=FG2, font=("",9)
                 ).pack(anchor="w", padx=20, pady=(0,4))

        # Open tearsheet banner
        banner = tk.Frame(inn, bg="#0d2137")
        banner.pack(fill="x", padx=20, pady=(4,10))
        tk.Label(banner,
                 text="  Full QuantStats tearsheet generated — "
                      "monthly heatmap · rolling Sharpe · drawdown · distribution",
                 bg="#0d2137", fg=ACCENT, font=("",9)
                 ).pack(side="left", padx=10, pady=8)
        ttk.Button(banner, text="🌐 Open Full Tearsheet",
                   style="Open.TButton",
                   command=self._open_tv_report).pack(side="right", padx=10, pady=6)

        ttk.Separator(inn).pack(fill="x", padx=20)

        def kpi_grid(frame, title, items, ncols=4):
            tk.Label(frame, text=title, bg=BG2, fg=FG2,
                     font=(HEAD[0],10,"bold")).pack(anchor="w", padx=20, pady=(14,6))
            g = tk.Frame(frame, bg=BG2); g.pack(fill="x", padx=20)
            for i,(lbl,val,col) in enumerate(items):
                r,c = divmod(i, ncols)
                cell = tk.Frame(g, bg=BG3)
                cell.grid(row=r, column=c, padx=4, pady=4, sticky="ew")
                g.columnconfigure(c, weight=1)
                tk.Label(cell, text=lbl, bg=BG3, fg=FG2,
                         font=("",8)).pack(anchor="w", padx=10, pady=(8,2))
                tk.Label(cell, text=str(val), bg=BG3, fg=col,
                         font=(HEAD[0],14,"bold")).pack(
                             anchor="w", padx=10, pady=(0,8))

        pf = d["profit_factor"]

        # QuantStats metrics (from tearsheet computation)
        kpi_grid(inn, "QUANTSTATS METRICS (from tearsheet)", [
            ("Total Return",    qs_sum.get("Total Return","—"),
             GREEN if not str(qs_sum.get("Total Return","")).startswith("-") else RED),
            ("CAGR",            qs_sum.get("CAGR","—"),
             GREEN if not str(qs_sum.get("CAGR","")).startswith("-") else RED),
            ("Sharpe Ratio",    qs_sum.get("Sharpe Ratio","—"), ACCENT),
            ("Sortino Ratio",   qs_sum.get("Sortino Ratio","—"), ACCENT),
            ("Max Drawdown",    qs_sum.get("Max Drawdown","—"), RED),
            ("Volatility",      qs_sum.get("Volatility (Ann)","—"), WARN),
            ("VaR (95%)",       qs_sum.get("VaR (95%)","—"), RED),
            (f"{bench_name} Return",
             qs_sum.get(f"Benchmark ({bench_name})","—"), FG2),
        ])

        kpi_grid(inn, "TRADE-LEVEL PERFORMANCE", [
            ("Net P&L",        f"${d['total_pnl']:,.2f}",
             GREEN if d["total_pnl"] >= 0 else RED),
            ("Gross Profit",   f"${d['gross_profit']:,.2f}",   GREEN),
            ("Gross Loss",     f"${abs(d['gross_loss']):,.2f}", RED),
            ("Profit Factor",  f"{pf:.3f}" if pf != float('inf') else "∞",
             GREEN if pf >= 1 else RED),
            ("Max Drawdown $", f"${d['max_drawdown']:,.2f}", RED),
            ("Win Rate",       f"{d['win_rate']:.1f}%",
             GREEN if d["win_rate"] >= 50 else WARN),
            ("Avg Win",        f"${d['avg_win']:,.2f}",        GREEN),
            ("Avg Loss",       f"${abs(d['avg_loss']):,.2f}",  RED),
        ])

        kpi_grid(inn, "TRADE COUNTS & RISK", [
            ("Total Trades",   str(d["total_trades"]),          FG),
            ("Winners",        str(d["winners"]),               GREEN),
            ("Losers",         str(d["losers"]),                RED),
            ("Even",           str(d["evens"]),                 FG2),
            ("Largest Win",    f"${d['largest_win']:,.2f}",     GREEN),
            ("Largest Loss",   f"${abs(d['largest_loss']):,.2f}", RED),
            ("Initial Capital",f"${d['init_cap']:,.0f}",        FG),
            ("Avg Duration",
             str(d["avg_duration"]).split(".")[0]
             if pd.notna(d["avg_duration"]) else "—",           FG),
        ])

        # Equity vs Benchmark preview
        tk.Label(inn, text="STRATEGY EQUITY  vs  BENCHMARK",
                 bg=BG2, fg=FG2,
                 font=(HEAD[0],10,"bold")).pack(anchor="w", padx=20, pady=(14,4))
        self._canvas_equity_vs_bench(inn, daily_ret, bench_ret, height=160, padx=20)

        tk.Label(inn, text="DRAWDOWN",
                 bg=BG2, fg=FG2,
                 font=(HEAD[0],10,"bold")).pack(anchor="w", padx=20, pady=(12,4))
        self._canvas_line(inn, d["drawdown"], RED, height=80, padx=20)
        tk.Frame(inn, bg=BG2, height=20).pack()

    # ── TV Trade Table ─────────────────────────────────────────────────────
    def _tv_table(self, parent, trades):
        tk.Label(parent, text="  All Closed Trades",
                 bg=BG2, fg=PURPLE,
                 font=(HEAD[0],12,"bold")).pack(anchor="w", padx=16, pady=(12,4))

        fb = tk.Frame(parent, bg=BG2); fb.pack(fill="x", padx=16, pady=(0,6))
        tk.Label(fb, text="Filter:", bg=BG2, fg=FG2, font=("",9)).pack(side="left")
        self._tv_filt = tk.StringVar()
        self._tv_filt.trace_add("write",
            lambda *a: self._tv_filter(trades))
        ttk.Entry(fb, textvariable=self._tv_filt,
                  width=20, font=("",9)).pack(side="left", padx=6)
        tk.Label(fb, text="(filters any column — press Enter or type)",
                 bg=BG2, fg=FG2, font=("",8)).pack(side="left")

        cols   = ["#","Entry Time","Exit Time","Dir","Signal",
                  "Entry $","Exit $","Qty","P&L $","P&L %",
                  "MFE $","MAE $","Duration","Cum P&L $"]
        widths = [38,132,132,55,130,68,68,55,75,62,62,62,100,80]
        frm    = tk.Frame(parent, bg=BG2)
        frm.pack(fill="both", expand=True, padx=16, pady=(0,12))
        xsb = ttk.Scrollbar(frm, orient="horizontal")
        ysb = ttk.Scrollbar(frm, orient="vertical")
        self._tv_tree = ttk.Treeview(
            frm, columns=cols, show="headings",
            yscrollcommand=ysb.set, xscrollcommand=xsb.set)
        ysb.configure(command=self._tv_tree.yview)
        xsb.configure(command=self._tv_tree.xview)
        xsb.pack(side="bottom", fill="x")
        ysb.pack(side="right", fill="y")
        self._tv_tree.pack(fill="both", expand=True)
        for col, w in zip(cols, widths):
            self._tv_tree.heading(col, text=col,
                command=lambda c=col: self._tv_sort(c))
            self._tv_tree.column(col, width=w, stretch=False,
                anchor="w" if col in ("Entry Time","Exit Time","Dir","Signal") else "e")
        self._tv_tree.tag_configure("win",  foreground=GREEN)
        self._tv_tree.tag_configure("loss", foreground=RED)
        self._tv_tree.tag_configure("even", foreground=FG2)
        self._tv_raw = trades
        self._tv_fill(trades)

    def _tv_fill(self, trades):
        self._tv_tree.delete(*self._tv_tree.get_children())
        for _, row in trades.iterrows():
            pnl = row.get("pnl_usd", 0) or 0
            tag = "win" if pnl > 0 else ("loss" if pnl < 0 else "even")
            vals = [
                int(row["trade_num"]) if pd.notna(row.get("trade_num")) else "",
                str(row.get("datetime_entry",""))[:16],
                str(row.get("datetime_exit",""))[:16],
                str(row.get("direction","")).capitalize(),
                str(row.get("signal_exit", row.get("signal",""))),
                f"{row.get('price_entry',0):.4f}" if pd.notna(row.get("price_entry")) else "",
                f"{row.get('price_exit',0):.4f}"  if pd.notna(row.get("price_exit"))  else "",
                f"{int(row.get('qty',0)):,}"       if pd.notna(row.get("qty"))         else "",
                f"${pnl:,.2f}",
                f"{row.get('pnl_pct',0):.2f}%"    if pd.notna(row.get("pnl_pct"))     else "",
                f"${row.get('mfe_usd',0):,.2f}"    if pd.notna(row.get("mfe_usd"))     else "",
                f"${row.get('mae_usd',0):,.2f}"    if pd.notna(row.get("mae_usd"))     else "",
                str(row.get("duration","")).split(".")[0],
                f"${row.get('cum_pnl_usd',0):,.2f}" if pd.notna(row.get("cum_pnl_usd")) else "",
            ]
            self._tv_tree.insert("","end", values=vals, tags=(tag,))

    def _tv_filter(self, trades):
        q = self._tv_filt.get().lower().strip()
        self._tv_fill(
            trades if not q else
            trades[trades.apply(lambda r: q in str(r.to_dict()).lower(), axis=1)])

    def _tv_sort(self, col):
        data = [(self._tv_tree.set(c, col), c)
                for c in self._tv_tree.get_children("")]
        try:
            data.sort(key=lambda x: float(
                x[0].replace("$","").replace("%","").replace(",","")))
        except ValueError:
            data.sort()
        for i,(_,c) in enumerate(data): self._tv_tree.move(c,"",i)

    # ── TV Equity & Drawdown ───────────────────────────────────────────────
    def _tv_equity_tab(self, parent, d, daily_ret, bench_ret):
        cvs = tk.Canvas(parent, bg=BG2, bd=0, highlightthickness=0)
        vsb = ttk.Scrollbar(parent, orient="vertical", command=cvs.yview)
        inn = tk.Frame(cvs, bg=BG2)
        inn.bind("<Configure>",
                 lambda e: cvs.configure(scrollregion=cvs.bbox("all")))
        cvs.create_window((0,0), window=inn, anchor="nw")
        cvs.configure(yscrollcommand=vsb.set)
        cvs.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        tk.Label(inn, text="Equity Curve — Strategy vs Benchmark",
                 bg=BG2, fg=PURPLE,
                 font=(HEAD[0],12,"bold")).pack(anchor="w", padx=20, pady=(16,4))
        self._canvas_equity_vs_bench(inn, daily_ret, bench_ret, height=180, padx=20)

        tk.Label(inn, text="Cumulative P&L  (USD)",
                 bg=BG2, fg=PURPLE,
                 font=(HEAD[0],12,"bold")).pack(anchor="w", padx=20, pady=(14,4))
        self._canvas_equity(inn, d["equity"], GREEN, height=140, padx=20)

        tk.Label(inn, text="Daily P&L  (USD)",
                 bg=BG2, fg=ORANGE,
                 font=(HEAD[0],12,"bold")).pack(anchor="w", padx=20, pady=(14,4))
        self._canvas_bars(inn, d["daily_pnl"], height=110, padx=20)

        tk.Label(inn, text="Drawdown from Equity Peak  (USD)",
                 bg=BG2, fg=RED,
                 font=(HEAD[0],12,"bold")).pack(anchor="w", padx=20, pady=(14,4))
        self._canvas_line(inn, d["drawdown"], RED, height=100, padx=20)

        # Metrics
        eq   = d["equity"]
        dd   = d["drawdown"]
        dpnl = d["daily_pnl"]
        tk.Label(inn, text="Key Metrics",
                 bg=BG2, fg=FG2,
                 font=(HEAD[0],10,"bold")).pack(anchor="w", padx=20, pady=(14,6))
        g = tk.Frame(inn, bg=BG2); g.pack(fill="x", padx=20, pady=(0,20))
        for i,(k,v,c) in enumerate([
            ("Final Equity",    f"${eq.iloc[-1]:,.2f}",              GREEN),
            ("Peak Equity",     f"${eq.max():,.2f}",                 GREEN),
            ("Max Drawdown $",  f"${dd.min():,.2f}",                 RED),
            ("Max DD %",        f"{dd.min()/max(eq.max(),1)*100:.2f}%", RED),
            ("Best Day",        f"${dpnl.max():,.2f}",               GREEN),
            ("Worst Day",       f"${dpnl.min():,.2f}",               RED),
            ("Avg Daily P&L",   f"${dpnl.mean():,.2f}",              FG),
            ("Profitable Days", f"{(dpnl>0).sum()} / {len(dpnl)}",  GREEN),
        ]):
            r,c_ = divmod(i,4)
            cell = tk.Frame(g, bg=BG3)
            cell.grid(row=r, column=c_, padx=4, pady=4, sticky="ew")
            g.columnconfigure(c_, weight=1)
            tk.Label(cell, text=k, bg=BG3, fg=FG2,
                     font=("",8)).pack(anchor="w", padx=8, pady=(7,1))
            tk.Label(cell, text=v, bg=BG3, fg=c,
                     font=(HEAD[0],12,"bold")).pack(anchor="w", padx=8, pady=(0,7))

    # ── TV Distribution ────────────────────────────────────────────────────
    def _tv_dist(self, parent, trades):
        cvs = tk.Canvas(parent, bg=BG2, bd=0, highlightthickness=0)
        vsb = ttk.Scrollbar(parent, orient="vertical", command=cvs.yview)
        inn = tk.Frame(cvs, bg=BG2)
        inn.bind("<Configure>",
                 lambda e: cvs.configure(scrollregion=cvs.bbox("all")))
        cvs.create_window((0,0), window=inn, anchor="nw")
        cvs.configure(yscrollcommand=vsb.set)
        cvs.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        pnl  = trades["pnl_usd"].dropna()
        w    = pnl[pnl > 0]; l = pnl[pnl < 0]; e = pnl[pnl == 0]
        tot  = max(len(pnl), 1)

        tk.Label(inn, text="P&L Distribution", bg=BG2, fg=PURPLE,
                 font=(HEAD[0],13,"bold")).pack(anchor="w", padx=20, pady=(16,8))
        self._stacked_bar(inn, len(w)/tot, len(e)/tot, len(l)/tot)

        tk.Label(inn, text="P&L Histogram  (USD per trade)",
                 bg=BG2, fg=FG2, font=(HEAD[0],10,"bold")).pack(
                     anchor="w", padx=20, pady=(14,4))
        self._histogram(inn, pnl.tolist(), height=180, padx=20)

        # Streak analysis
        s = self._streaks(pnl)
        tk.Label(inn, text="Streak Analysis", bg=BG2, fg=FG2,
                 font=(HEAD[0],10,"bold")).pack(anchor="w", padx=20, pady=(14,4))
        sg = tk.Frame(inn, bg=BG2); sg.pack(fill="x", padx=20)
        for i,(k,v,c) in enumerate([
            ("Max Win Streak",  str(s["max_win"]),       GREEN),
            ("Max Loss Streak", str(s["max_loss"]),      RED),
            ("Current Streak",  str(s["current"]),       FG),
            ("Avg Win Streak",  f"{s['avg_win']:.1f}",  GREEN),
        ]):
            cell = tk.Frame(sg, bg=BG3)
            cell.grid(row=0, column=i, padx=4, pady=4, sticky="ew")
            sg.columnconfigure(i, weight=1)
            tk.Label(cell, text=k, bg=BG3, fg=FG2,
                     font=("",8)).pack(anchor="w", padx=8, pady=(7,1))
            tk.Label(cell, text=v, bg=BG3, fg=c,
                     font=(HEAD[0],14,"bold")).pack(anchor="w", padx=8, pady=(0,7))

        # Monthly breakdown
        tk.Label(inn, text="Monthly P&L Breakdown",
                 bg=BG2, fg=FG2, font=(HEAD[0],10,"bold")).pack(
                     anchor="w", padx=20, pady=(14,4))
        self._monthly_table(inn, trades)
        tk.Frame(inn, bg=BG2, height=20).pack()

    # ── TV All Stats ───────────────────────────────────────────────────────
    def _tv_all_stats(self, parent, data):
        cvs = tk.Canvas(parent, bg=BG2, bd=0, highlightthickness=0)
        vsb = ttk.Scrollbar(parent, orient="vertical", command=cvs.yview)
        inn = tk.Frame(cvs, bg=BG2)
        inn.bind("<Configure>",
                 lambda e: cvs.configure(scrollregion=cvs.bbox("all")))
        cvs.create_window((0,0), window=inn, anchor="nw")
        cvs.configure(yscrollcommand=vsb.set)
        cvs.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        tk.Label(inn, text="All TradingView Summary Data",
                 bg=BG2, fg=PURPLE,
                 font=(HEAD[0],13,"bold")).pack(anchor="w", padx=20, pady=(16,10))

        # Column headers explanation
        tk.Label(inn,
                 text="  Columns: All USD  |  All %  |  Long USD  |"
                      "  Long %  |  Short USD  |  Short %",
                 bg=BG2, fg=FG2, font=("",8)).pack(anchor="w", padx=20, pady=(0,10))

        for name, sheet, col in [
            ("📋  Properties",                data["props"],  ACCENT),
            ("💰  Performance",               data["perf"],   GREEN),
            ("📈  Trades Analysis",           data["tana"],   ORANGE),
            ("⚖️   Risk-Adjusted Performance", data["risk"],  PURPLE),
        ]:
            if not sheet: continue
            tk.Label(inn, text=name, bg=BG2, fg=col,
                     font=(HEAD[0],11,"bold")).pack(anchor="w", padx=20, pady=(16,4))
            f = tk.Frame(inn, bg=BG3); f.pack(fill="x", padx=20, pady=(0,8))

            # Table header
            hrow = tk.Frame(f, bg=BG2)
            hrow.pack(fill="x")
            for htxt, hw in [("Metric",220),("All USD",80),("All %",65),
                              ("Long USD",80),("Long %",65),
                              ("Short USD",80),("Short %",65),]:
                tk.Label(hrow, text=htxt, bg=BG2, fg=FG2, font=("",8,"bold"),
                         width=hw//7, anchor="w").pack(side="left", padx=4, pady=3)

            for key, vals in sheet.items():
                if key in ("name","value"): continue
                r = tk.Frame(f, bg=BG3); r.pack(fill="x")
                r.bind("<Enter>",  lambda e, fr=r: fr.configure(bg="#1d2230"))
                r.bind("<Leave>",  lambda e, fr=r: fr.configure(bg=BG3))
                tk.Label(r, text=str(key), bg=BG3, fg=FG, font=("",9),
                         width=30, anchor="w").pack(side="left", padx=(10,4), pady=2)
                for v in vals:
                    disp = str(v) if v is not None else "—"
                    col2 = (GREEN if isinstance(v,(int,float)) and v > 0
                            else RED if isinstance(v,(int,float)) and v < 0
                            else FG2)
                    tk.Label(r, text=disp, bg=BG3, fg=col2,
                             font=MONO, width=9, anchor="e"
                             ).pack(side="left", padx=4)
                ttk.Separator(f).pack(fill="x")
        tk.Frame(inn, bg=BG2, height=20).pack()

    # ════════════════════════════════════════════════════════════════════════
    # SHARED CANVAS HELPERS
    # ════════════════════════════════════════════════════════════════════════

    def _canvas_equity(self, parent, series, color, height=120, padx=16):
        if series is None or len(series) == 0: return
        f = tk.Frame(parent, bg=BG2); f.pack(fill="x", padx=padx, pady=4)
        c = tk.Canvas(f, height=height, bg=BG3, highlightthickness=0, bd=0)
        c.pack(fill="x"); c.update()
        W = c.winfo_width() or 900; PAD = 12
        vals = np.array(series.values, dtype=float)
        mn, mx = vals.min(), vals.max(); rng = (mx-mn) or 1e-9
        pts = []
        for i,v in enumerate(vals):
            x = PAD+(i/max(len(vals)-1,1))*(W-2*PAD)
            y = height-PAD-((v-mn)/rng)*(height-2*PAD)
            pts.append((x,y))
        flat = [n for pt in pts for n in pt]
        for i in range(len(pts)-1):
            x1,y1=pts[i]; x2,y2=pts[i+1]
            c.create_polygon(x1,y1,x2,y2,x2,height-PAD,x1,height-PAD,
                             fill="#1a3a5c",outline="")
        if len(flat) >= 4:
            c.create_line(*flat, fill=color, width=2, smooth=True)
        c.create_text(W-PAD, PAD,       text=f"{mx:,.2f}", fill=FG2,
                      anchor="ne", font=("",7))
        c.create_text(W-PAD, height-4,  text=f"{mn:,.2f}", fill=FG2,
                      anchor="se", font=("",7))

    def _canvas_equity_vs_bench(self, parent, strat_ret, bench_ret,
                                 height=150, padx=16):
        """Overlay equity curves: strategy (green) vs benchmark (orange)."""
        if strat_ret is None or len(strat_ret) == 0: return
        f = tk.Frame(parent, bg=BG2); f.pack(fill="x", padx=padx, pady=4)
        c = tk.Canvas(f, height=height, bg=BG3, highlightthickness=0, bd=0)
        c.pack(fill="x"); c.update()
        W = c.winfo_width() or 900; PAD = 12

        strat_eq = (1 + strat_ret).cumprod().values.astype(float)
        bench_eq = (1 + bench_ret.reindex(strat_ret.index).fillna(0)
                    ).cumprod().values.astype(float)

        all_vals = np.concatenate([strat_eq, bench_eq])
        mn, mx   = all_vals.min(), all_vals.max(); rng = (mx-mn) or 1e-9
        N        = len(strat_eq)

        def pts(arr):
            return [(PAD+(i/max(N-1,1))*(W-2*PAD),
                     height-PAD-((v-mn)/rng)*(height-2*PAD))
                    for i,v in enumerate(arr)]

        def draw(arr, col, lw=2):
            p  = pts(arr)
            fl = [n for pt in p for n in pt]
            if len(fl) >= 4:
                c.create_line(*fl, fill=col, width=lw, smooth=True)

        # Shade strategy area
        sp  = pts(strat_eq)
        sfl = []
        for x,y in sp: sfl += [x,y]
        sfl += [sp[-1][0], height-PAD, sp[0][0], height-PAD]
        c.create_polygon(*sfl, fill="#1a3a5c", outline="")

        draw(bench_eq, ORANGE, lw=1)
        draw(strat_eq, GREEN,  lw=2)

        # 1.0 baseline
        base_y = height-PAD-((1.0-mn)/rng)*(height-2*PAD)
        c.create_line(PAD, base_y, W-PAD, base_y, fill=BORDER, dash=(3,3))

        # Legend
        bn = bench_ret.name if hasattr(bench_ret,"name") else "Benchmark"
        c.create_rectangle(W-190, 8, W-180, 18, fill=GREEN, outline="")
        c.create_text(W-178, 13, text="Strategy", fill=FG2,
                      anchor="w", font=("",7))
        c.create_rectangle(W-110, 8, W-100, 18, fill=ORANGE, outline="")
        c.create_text(W-98,  13, text=bn, fill=FG2,
                      anchor="w", font=("",7))

        c.create_text(W-PAD, PAD,       text=f"{mx:.3f}x", fill=FG2,
                      anchor="ne", font=("",7))
        c.create_text(W-PAD, height-4,  text=f"{mn:.3f}x", fill=FG2,
                      anchor="se", font=("",7))

    def _canvas_line(self, parent, series, color, height=100, padx=16):
        if series is None or len(series) == 0: return
        f = tk.Frame(parent, bg=BG2); f.pack(fill="x", padx=padx, pady=4)
        c = tk.Canvas(f, height=height, bg=BG3, highlightthickness=0, bd=0)
        c.pack(fill="x"); c.update()
        W = c.winfo_width() or 900; PAD = 10
        vals = np.array(series.values, dtype=float)
        mn, mx = vals.min(), vals.max(); rng = (mx-mn) or 1e-9
        xs = [PAD+(i/max(len(vals)-1,1))*(W-2*PAD) for i in range(len(vals))]
        ys = [height-PAD-((v-mn)/rng)*(height-2*PAD) for v in vals]
        fl = [n for xy in zip(xs,ys) for n in xy]
        if len(fl) >= 4:
            c.create_line(*fl, fill=color, width=1.5, smooth=True)
        c.create_text(W-PAD, PAD+2,    text=f"{mx:,.2f}", fill=FG2,
                      anchor="ne", font=("",7))
        c.create_text(W-PAD, height-4, text=f"{mn:,.2f}", fill=FG2,
                      anchor="se", font=("",7))

    def _canvas_bars(self, parent, series, height=100, padx=16):
        if series is None or len(series) == 0: return
        f = tk.Frame(parent, bg=BG2); f.pack(fill="x", padx=padx, pady=4)
        c = tk.Canvas(f, height=height, bg=BG3, highlightthickness=0, bd=0)
        c.pack(fill="x"); c.update()
        W   = c.winfo_width() or 900; PAD = 10
        vals = np.array(series.values, dtype=float)
        rng  = max(abs(vals.min()), abs(vals.max())) or 1e-9
        bw   = max((W-2*PAD)/len(vals)-1, 2)
        mid  = height//2
        for i,v in enumerate(vals):
            x1 = PAD+i*((W-2*PAD)/len(vals)); x2 = x1+bw
            h_ = abs(v)/rng*(height//2-PAD)
            col = GREEN if v >= 0 else RED
            if v >= 0: c.create_rectangle(x1, mid-h_, x2, mid, fill=col, outline="")
            else:      c.create_rectangle(x1, mid,    x2, mid+h_, fill=col, outline="")
        c.create_line(PAD, mid, W-PAD, mid, fill=BORDER, width=1)

    def _stacked_bar(self, parent, wr, er, lr):
        f = tk.Frame(parent, bg=BG2); f.pack(fill="x", padx=20, pady=(0,4))
        bar = tk.Canvas(f, height=28, bg=BG3, highlightthickness=0)
        bar.pack(fill="x"); bar.update()
        W = bar.winfo_width() or 900
        xw = int(W*wr); xe = int(W*(wr+er))
        if xw:    bar.create_rectangle(0, 0, xw, 28, fill=GREEN,  outline="")
        if xe-xw: bar.create_rectangle(xw,0, xe, 28, fill=FG2,    outline="")
        if W-xe:  bar.create_rectangle(xe,0, W,  28, fill=RED,    outline="")
        bar.create_text(max(xw//2,20), 14, fill=BG,
                        text=f"W {wr*100:.0f}%", font=("",8,"bold"))
        if xe-xw > 30:
            bar.create_text(xw+(xe-xw)//2, 14, fill=BG,
                            text=f"E {er*100:.0f}%", font=("",8))
        bar.create_text(xe+(W-xe)//2 if W-xe>20 else W-25, 14, fill=BG,
                        text=f"L {lr*100:.0f}%", font=("",8,"bold"))
        lab = tk.Frame(f, bg=BG2); lab.pack(fill="x")
        for t,cl in [("■ Wins",GREEN),("■ Even",FG2),("■ Losses",RED)]:
            tk.Label(lab, text=t, bg=BG2, fg=cl, font=("",8)).pack(side="left", padx=8)

    def _histogram(self, parent, values, height=160, padx=20):
        if not values: return
        f = tk.Frame(parent, bg=BG2); f.pack(fill="x", padx=padx, pady=4)
        c = tk.Canvas(f, height=height, bg=BG3, highlightthickness=0)
        c.pack(fill="x"); c.update()
        W = c.winfo_width() or 900
        mn, mx = min(values), max(values); rng = (mx-mn) or 1e-9
        BINS = 30; bw = rng/BINS; cnts = [0]*BINS
        for v in values:
            cnts[min(int((v-mn)/bw), BINS-1)] += 1
        mc = max(cnts) or 1; BAR = W/BINS; PAD = 10
        for i,cnt in enumerate(cnts):
            x1=i*BAR; x2=x1+BAR-1
            col = GREEN if mn+i*bw >= 0 else RED
            y1 = height-PAD-(cnt/mc)*(height-2*PAD)
            c.create_rectangle(x1,y1,x2,height-PAD,fill=col,outline="")
        c.create_line(0,height-PAD,W,height-PAD,fill=BORDER)
        zx = (-mn/rng)*W
        if 0 <= zx <= W: c.create_line(zx,0,zx,height,fill=FG2,dash=(4,4))
        c.create_text(6, height-3, text=f"${mn:.2f}", fill=FG2, anchor="sw", font=("",7))
        c.create_text(W-4,height-3,text=f"${mx:.2f}", fill=FG2, anchor="se", font=("",7))

    def _streaks(self, pnl):
        mw=ml=cw=cl=0; ws=[]
        for v in pnl:
            if v > 0:
                cw+=1; mw=max(mw,cw)
                if cl: cl=0
            elif v < 0:
                cl+=1; ml=max(ml,cl)
                if cw: ws.append(cw); cw=0
        cur = f"+{cw}W" if cw else (f"-{cl}L" if cl else "0")
        return dict(max_win=mw, max_loss=ml, current=cur,
                    avg_win=float(np.mean(ws)) if ws else 0.0)

    def _monthly_table(self, parent, trades):
        if "datetime_exit" not in trades.columns: return
        t = trades.copy()
        t["month"] = pd.to_datetime(
            t["datetime_exit"], errors="coerce").dt.to_period("M")
        mo = t.groupby("month").agg(
            n=("pnl_usd","count"),
            pnl=("pnl_usd","sum"),
            wins=("pnl_usd", lambda x:(x>0).sum()),
        ).reset_index()
        mo["wr"] = mo["wins"]/mo["n"]*100
        f  = tk.Frame(parent, bg=BG2); f.pack(fill="x", padx=20, pady=(0,6))
        cols = ["Month","Trades","Win%","P&L $","Cumulative P&L $"]
        tree = ttk.Treeview(f, columns=cols, show="headings", height=min(len(mo)+1,12))
        for col, w in zip(cols, [110,70,70,100,130]):
            tree.heading(col, text=col)
            tree.column(col, width=w, anchor="w" if col=="Month" else "e")
        cum = 0
        for _,r in mo.iterrows():
            pnl = r["pnl"]; cum += pnl
            tag = "mw" if pnl >= 0 else "ml"
            tree.insert("","end", tags=(tag,),
                values=[str(r["month"]), int(r["n"]),
                        f"{r['wr']:.1f}%", f"${pnl:,.2f}", f"${cum:,.2f}"])
        tree.tag_configure("mw", foreground=GREEN)
        tree.tag_configure("ml", foreground=RED)
        tree.pack(fill="x")

    def _reset_ed(self):
        self._ed.delete("1.0","end")
        self._ed.insert("1.0", DEFAULT_STRATEGY)


# ════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    App().mainloop()

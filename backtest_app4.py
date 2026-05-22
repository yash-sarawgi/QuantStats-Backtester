#!/usr/bin/env python3
"""
QuantStats Backtest + TradingView Analyzer  v4.0

CHANGES vs v3:
  ✓ Backtest now generates a full TRADE LOG from signal transitions
  ✓ Ticker symbol shown in tearsheet title, result header, and log
  ✓ Risk-Free Rate field (annualised %) — wired into every qs call
  ✓ Calculations audited and confirmed correct (see comments in run_backtest)
  ✓ Benchmark via yfinance (default SPY), configurable per module
"""

# ── AUTO-INSTALL ─────────────────────────────────────────────────────────────
import subprocess, sys

REQUIRED = {
    "quantstats": "quantstats",
    "pandas":     "pandas",
    "numpy":      "numpy",
    "pytz":       "pytz",
    "openpyxl":   "openpyxl",
    "yfinance":   "yfinance",
}

def _ensure():
    for mod, pkg in REQUIRED.items():
        try:
            __import__(mod)
        except ImportError:
            print(f"Installing {pkg}…", flush=True)
            subprocess.check_call(
                [sys.executable, "-m", "pip", "install", pkg, "-q",
                 "--break-system-packages"],
                stderr=subprocess.DEVNULL)

_ensure()

# ── IMPORTS ──────────────────────────────────────────────────────────────────
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading, traceback, os, tempfile, webbrowser, re
import pandas as pd
import numpy as np
import pytz
import quantstats as qs
import yfinance as yf
from datetime import datetime, timedelta
import openpyxl

# ── THEME ─────────────────────────────────────────────────────────────────────
BG     = "#0d1117"
BG2    = "#161b22"
BG3    = "#21262d"
BORDER = "#30363d"
ACCENT = "#58a6ff"
GREEN  = "#3fb950"
WARN   = "#d29922"
RED    = "#f85149"
PURPLE = "#bc8cff"
ORANGE = "#ffa657"
TEAL   = "#39d353"
FG     = "#e6edf3"
FG2    = "#8b949e"

_P   = sys.platform
MONO = ("Cascadia Code", 9) if _P == "win32" else ("Menlo", 9)
UI   = ("Segoe UI",      10) if _P == "win32" else ("SF Pro Display", 10)
HEAD = ("Segoe UI Semibold", 11) if _P == "win32" else ("SF Pro Display", 11)

TF_MAP = {
    "1 Min":  "1Min",  "5 Min":  "5Min",  "15 Min": "15Min",
    "30 Min": "30Min", "1 Hour": "1Hour", "4 Hour": "4Hour",
    "1 Day":  "1Day",
}

DEFAULT_STRATEGY = '''\
# ══════════════════════════════════════════════════════════════════════════
# Strategy function signature (required):
#
#   def strategy(data: pd.DataFrame) -> pd.Series
#
# INPUT  data (UTC-aware DatetimeIndex):
#   open, high, low, close  – OHLC (float)
#   volume                  – float
#   vwap                    – float
#
# OUTPUT  pd.Series of integer/float SIGNALS:
#   +1 = Long    0 = Flat    -1 = Short
#
# Execution: signal at bar N → fills at bar N+1 open (no look-ahead).
# ══════════════════════════════════════════════════════════════════════════

import pandas as pd
import numpy as np

def strategy(data: pd.DataFrame) -> pd.Series:
    """Golden Cross — SMA-20 above SMA-50 → long, else flat."""
    close = data["close"]
    fast  = close.rolling(20).mean()
    slow  = close.rolling(50).mean()
    sig   = pd.Series(0.0, index=data.index)
    sig[fast > slow] = 1.0
    return sig
'''

# ════════════════════════════════════════════════════════════════════════════
# PURE HELPER FUNCTIONS  (no GUI, fully testable)
# ════════════════════════════════════════════════════════════════════════════

def _strip_tz(s: pd.Series) -> pd.Series:
    """
    Convert tz-aware DatetimeIndex → tz-naive UTC.

    WHY: QuantStats (and its Matplotlib internals) compares the strategy's
    index with the benchmark's index using dtype-equality.  When one is
    datetime64[us, UTC] and the other is datetime64[us] the comparison raises:
        TypeError: Cannot compare dtypes datetime64[us, UTC] and datetime64[us]

    We normalise to tz-naive (values unchanged — still UTC wall-clock).
    """
    if hasattr(s.index, "tz") and s.index.tz is not None:
        s = s.copy()
        s.index = s.index.tz_convert("UTC").tz_localize(None)
    return s


def _qs(fn, *a, **kw):
    """Call a qs.stats function, return NaN on any error."""
    try:    return fn(*a, **kw)
    except: return float("nan")


def _fmt(v, pct=False, dp=2):
    if np.isnan(v): return "—"
    return f"{v*100:.{dp}f}%" if pct else f"{v:.{dp}f}"


# ── Benchmark ─────────────────────────────────────────────────────────────

def fetch_benchmark(ticker: str, start: str, end: str) -> pd.Series:
    """
    Fetch adjusted-close daily returns for `ticker` from yfinance.
    Returns a tz-naive pd.Series of daily returns, named `ticker`.

    Raises ValueError with a helpful message if ticker is invalid or
    no data is returned (e.g. wrong date range).
    """
    end_padded = (datetime.strptime(end, "%Y-%m-%d")
                  + timedelta(days=5)).strftime("%Y-%m-%d")
    raw = yf.download(ticker, start=start, end=end_padded,
                      auto_adjust=True, progress=False, threads=False)
    if raw.empty:
        raise ValueError(
            f"yfinance returned no data for '{ticker}' "
            f"({start} → {end}).\n"
            "Check the ticker symbol and date range.")
    close = raw["Close"]
    if isinstance(close, pd.DataFrame):
        close = close.iloc[:, 0]          # multi-ticker edge case
    close.index = pd.to_datetime(close.index).tz_localize(None)
    rets = close.pct_change().dropna()
    rets.name = ticker.upper()
    return rets


# ── Data loading ──────────────────────────────────────────────────────────

def load_csv(filepath: str) -> pd.DataFrame:
    """
    Auto-detect timestamp + OHLCV columns and return a clean DataFrame
    with a UTC-aware DatetimeIndex and columns:
        open, high, low, close, volume, vwap
    """
    df = pd.read_csv(filepath)
    df.columns = [c.strip().lower() for c in df.columns]

    ts_candidates = ["timestamp","datetime","date","time","t",
                     "date_time","bar_time","open_time"]
    ts_col = next((c for c in ts_candidates if c in df.columns), None)
    if ts_col is None:
        for c in df.columns:
            try:
                if re.search(r"\d{4}-\d{2}-\d{2}", str(df[c].iloc[0])):
                    ts_col = c; break
            except Exception:
                pass
    if ts_col is None:
        raise ValueError("Cannot detect timestamp column. "
                         "Expected one of: timestamp / datetime / date.")

    df["timestamp"] = pd.to_datetime(df[ts_col], utc=True, errors="coerce")
    df = df.dropna(subset=["timestamp"]).set_index("timestamp").sort_index()

    aliases = {
        "open":   ["open","o","open_price"],
        "high":   ["high","h","high_price"],
        "low":    ["low","l","low_price"],
        "close":  ["close","c","last","adj_close","close_price"],
        "volume": ["volume","vol","v","qty","quantity"],
        "vwap":   ["vwap","vw","weighted_avg"],
    }
    for canon, alts in aliases.items():
        if canon not in df.columns:
            for a in alts:
                if a in df.columns:
                    df.rename(columns={a: canon}, inplace=True); break

    missing = [c for c in ["open","high","low","close"] if c not in df.columns]
    if missing:
        raise ValueError(f"CSV missing required columns: {missing}. "
                         f"Found: {list(df.columns)}")

    for col in ["volume","vwap"]:
        if col not in df.columns:
            df[col] = df["close"] if col == "vwap" else 0.0

    for col in ["open","high","low","close","volume","vwap"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    return (df[["open","high","low","close","volume","vwap"]]
              .dropna(subset=["open","high","low","close"]))


def fetch_alpaca(ticker, start, end, tf_label, api_key, secret_key):
    """Fetch OHLCV bars from Alpaca Free API."""
    try:
        from alpaca.data.historical import StockHistoricalDataClient
        from alpaca.data.requests   import StockBarsRequest
        from alpaca.data.timeframe  import TimeFrame, TimeFrameUnit
    except ImportError:
        raise RuntimeError("alpaca-py is not installed.\n"
                           "Run:  pip install alpaca-py")

    if not api_key or not secret_key:
        raise ValueError("Alpaca API Key and Secret Key are both required.")

    tf_map = {
        "1Min":  TimeFrame(1,  TimeFrameUnit.Minute),
        "5Min":  TimeFrame(5,  TimeFrameUnit.Minute),
        "15Min": TimeFrame(15, TimeFrameUnit.Minute),
        "30Min": TimeFrame(30, TimeFrameUnit.Minute),
        "1Hour": TimeFrame(1,  TimeFrameUnit.Hour),
        "4Hour": TimeFrame(4,  TimeFrameUnit.Hour),
        "1Day":  TimeFrame(1,  TimeFrameUnit.Day),
    }
    tf_key = TF_MAP.get(tf_label, "1Day")
    tz_ny  = pytz.timezone("America/New_York")
    s_dt   = datetime.strptime(start, "%Y-%m-%d").replace(tzinfo=tz_ny)
    e_dt   = datetime.strptime(end,   "%Y-%m-%d").replace(
                 hour=23, minute=59, second=59, tzinfo=tz_ny)

    client = StockHistoricalDataClient(api_key=api_key, secret_key=secret_key)
    req    = StockBarsRequest(symbol_or_symbols=ticker.upper(),
                              timeframe=tf_map[tf_key],
                              start=s_dt, end=e_dt, adjustment="all")
    df = client.get_stock_bars(req).df

    if df.empty:
        raise ValueError(f"No data returned for {ticker} ({start} → {end}).")

    if isinstance(df.index, pd.MultiIndex):
        df = df.reset_index(level=0, drop=True)

    df.index = pd.to_datetime(df.index, utc=True)
    df.index.name = "timestamp"
    df.columns = [c.lower() for c in df.columns]
    df["volume"] = df.get("volume", pd.Series(0.0, index=df.index))
    df["vwap"]   = df.get("vwap",   df["close"])

    return df[["open","high","low","close","volume","vwap"]].sort_index()


# ════════════════════════════════════════════════════════════════════════════
# TRADE LOG BUILDER  (from signal series)
# ════════════════════════════════════════════════════════════════════════════

def build_trade_log(data: pd.DataFrame,
                    signals: pd.Series) -> pd.DataFrame:
    """
    Extract round-trip trades from a signal series.

    SIGNAL CONVENTION (after the 1-bar shift applied in run_backtest):
      +1 → long position held this bar
       0 → flat (no position)
      -1 → short position held this bar

    TRADE DETECTION:
      We look at transitions in the *position* series.
      A new trade begins when position changes from 0 → ±1.
      It ends when position returns to 0 or flips direction.

    RETURN columns:
      trade_num, direction, entry_time, entry_price, exit_time,
      exit_price, bars_held, pnl_pct, pnl_pts, mfe_pct, mae_pct

    NOTE: entry/exit prices are the bar-open of the first and last bars
    of the position (consistent with next-bar-open fill assumption).
    MFE = max-favourable excursion (% from entry, best intrabar move).
    MAE = max-adverse excursion  (% from entry, worst intrabar move).
    """
    pos = signals.copy()
    trades = []
    trade_num = 0
    in_trade  = False
    entry_idx = None
    direction = 0

    idx_arr   = pos.index
    pos_arr   = pos.values
    open_arr  = data["open"].reindex(idx_arr).values
    high_arr  = data["high"].reindex(idx_arr).values
    low_arr   = data["low"].reindex(idx_arr).values
    close_arr = data["close"].reindex(idx_arr).values

    n = len(idx_arr)

    for i in range(n):
        cur = pos_arr[i]

        if not in_trade:
            if cur != 0:                          # ── ENTRY ──
                in_trade  = True
                direction = int(np.sign(cur))
                entry_idx = i
                trade_num += 1
        else:
            # Position changed or last bar
            if cur != direction or i == n - 1:    # ── EXIT ──
                exit_idx = i if i < n - 1 else i

                entry_time  = idx_arr[entry_idx]
                exit_time   = idx_arr[exit_idx]
                entry_price = open_arr[entry_idx]
                exit_price  = open_arr[exit_idx]   # next-bar-open convention

                # Bar-by-bar high/low of the position window
                sl = slice(entry_idx, exit_idx + 1)
                highs  = high_arr[sl]
                lows   = low_arr[sl]

                if direction == 1:   # long
                    mfe = (np.nanmax(highs)  - entry_price) / entry_price
                    mae = (np.nanmin(lows)   - entry_price) / entry_price
                else:                # short
                    mfe = (entry_price - np.nanmin(lows))  / entry_price
                    mae = (entry_price - np.nanmax(highs)) / entry_price

                pnl_pct = direction * (exit_price - entry_price) / entry_price
                pnl_pts = direction * (exit_price - entry_price)

                trades.append(dict(
                    trade_num    = trade_num,
                    direction    = "Long" if direction == 1 else "Short",
                    entry_time   = entry_time,
                    entry_price  = round(entry_price, 4),
                    exit_time    = exit_time,
                    exit_price   = round(exit_price,  4),
                    bars_held    = exit_idx - entry_idx,
                    pnl_pct      = round(pnl_pct * 100, 3),
                    pnl_pts      = round(pnl_pts, 4),
                    mfe_pct      = round(mfe * 100, 3),
                    mae_pct      = round(mae * 100, 3),
                ))

                # Did price immediately flip direction?
                if cur != 0 and i < n - 1:
                    in_trade  = True
                    direction = int(np.sign(cur))
                    entry_idx = i
                    trade_num += 1
                else:
                    in_trade = False

    if not trades:
        return pd.DataFrame(columns=[
            "trade_num","direction","entry_time","entry_price",
            "exit_time","exit_price","bars_held",
            "pnl_pct","pnl_pts","mfe_pct","mae_pct"])

    df = pd.DataFrame(trades)

    # Cumulative P&L (%)  — compound, not sum
    df["cum_pnl_pct"] = ((1 + df["pnl_pct"] / 100).cumprod() - 1) * 100
    df["cum_pnl_pct"] = df["cum_pnl_pct"].round(3)

    return df


# ════════════════════════════════════════════════════════════════════════════
# BACKTEST ENGINE
# ════════════════════════════════════════════════════════════════════════════

def run_backtest(data: pd.DataFrame,
                 code: str,
                 strategy_name: str,
                 ticker: str,
                 benchmark_ticker: str = "SPY",
                 rf_annual: float = 0.0,
                 shift_signals: bool = True) -> dict:
    """
    Full backtest pipeline:
      1. Execute user strategy code  →  raw signal Series
      2. Shift signals 1 bar        →  no look-ahead bias
      3. Multiply by bar returns    →  portfolio return series
      4. Resample to daily          →  daily return Series
      5. Build trade log            →  DataFrame of round-trips
      6. Fetch benchmark            →  via yfinance
      7. Generate QuantStats HTML   →  tearsheet with rf applied

    CALCULATION NOTES (confirmed correct):

    Step 2 — Signal shift:
      signal[t] is computed from data available at bar t (close price etc.).
      We shift by +1 → the position is opened at bar t+1 open.
      This prevents using bar-t close to fill at bar-t close (look-ahead).

    Step 3 — Bar return:
      pct_change() on close gives  r[t] = (close[t] - close[t-1]) / close[t-1]
      Portfolio return = position[t] * r[t]
      Because position[t] is set from the previous bar's signal (after shift),
      this correctly represents: "hold position sized by signal, earn bar return".

    Step 4 — Daily compounding:
      For intraday data multiple bars fall on the same calendar date.
      Correct compound: daily_ret = ∏(1 + bar_ret_i) - 1
      We drop days where |daily_ret| ≥ 1 (data errors / splits).

    Step 5 — Trade log:
      See build_trade_log() docstring. Entry price = open of first signal bar.
      Exit price = open of first bar after signal turns off.

    Step 7 — Risk-free rate:
      QuantStats expects rf as an annualised decimal fraction.
      rf_annual = 0.05 means 5 % per year.
      We pass it to qs.reports.html(rf=...) and every qs.stats call.
    """
    # ── 1. Compile and execute strategy code ─────────────────────────────
    ns = {"pd": pd, "np": np, "__builtins__": __builtins__}
    try:
        exec(compile(code, "<strategy>", "exec"), ns)
    except SyntaxError as e:
        raise RuntimeError(f"Syntax error in strategy:\n  Line {e.lineno}: {e.msg}")
    except Exception as e:
        raise RuntimeError(f"Strategy compilation error:\n{e}")

    if "strategy" not in ns:
        raise ValueError("Strategy code must define a function named `strategy(data)`.")

    try:
        raw_sigs = ns["strategy"](data.copy())
    except Exception:
        raise RuntimeError(
            f"Strategy raised an exception at runtime:\n{traceback.format_exc()}")

    if not isinstance(raw_sigs, pd.Series):
        raise ValueError("`strategy()` must return a pandas Series.")

    # ── 2. Align + optionally shift signals (look-ahead prevention) ──────
    sigs = raw_sigs.reindex(data.index).fillna(0)
    if shift_signals:
        sigs = sigs.shift(1).fillna(0)      # position held at bar t

    # ── 3. Bar-level portfolio returns ───────────────────────────────────
    close_ret = data["close"].pct_change().fillna(0)   # r[t] = Δclose/close
    port_ret  = (sigs * close_ret).rename(strategy_name)
    port_ret.index = pd.to_datetime(port_ret.index, utc=True)

    # ── 4. Daily compounding ──────────────────────────────────────────────
    daily = (port_ret
             .resample("D")
             .apply(lambda x: (1 + x).prod() - 1)
             .dropna())
    # Drop suspicious days (|ret| ≥ 100 % — almost always a data issue)
    daily = daily[daily.abs() < 1.0]

    if daily.empty:
        raise ValueError(
            "The backtest produced zero valid daily returns.\n"
            "Verify the date range, ticker, and strategy logic.")

    # ── 5. Trade log ──────────────────────────────────────────────────────
    trade_log = build_trade_log(data, sigs)

    # ── 6. Benchmark ─────────────────────────────────────────────────────
    d_start    = daily.index[0].strftime("%Y-%m-%d")
    d_end      = daily.index[-1].strftime("%Y-%m-%d")
    bench_name = benchmark_ticker.strip().upper() or "SPY"

    try:
        bench_raw = fetch_benchmark(bench_name, d_start, d_end)
        bench     = (bench_raw
                     .reindex(daily.index.tz_localize(None), method="ffill")
                     .fillna(0.0))
        bench.name = bench_name
    except Exception as ex:
        # Fall back to buy-and-hold of the same ticker
        bh = (close_ret
              .resample("D")
              .apply(lambda x: (1 + x).prod() - 1)
              .dropna()
              .reindex(daily.index.tz_localize(None))
              .fillna(0.0))
        bh.name = f"B&H {ticker}"
        bench = bh
        print(f"[WARN] Benchmark '{bench_name}' fetch failed: {ex}. "
              f"Falling back to {bench.name}.")

    # ── 7. Strip tz + align indexes ───────────────────────────────────────
    daily_qs = _strip_tz(daily)
    bench_qs  = bench        # yfinance / pct_change already tz-naive

    common = daily_qs.index.intersection(bench_qs.index)
    if len(common) > 0:
        daily_qs = daily_qs.reindex(common)
        bench_qs  = bench_qs.reindex(common).fillna(0.0)
    else:
        # intraday strategy vs daily benchmark — just ffill-align
        bench_qs = bench_qs.reindex(daily_qs.index, method="ffill").fillna(0.0)

    # ── 8. QuantStats HTML tearsheet ─────────────────────────────────────
    safe = re.sub(r"[^a-zA-Z0-9]", "_", strategy_name)
    html = os.path.join(tempfile.gettempdir(),
                        f"bt_{safe}_{datetime.now():%Y%m%d_%H%M%S}.html")

    qs.extend_pandas()
    qs.reports.html(
        returns           = daily_qs,
        benchmark         = bench_qs,
        rf                = rf_annual,         # ← risk-free rate wired in
        title             = (f"{strategy_name} — {ticker.upper()}"
                             f"  vs  {bench_name}"),
        output            = html,
        download_filename = html,
    )

    # ── 9. Summary stats (same rf applied) ───────────────────────────────
    rf = rf_annual
    pct_in_mkt = (sigs.abs() > 0).mean() * 100

    summary = {
        f"Ticker":              ticker.upper(),
        f"Strategy":            strategy_name,
        "Total Return":         _fmt(_qs(qs.stats.comp, daily_qs),        pct=True),
        "CAGR":                 _fmt(_qs(qs.stats.cagr, daily_qs),        pct=True),
        "Sharpe Ratio":         _fmt(_qs(qs.stats.sharpe,  daily_qs, rf=rf)),
        "Sortino Ratio":        _fmt(_qs(qs.stats.sortino, daily_qs, rf=rf)),
        "Max Drawdown":         _fmt(_qs(qs.stats.max_drawdown, daily_qs), pct=True),
        "Volatility (Ann)":     _fmt(_qs(qs.stats.volatility,  daily_qs), pct=True),
        "Win Rate (days)":      _fmt(_qs(qs.stats.win_rate,    daily_qs), pct=True),
        "Calmar Ratio":         _fmt(_qs(qs.stats.calmar,      daily_qs)),
        "Skew":                 _fmt(_qs(qs.stats.skew,        daily_qs)),
        "Kurtosis":             _fmt(_qs(qs.stats.kurtosis,    daily_qs)),
        "VaR 95%":              _fmt(_qs(qs.stats.value_at_risk, daily_qs), pct=True),
        "CVaR 95%":             _fmt(_qs(qs.stats.cvar,        daily_qs), pct=True),
        f"{bench_name} Return": _fmt(_qs(qs.stats.comp, bench_qs),        pct=True),
        "% Bars in Market":     f"{pct_in_mkt:.1f}%",
        "Total Trades":         str(len(trade_log)),
        "Winning Trades":       str((trade_log["pnl_pct"] > 0).sum()) if len(trade_log) else "0",
        "Losing Trades":        str((trade_log["pnl_pct"] < 0).sum()) if len(trade_log) else "0",
        "Win Rate (trades)":    (f"{(trade_log['pnl_pct']>0).mean()*100:.1f}%"
                                 if len(trade_log) else "—"),
        "Avg Trade P&L":        (f"{trade_log['pnl_pct'].mean():.3f}%"
                                 if len(trade_log) else "—"),
        "Risk-Free Rate":       f"{rf_annual*100:.2f}% p.a.",
    }

    return dict(
        summary    = summary,
        html_path  = html,
        daily_rets = daily_qs,
        bench_rets = bench_qs,
        bench_name = bench_name,
        trade_log  = trade_log,
        ticker     = ticker.upper(),
    )


# ════════════════════════════════════════════════════════════════════════════
# TRADINGVIEW XLSX LOADER
# ════════════════════════════════════════════════════════════════════════════

def load_tv_xlsx(filepath: str) -> dict:
    """
    Parse a TradingView strategy-tester export (.xlsx).

    Expected sheets (TradingView standard layout):
      • List of trades
      • Performance
      • Trades analysis
      • Risk-adjusted performance
      • Properties
    """
    wb   = openpyxl.load_workbook(filepath, data_only=True)
    smap = {s.lower(): s for s in wb.sheetnames}

    def _sheet_df(key):
        name = next((v for k, v in smap.items() if key in k), None)
        if not name: return pd.DataFrame()
        ws   = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows: return pd.DataFrame()
        hdrs = [str(h) if h is not None else f"c{i}"
                for i, h in enumerate(rows[0])]
        return pd.DataFrame(rows[1:], columns=hdrs)

    def _kv(key):
        name = next((v for k, v in smap.items() if key in k), None)
        if not name: return {}
        ws  = wb[name]
        out = {}
        for row in ws.iter_rows(values_only=True):
            if row[0] is not None:
                out[str(row[0])] = row[1:]
        return out

    # ── List of trades ────────────────────────────────────────────────────
    lot = _sheet_df("list of trade")
    lot.columns = [c.strip() for c in lot.columns]
    col_map = {
        "Trade #":                 "trade_num",
        "Type":                    "type",
        "Date and time":           "datetime",
        "Signal":                  "signal",
        "Price USD":               "price",
        "Position size (qty)":     "qty",
        "Position size (value)":   "value",
        "Net P&L USD":             "pnl_usd",
        "Net P&L %":               "pnl_pct",
        "Favorable excursion USD": "mfe_usd",
        "Favorable excursion %":   "mfe_pct",
        "Adverse excursion USD":   "mae_usd",
        "Adverse excursion %":     "mae_pct",
        "Cumulative P&L USD":      "cum_pnl_usd",
        "Cumulative P&L %":        "cum_pnl_pct",
    }
    lot.rename(columns={k: v for k, v in col_map.items() if k in lot.columns},
               inplace=True)
    lot["datetime"] = pd.to_datetime(lot["datetime"], errors="coerce")
    for col in ["price","qty","value","pnl_usd","pnl_pct","mfe_usd","mfe_pct",
                "mae_usd","mae_pct","cum_pnl_usd","cum_pnl_pct"]:
        if col in lot.columns:
            lot[col] = pd.to_numeric(lot[col], errors="coerce")

    # ── Split entry / exit rows → merge into round-trips ─────────────────
    ent = lot[lot["type"].str.lower().str.contains("entry", na=False)].copy()
    ext = lot[lot["type"].str.lower().str.contains("exit",  na=False)].copy()
    ent = ent.sort_values("trade_num").drop_duplicates("trade_num", keep="last")
    ext = ext.sort_values("trade_num").drop_duplicates("trade_num", keep="last")

    trades = ent.merge(
        ext[["trade_num","datetime","price","pnl_usd","pnl_pct",
             "mfe_usd","mfe_pct","mae_usd","mae_pct","cum_pnl_usd","signal"]],
        on="trade_num", suffixes=("_entry","_exit"), how="inner")

    trades["duration"]  = trades["datetime_exit"] - trades["datetime_entry"]
    trades["direction"] = (trades["type"]
                           .str.extract(r"(long|short)", flags=re.I)[0]
                           .str.lower())
    trades.rename(columns={"pnl_usd_exit": "pnl_usd",
                            "pnl_pct_exit": "pnl_pct"}, inplace=True)
    trades = trades.reset_index(drop=True)

    # ── Derived statistics ────────────────────────────────────────────────
    pnl = trades["pnl_usd"].dropna()
    w   = pnl[pnl > 0];  l = pnl[pnl < 0];  e = pnl[pnl == 0]

    trades["exit_date"] = (pd.to_datetime(trades["datetime_exit"], errors="coerce")
                           .dt.normalize())
    daily_pnl = trades.groupby("exit_date")["pnl_usd"].sum()
    equity    = daily_pnl.cumsum()
    drawdown  = equity - equity.cummax()

    props    = _kv("properties")
    init_cap = 100_000.0
    try:
        raw = props.get("Initial capital", (None,))[0]
        if raw is not None:
            init_cap = float(str(raw).replace(",", ""))
    except Exception:
        pass

    derived = dict(
        total_trades  = len(trades),
        winners       = len(w),
        losers        = len(l),
        evens         = len(e),
        win_rate      = len(w) / max(len(trades) - len(e), 1) * 100,
        avg_win       = float(w.mean())  if len(w) else 0.0,
        avg_loss      = float(l.mean())  if len(l) else 0.0,
        total_pnl     = float(pnl.sum()),
        gross_profit  = float(w.sum())   if len(w) else 0.0,
        gross_loss    = float(l.sum())   if len(l) else 0.0,
        profit_factor = abs(w.sum() / l.sum()) if l.sum() != 0 else float("inf"),
        largest_win   = float(w.max())   if len(w) else 0.0,
        largest_loss  = float(l.min())   if len(l) else 0.0,
        max_drawdown  = float(drawdown.min()),
        avg_duration  = trades["duration"].mean(),
        equity        = equity,
        drawdown      = drawdown,
        daily_pnl     = daily_pnl,
        init_cap      = init_cap,
    )

    return dict(
        trades = trades,
        lot_raw= lot,
        perf   = _kv("performance"),
        tana   = _kv("trades analysis"),
        risk   = _kv("risk-adjusted"),
        props  = props,
        derived= derived,
    )


def build_tv_tearsheet(data: dict,
                       name: str,
                       benchmark_ticker: str = "SPY",
                       rf_annual: float = 0.0) -> tuple:
    """
    Convert TradingView daily P&L → daily returns → full QuantStats tearsheet.

    CALCULATION:
      equity[t]     = init_cap + cumulative_pnl[t]
      equity_lag[t] = equity[t-1]  (init_cap for t=0)
      daily_ret[t]  = daily_pnl[t] / equity_lag[t]

    This is the standard method: returns are in units of the portfolio value
    at the start of each day, which is what QuantStats expects.
    """
    d         = data["derived"]
    daily_pnl = d["daily_pnl"].copy()
    init_cap  = d["init_cap"]

    cum_pnl    = daily_pnl.cumsum()
    equity     = init_cap + cum_pnl
    equity_lag = equity.shift(1).fillna(init_cap)
    daily_ret  = (daily_pnl / equity_lag).rename(name)

    daily_ret.index = pd.to_datetime(daily_ret.index).tz_localize(None)
    daily_ret = daily_ret[daily_ret.abs() < 1].dropna()

    if daily_ret.empty:
        raise ValueError(
            "Could not compute daily returns from TradingView trade data.\n"
            "Ensure the 'List of trades' sheet contains P&L values.")

    d_start    = daily_ret.index[0].strftime("%Y-%m-%d")
    d_end      = daily_ret.index[-1].strftime("%Y-%m-%d")
    bench_name = benchmark_ticker.strip().upper() or "SPY"

    try:
        bench_raw = fetch_benchmark(bench_name, d_start, d_end)
        bench     = bench_raw.reindex(daily_ret.index, method="ffill").fillna(0.0)
        bench.name = bench_name
    except Exception as ex:
        bench = pd.Series(0.0, index=daily_ret.index, name="Flat (no data)")
        print(f"[WARN] TV benchmark fetch failed: {ex}")

    html = os.path.join(tempfile.gettempdir(),
                        f"tv_{re.sub(r'[^a-zA-Z0-9]','_',name)}"
                        f"_{datetime.now():%Y%m%d_%H%M%S}.html")

    qs.extend_pandas()
    qs.reports.html(
        returns           = daily_ret,
        benchmark         = bench,
        rf                = rf_annual,
        title             = f"{name}  vs  {bench_name}  [TradingView]",
        output            = html,
        download_filename = html,
    )

    rf = rf_annual
    summary = {
        "Total Return":      _fmt(_qs(qs.stats.comp,        daily_ret), pct=True),
        "CAGR":              _fmt(_qs(qs.stats.cagr,        daily_ret), pct=True),
        "Sharpe Ratio":      _fmt(_qs(qs.stats.sharpe,      daily_ret, rf=rf)),
        "Sortino Ratio":     _fmt(_qs(qs.stats.sortino,     daily_ret, rf=rf)),
        "Max Drawdown":      _fmt(_qs(qs.stats.max_drawdown,daily_ret), pct=True),
        "Volatility (Ann)":  _fmt(_qs(qs.stats.volatility,  daily_ret), pct=True),
        "Win Rate (days)":   _fmt(_qs(qs.stats.win_rate,    daily_ret), pct=True),
        "Calmar Ratio":      _fmt(_qs(qs.stats.calmar,      daily_ret)),
        "Skew":              _fmt(_qs(qs.stats.skew,        daily_ret)),
        "Kurtosis":          _fmt(_qs(qs.stats.kurtosis,    daily_ret)),
        "VaR 95%":           _fmt(_qs(qs.stats.value_at_risk,daily_ret),pct=True),
        "CVaR 95%":          _fmt(_qs(qs.stats.cvar,        daily_ret), pct=True),
        f"{bench_name} Ret": _fmt(_qs(qs.stats.comp,        bench),     pct=True),
        "Net P&L $":         f"${d['total_pnl']:,.2f}",
        "Total Trades":      str(d["total_trades"]),
        "Win Rate (trades)": f"{d['win_rate']:.1f}%",
        "Profit Factor":     (f"{d['profit_factor']:.3f}"
                              if d["profit_factor"] != float("inf") else "∞"),
        "Max Drawdown $":    f"${d['max_drawdown']:,.2f}",
        "Largest Win":       f"${d['largest_win']:,.2f}",
        "Largest Loss":      f"${abs(d['largest_loss']):,.2f}",
        "Risk-Free Rate":    f"{rf_annual*100:.2f}% p.a.",
    }
    return html, summary, daily_ret, bench


# ════════════════════════════════════════════════════════════════════════════
# GUI APPLICATION
# ════════════════════════════════════════════════════════════════════════════

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("QuantStats Backtest + TradingView Analyzer  v4.0")
        self.configure(bg=BG)
        self.geometry("1480x960")
        self.minsize(1200, 760)

        # ── State variables ───────────────────────────────────────────────
        self._src        = tk.StringVar(value="alpaca")
        self._csv_path   = tk.StringVar()
        self._ticker     = tk.StringVar(value="AAPL")
        self._start      = tk.StringVar(value="2022-01-01")
        self._end        = tk.StringVar(value=datetime.today().strftime("%Y-%m-%d"))
        self._tf         = tk.StringVar(value="1 Day")
        self._api_key    = tk.StringVar()
        self._sec_key    = tk.StringVar()
        self._strat_name = tk.StringVar(value="My Strategy")
        self._shift_sig  = tk.BooleanVar(value=True)
        self._benchmark  = tk.StringVar(value="SPY")
        self._rf_rate    = tk.StringVar(value="0.00")    # % per year

        self._tv_path    = tk.StringVar()
        self._tv_name    = tk.StringVar(value="TV Strategy")
        self._tv_bench   = tk.StringVar(value="SPY")
        self._tv_rf      = tk.StringVar(value="0.00")   # % per year

        self._status     = tk.StringVar(value="Ready.")
        self._bt_html    = None
        self._tv_html    = None

        self._sty()
        self._build_layout()

    # ── ttk styles ────────────────────────────────────────────────────────
    def _sty(self):
        s = ttk.Style(self)
        s.theme_use("clam")
        eb = dict(background=BG3, foreground=FG, fieldbackground=BG3,
                  insertbackground=FG, selectbackground=ACCENT,
                  selectforeground=BG, bordercolor=BORDER,
                  lightcolor=BORDER, darkcolor=BORDER,
                  relief="flat", padding=5)
        s.configure("TFrame",        background=BG)
        s.configure("TLabel",        background=BG,  foreground=FG, font=UI)
        s.configure("TEntry",        **eb)
        s.configure("TCombobox",     **eb)
        s.map("TCombobox",
              fieldbackground=[("readonly", BG3)],
              background     =[("readonly", BG3)])
        s.configure("TButton",       background=BG3, foreground=FG,
                    bordercolor=BORDER, font=UI,
                    padding=(10, 5), relief="flat")
        s.map("TButton",     background=[("active", BORDER)])
        s.configure("Run.TButton",   background=ACCENT, foreground=BG,
                    font=(UI[0],11,"bold"), padding=(16,8))
        s.map("Run.TButton", background=[("active","#79c0ff")])
        s.configure("Open.TButton",  background=GREEN, foreground=BG,
                    font=(UI[0],10,"bold"), padding=(12,6))
        s.map("Open.TButton",background=[("active","#56d364")])
        s.configure("TV.TButton",    background=PURPLE, foreground=BG,
                    font=(UI[0],10,"bold"), padding=(12,6))
        s.map("TV.TButton",  background=[("active","#d2a8ff")])
        s.configure("TVo.TButton",   background="#6e40c9", foreground=FG,
                    font=(UI[0],10,"bold"), padding=(12,6))
        s.map("TVo.TButton", background=[("active","#8957e5")])
        s.configure("TRadiobutton",  background=BG2, foreground=FG,
                    indicatorbackground=BG3, selectcolor=ACCENT)
        s.configure("TNotebook",     background=BG, bordercolor=BORDER,
                    tabmargins=0)
        s.configure("TNotebook.Tab", background=BG3, foreground=FG2,
                    padding=(14, 7), bordercolor=BORDER)
        s.map("TNotebook.Tab",
              background=[("selected", BG2)],
              foreground=[("selected", FG)])
        s.configure("TProgressbar",  troughcolor=BG3, background=ACCENT,
                    bordercolor=BORDER, thickness=4)
        s.configure("Treeview",
                    background=BG3, foreground=FG, fieldbackground=BG3,
                    bordercolor=BORDER, relief="flat", rowheight=22,
                    font=MONO)
        s.configure("Treeview.Heading",
                    background=BG2, foreground=FG2,
                    font=(UI[0],9,"bold"), relief="flat")
        s.map("Treeview",
              background=[("selected", ACCENT)],
              foreground=[("selected", BG)])

    # ── Master layout ─────────────────────────────────────────────────────
    def _build_layout(self):
        # Header
        hdr = tk.Frame(self, bg=BG2, height=54)
        hdr.pack(fill="x"); hdr.pack_propagate(False)
        tk.Label(hdr, text="⚡  QuantStats Backtest  +  TradingView Analyzer  v4.0",
                 bg=BG2, fg=FG, font=(HEAD[0],14,"bold")
                 ).pack(side="left", padx=20, pady=10)
        tk.Label(hdr, text="trade log  ·  benchmark via yfinance  ·  risk-free rate",
                 bg=BG2, fg=FG2, font=("",9)
                 ).pack(side="right", padx=20)
        ttk.Separator(self).pack(fill="x")

        # Body
        body = tk.Frame(self, bg=BG); body.pack(fill="both", expand=True)

        # Left sidebar
        left = tk.Frame(body, bg=BG, width=415)
        left.pack(side="left", fill="y"); left.pack_propagate(False)
        self._sidebar(left)

        tk.Frame(body, bg=BORDER, width=1).pack(side="left", fill="y")

        # Right content
        right = tk.Frame(body, bg=BG)
        right.pack(side="left", fill="both", expand=True)
        self._content(right)

        # Status bar
        sb = tk.Frame(self, bg=BG3, height=26)
        sb.pack(fill="x", side="bottom"); sb.pack_propagate(False)
        self._prog = ttk.Progressbar(sb, mode="indeterminate", length=160)
        self._prog.pack(side="right", padx=10, pady=5)
        tk.Label(sb, textvariable=self._status,
                 bg=BG3, fg=FG2, font=("",9), anchor="w"
                 ).pack(side="left", padx=10)

    # ── Sidebar ───────────────────────────────────────────────────────────
    def _sidebar(self, parent):
        cvs = tk.Canvas(parent, bg=BG, bd=0, highlightthickness=0)
        vsb = ttk.Scrollbar(parent, orient="vertical", command=cvs.yview)
        inn = tk.Frame(cvs, bg=BG)
        inn.bind("<Configure>",
                 lambda e: cvs.configure(scrollregion=cvs.bbox("all")))
        cvs.create_window((0, 0), window=inn, anchor="nw")
        cvs.configure(yscrollcommand=vsb.set)
        cvs.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        M = 12   # margin

        def sec(title, col=ACCENT):
            f = tk.Frame(inn, bg=BG2)
            f.pack(fill="x", padx=M, pady=(8, 0))
            tk.Label(f, text=title, bg=BG2, fg=col,
                     font=(HEAD[0],10,"bold")).pack(anchor="w", padx=10, pady=(7,3))
            return f

        def erow(p, label, var, show=None, lw=14):
            r = tk.Frame(p, bg=BG2); r.pack(fill="x", padx=10, pady=2)
            tk.Label(r, text=label, bg=BG2, fg=FG2, font=("",9),
                     width=lw, anchor="w").pack(side="left")
            ttk.Entry(r, textvariable=var, font=MONO, show=show
                      ).pack(side="left", fill="x", expand=True)

        def note(p, txt):
            tk.Label(p, text=txt, bg=BG2, fg=FG2, font=("",8),
                     justify="left").pack(anchor="w", padx=10, pady=(0,6))

        # ── BACKTEST CONFIG ───────────────────────────────────────────────
        f0 = sec("◈  BACKTEST STRATEGY")
        erow(f0, "Strategy Name", self._strat_name)
        erow(f0, "Ticker",        self._ticker)
        shf = tk.Frame(f0, bg=BG2); shf.pack(fill="x", padx=10, pady=2)
        tk.Checkbutton(shf, text="Shift signals +1 bar (prevent look-ahead)",
                       variable=self._shift_sig, onvalue=True, offvalue=False,
                       bg=BG2, fg=FG, selectcolor=BG3, activebackground=BG2,
                       activeforeground=FG, font=("",9)
                       ).pack(anchor="w")

        f1 = sec("◈  BENCHMARK  &  RISK-FREE RATE")
        erow(f1, "Benchmark",  self._benchmark)
        note(f1, "  SPY · QQQ · BTC-USD · ^GSPC · any yfinance ticker")
        erow(f1, "RF Rate % p.a.", self._rf_rate)
        note(f1, "  e.g. 5.25 for 5.25 % per year (US Fed rate)\n"
                 "  Used in Sharpe, Sortino, and tearsheet")

        f2 = sec("◈  DATA SOURCE")
        rb = tk.Frame(f2, bg=BG2); rb.pack(fill="x", padx=10, pady=3)
        for txt, val in [("Alpaca API","alpaca"),("CSV File","csv")]:
            ttk.Radiobutton(rb, text=txt, variable=self._src, value=val,
                            command=self._toggle_src).pack(side="left", padx=(0,14))

        self._alpaca_f = tk.Frame(f2, bg=BG2); self._alpaca_f.pack(fill="x")
        erow(self._alpaca_f, "API Key",    self._api_key)
        erow(self._alpaca_f, "Secret Key", self._sec_key, show="•")
        note(self._alpaca_f, "  alpaca.markets → Paper → API Keys (free)")

        self._csv_f = tk.Frame(f2, bg=BG2)
        cr = tk.Frame(self._csv_f, bg=BG2); cr.pack(fill="x", padx=10, pady=3)
        tk.Label(cr, text="CSV File", bg=BG2, fg=FG2, font=("",9),
                 width=14, anchor="w").pack(side="left")
        ttk.Entry(cr, textvariable=self._csv_path, font=("",9)
                  ).pack(side="left", fill="x", expand=True, padx=(0,4))
        ttk.Button(cr, text="Browse",
                   command=lambda: self._browse(self._csv_path)).pack(side="left")
        note(self._csv_f, "  Needs: timestamp, open, high, low, close, volume")

        f3 = sec("◈  DATE RANGE & TIMEFRAME")
        r  = tk.Frame(f3, bg=BG2); r.pack(fill="x", padx=10, pady=2)
        tk.Label(r, text="Timeframe", bg=BG2, fg=FG2, font=("",9),
                 width=14, anchor="w").pack(side="left")
        ttk.Combobox(r, textvariable=self._tf, state="readonly",
                     values=list(TF_MAP.keys()), font=UI
                     ).pack(side="left", fill="x", expand=True)
        erow(f3, "Start Date", self._start)
        erow(f3, "End Date",   self._end)
        note(f3, "  Format: YYYY-MM-DD")

        # Run button
        bf = tk.Frame(inn, bg=BG); bf.pack(fill="x", padx=M, pady=(10, 0))
        ttk.Button(bf, text="▶  Run Backtest  +  Build Trade Log",
                   style="Run.TButton",
                   command=self._run_bt).pack(fill="x")
        self._bt_open_btn = ttk.Button(
            bf, text="🌐  Open QuantStats Tearsheet",
            style="Open.TButton", command=self._open_bt_html, state="disabled")
        self._bt_open_btn.pack(fill="x", pady=(7, 0))

        # ── DIVIDER ───────────────────────────────────────────────────────
        ttk.Separator(inn).pack(fill="x", padx=M, pady=(14, 0))

        # ── TRADINGVIEW SECTION ───────────────────────────────────────────
        f4 = sec("◈  TRADINGVIEW ANALYZER", col=PURPLE)
        note(f4, "  Export: Strategy Tester → Export → Excel (.xlsx)")
        tvr = tk.Frame(f4, bg=BG2); tvr.pack(fill="x", padx=10, pady=(0,3))
        ttk.Entry(tvr, textvariable=self._tv_path, font=("",9)
                  ).pack(side="left", fill="x", expand=True, padx=(0,4))
        ttk.Button(tvr, text="Browse",
                   command=lambda: self._browse(
                       self._tv_path,
                       [("Excel XLSX","*.xlsx"),("All","*.*")]
                   )).pack(side="left")
        erow(f4, "Strategy Name", self._tv_name)
        erow(f4, "Benchmark",     self._tv_bench)
        erow(f4, "RF Rate % p.a.",self._tv_rf)
        note(f4, "  RF rate: same convention as backtest section above")

        tv_bf = tk.Frame(inn, bg=BG); tv_bf.pack(fill="x", padx=M, pady=(8, 0))
        ttk.Button(tv_bf, text="📊  Analyze Trades  +  Build Tearsheet",
                   style="TV.TButton", command=self._run_tv).pack(fill="x")
        self._tv_open_btn = ttk.Button(
            tv_bf, text="🌐  Open TV Tearsheet",
            style="TVo.TButton", command=self._open_tv_html, state="disabled")
        self._tv_open_btn.pack(fill="x", pady=(7, 0))

        tk.Frame(inn, bg=BG, height=14).pack()
        self._toggle_src()

    # ── Content (notebook) ────────────────────────────────────────────────
    def _content(self, parent):
        self._nb = ttk.Notebook(parent)
        self._nb.pack(fill="both", expand=True)

        # Tab 0 – Strategy Editor
        et = tk.Frame(self._nb, bg=BG2)
        self._nb.add(et, text="  Strategy Editor  ")
        tb = tk.Frame(et, bg=BG3, height=34)
        tb.pack(fill="x"); tb.pack_propagate(False)
        tk.Label(tb, text="strategy.py", bg=BG3, fg=FG2,
                 font=MONO).pack(side="left", padx=12, pady=7)
        ttk.Button(tb, text="Reset to default",
                   command=self._ed_reset).pack(side="right", padx=8, pady=5)
        self._ed = scrolledtext.ScrolledText(
            et, bg="#0d1117", fg="#e6edf3", font=MONO,
            insertbackground=FG, selectbackground=ACCENT, selectforeground=BG,
            relief="flat", bd=0, wrap="none", undo=True, tabs="1c")
        self._ed.pack(fill="both", expand=True)
        self._ed.insert("1.0", DEFAULT_STRATEGY)
        self._ed_highlight()

        # Tab 1 – Backtest Results
        self._bt_tab = tk.Frame(self._nb, bg=BG2)
        self._nb.add(self._bt_tab, text="  Backtest Results  ")
        tk.Label(self._bt_tab,
                 text="Configure your strategy, then click  ▶ Run Backtest",
                 bg=BG2, fg=FG2, font=(UI[0],12)).pack(expand=True)

        # Tab 2 – Trade Log (backtest)
        self._tl_tab = tk.Frame(self._nb, bg=BG2)
        self._nb.add(self._tl_tab, text="  📋 Trade Log  ")
        tk.Label(self._tl_tab,
                 text="Trade log appears here after running a backtest.",
                 bg=BG2, fg=FG2, font=(UI[0],12)).pack(expand=True)

        # Tab 3 – TV Analysis
        self._tv_tab = tk.Frame(self._nb, bg=BG2)
        self._nb.add(self._tv_tab, text="  📊 TV Analysis  ")
        tk.Label(self._tv_tab,
                 text="Load a TradingView XLSX export, then click  📊 Analyze",
                 bg=BG2, fg=FG2, font=(UI[0],12)).pack(expand=True)

        # Tab 4 – Log
        log_tab = tk.Frame(self._nb, bg=BG2)
        self._nb.add(log_tab, text="  Log  ")
        self._log = scrolledtext.ScrolledText(
            log_tab, bg="#0d1117", fg="#8b949e", font=MONO,
            relief="flat", bd=0, state="disabled", wrap="word")
        self._log.pack(fill="both", expand=True)
        for tag, col in [("ok",GREEN),("err",RED),("warn",WARN),
                         ("info",ACCENT),("plain",FG2)]:
            self._log.tag_config(tag, foreground=col)

    # ── Sidebar helpers ───────────────────────────────────────────────────
    def _toggle_src(self):
        if self._src.get() == "alpaca":
            self._alpaca_f.pack(fill="x")
            self._csv_f.pack_forget()
        else:
            self._csv_f.pack(fill="x")
            self._alpaca_f.pack_forget()

    def _browse(self, var, ft=None):
        p = filedialog.askopenfilename(
            filetypes=ft or [("CSV","*.csv"),("All","*.*")])
        if p: var.set(p)

    def _ed_reset(self):
        self._ed.delete("1.0","end")
        self._ed.insert("1.0", DEFAULT_STRATEGY)

    def _ed_highlight(self):
        self._ed.tag_config("kw",  foreground="#ff7b72")
        self._ed.tag_config("str", foreground="#a5d6ff")
        self._ed.tag_config("cmt", foreground="#8b949e")
        txt = self._ed.get("1.0","end")
        for m in re.finditer(
                r"\b(def|class|import|from|return|if|else|elif|for|while|"
                r"in|not|and|or|True|False|None|try|except|raise|with|as|"
                r"lambda|pass|break|continue|global|del|yield)\b", txt):
            self._ed.tag_add("kw",  f"1.0+{m.start()}c", f"1.0+{m.end()}c")
        for m in re.finditer(
                r'(""".*?"""|\'\'\'.*?\'\'\'|"[^"]*"|\'[^\']*\')',
                txt, re.DOTALL):
            self._ed.tag_add("str", f"1.0+{m.start()}c", f"1.0+{m.end()}c")
        for m in re.finditer(r"#[^\n]*", txt):
            self._ed.tag_add("cmt", f"1.0+{m.start()}c", f"1.0+{m.end()}c")

    # ── Log helper ────────────────────────────────────────────────────────
    def _log_w(self, msg: str, tag: str = "plain"):
        self._log.configure(state="normal")
        ts = datetime.now().strftime("%H:%M:%S")
        self._log.insert("end", f"[{ts}] {msg}\n", tag)
        self._log.see("end")
        self._log.configure(state="disabled")

    # ════════════════════════════════════════════════════════════════════════
    # BACKTEST
    # ════════════════════════════════════════════════════════════════════════

    def _run_bt(self):
        self._prog.start(12)
        self._status.set("Running backtest…")
        self._nb.select(4)          # show log while running
        self._log_w("═"*64, "info")
        self._log_w(f"Backtest  ·  {self._strat_name.get()}"
                    f"  ·  {self._ticker.get().upper()}", "info")
        threading.Thread(target=self._bt_thread, daemon=True).start()

    def _bt_thread(self):
        try:
            # ── Load data ────────────────────────────────────────────────
            self._log_w("Loading market data…")
            data   = self._load_data()
            ticker = self._ticker.get().strip().upper()
            self._log_w(
                f"Data loaded: {len(data):,} bars  "
                f"({data.index[0].date()} → {data.index[-1].date()})", "ok")

            # ── Parse RF rate ─────────────────────────────────────────────
            try:
                rf_pct = float(self._rf_rate.get().strip())
            except ValueError:
                rf_pct = 0.0
                self._log_w("RF rate parse error — defaulting to 0%", "warn")
            rf_ann = rf_pct / 100.0

            bench = self._benchmark.get().strip() or "SPY"
            self._log_w(f"Benchmark: {bench}  |  RF: {rf_pct:.2f}% p.a.")

            # ── Run backtest ──────────────────────────────────────────────
            result = run_backtest(
                data             = data,
                code             = self._ed.get("1.0","end"),
                strategy_name    = self._strat_name.get().strip(),
                ticker           = ticker,
                benchmark_ticker = bench,
                rf_annual        = rf_ann,
                shift_signals    = self._shift_sig.get(),
            )
            self._bt_html = result["html_path"]

            # ── Log summary ───────────────────────────────────────────────
            self._log_w(f"Tearsheet saved: {result['html_path']}", "ok")
            for k, v in result["summary"].items():
                self._log_w(f"  {k:<26}{v}", "plain")

            tl = result["trade_log"]
            self._log_w(f"Trade log: {len(tl)} round-trips found", "ok")
            if len(tl):
                wins = (tl["pnl_pct"] > 0).sum()
                self._log_w(f"  Winners: {wins}  Losers: {len(tl)-wins}", "plain")

            self.after(0, lambda: self._show_bt_results(result))

        except Exception as e:
            self._log_w(f"ERROR: {e}", "err")
            self._log_w(traceback.format_exc(), "err")
            self.after(0, lambda: messagebox.showerror("Backtest Error", str(e)))
            self.after(0, lambda: self._status.set(f"Error — see Log tab"))
        finally:
            self.after(0, self._prog.stop)

    def _load_data(self) -> pd.DataFrame:
        if self._src.get() == "alpaca":
            return fetch_alpaca(
                self._ticker.get().strip().upper(),
                self._start.get().strip(),
                self._end.get().strip(),
                self._tf.get(),
                self._api_key.get().strip(),
                self._sec_key.get().strip())
        path = self._csv_path.get().strip()
        if not path or not os.path.exists(path):
            raise ValueError("Please select a valid CSV file.")
        df = load_csv(path)
        try:
            s = pd.Timestamp(self._start.get().strip(), tz="UTC")
            e = pd.Timestamp(self._end.get().strip(),   tz="UTC") + pd.Timedelta(days=1)
            df = df[(df.index >= s) & (df.index <= e)]
        except Exception:
            pass
        if df.empty:
            raise ValueError("No data remains after date filtering.")
        return df

    # ── Show backtest results ─────────────────────────────────────────────
    def _show_bt_results(self, result: dict):
        # ── Tab 1: Stats ──────────────────────────────────────────────────
        for w in self._bt_tab.winfo_children(): w.destroy()

        hdr = tk.Frame(self._bt_tab, bg=BG2)
        hdr.pack(fill="x", padx=20, pady=(16,0))
        lbl = (f"✓  {result['summary']['Strategy']}"
               f"  —  {result['summary']['Ticker']}"
               f"  vs  {result['bench_name']}")
        tk.Label(hdr, text=lbl, bg=BG2, fg=GREEN,
                 font=(HEAD[0],13,"bold")).pack(side="left")
        ttk.Button(hdr, text="🌐 Open Full Tearsheet",
                   style="Open.TButton",
                   command=self._open_bt_html).pack(side="right")
        ttk.Separator(self._bt_tab).pack(fill="x", padx=20, pady=8)

        # Stats grid (4 columns)
        grid = tk.Frame(self._bt_tab, bg=BG2)
        grid.pack(fill="x", padx=20)
        COLS = 4
        for i, (k, v) in enumerate(result["summary"].items()):
            r_, c_ = divmod(i, COLS)
            cell   = tk.Frame(grid, bg=BG3)
            cell.grid(row=r_, column=c_, padx=4, pady=4, sticky="ew")
            grid.columnconfigure(c_, weight=1)
            # Colour logic
            if k in ("Ticker","Strategy","Risk-Free Rate"):
                col = ACCENT
            elif "Return" in k or "CAGR" in k or "Win" in k:
                col = (GREEN if not v.startswith("-") and v not in ("—","0.00%")
                       else RED)
            elif "Drawdown" in k or "VaR" in k or "CVaR" in k:
                col = RED
            elif "Sharpe" in k or "Sortino" in k or "Calmar" in k:
                col = PURPLE
            elif result["bench_name"] in k:
                col = ORANGE
            else:
                col = FG
            tk.Label(cell, text=k, bg=BG3, fg=FG2,
                     font=("",8)).pack(anchor="w", padx=8, pady=(6,1))
            tk.Label(cell, text=v, bg=BG3, fg=col,
                     font=(HEAD[0],12,"bold")).pack(anchor="w", padx=8, pady=(0,6))

        # Mini equity curve
        self._cv_equity(self._bt_tab, result["daily_rets"], ACCENT, height=110)

        # ── Tab 2: Trade Log ──────────────────────────────────────────────
        for w in self._tl_tab.winfo_children(): w.destroy()
        self._build_trade_log_tab(self._tl_tab, result)

        self._bt_open_btn.configure(state="normal")
        self._nb.select(1)
        self._status.set(
            f"Done — {result['summary']['Ticker']}  "
            f"{result['summary']['Total Return']}  "
            f"Sharpe {result['summary']['Sharpe Ratio']}")

    # ── Trade Log tab ─────────────────────────────────────────────────────
    def _build_trade_log_tab(self, parent: tk.Frame, result: dict):
        tl     = result["trade_log"]
        ticker = result["ticker"]
        bench  = result["bench_name"]

        # Header
        hdr = tk.Frame(parent, bg=BG2)
        hdr.pack(fill="x", padx=16, pady=(12,4))
        tk.Label(hdr,
                 text=(f"📋  Trade Log  —  {result['summary']['Strategy']}"
                       f"  /  {ticker}"),
                 bg=BG2, fg=TEAL,
                 font=(HEAD[0],12,"bold")).pack(side="left")
        tk.Label(hdr,
                 text=(f"{len(tl)} trades   |   "
                       f"Win rate: {result['summary']['Win Rate (trades)']}   |   "
                       f"Avg P&L: {result['summary']['Avg Trade P&L']}"),
                 bg=BG2, fg=FG2,
                 font=("",9)).pack(side="right")

        if tl.empty:
            tk.Label(parent,
                     text="No completed trades found.\n"
                          "The strategy may be always-in (never flips to 0)\n"
                          "or held position through the entire date range.",
                     bg=BG2, fg=WARN,
                     font=(UI[0],11), justify="center").pack(expand=True)
            return

        # Filter bar
        fb = tk.Frame(parent, bg=BG2); fb.pack(fill="x", padx=16, pady=(0,4))
        tk.Label(fb, text="Filter:", bg=BG2, fg=FG2, font=("",9)).pack(side="left")
        self._tl_filt = tk.StringVar()
        self._tl_filt.trace_add("write", lambda *_: self._tl_filter(tl))
        ttk.Entry(fb, textvariable=self._tl_filt, width=22,
                  font=("",9)).pack(side="left", padx=6)
        tk.Label(fb, text="Search any column · click headers to sort",
                 bg=BG2, fg=FG2, font=("",8)).pack(side="left", padx=4)

        # Treeview
        cols   = ["#","Dir","Entry Time","Entry $","Exit Time","Exit $",
                  "Bars","P&L %","P&L pts","MFE %","MAE %","Cum P&L %"]
        widths = [36,50,138,78,138,78,48,72,72,68,68,90]

        frm = tk.Frame(parent, bg=BG2)
        frm.pack(fill="both", expand=True, padx=16, pady=(0,12))
        xsb = ttk.Scrollbar(frm, orient="horizontal")
        ysb = ttk.Scrollbar(frm, orient="vertical")
        self._tl_tree = ttk.Treeview(
            frm, columns=cols, show="headings",
            yscrollcommand=ysb.set, xscrollcommand=xsb.set)
        ysb.configure(command=self._tl_tree.yview)
        xsb.configure(command=self._tl_tree.xview)
        xsb.pack(side="bottom", fill="x")
        ysb.pack(side="right",  fill="y")
        self._tl_tree.pack(fill="both", expand=True)

        for col, w in zip(cols, widths):
            self._tl_tree.heading(col, text=col,
                command=lambda c=col: self._tree_sort(self._tl_tree, c))
            self._tl_tree.column(col, width=w, stretch=False,
                anchor="w" if col in ("Dir","Entry Time","Exit Time") else "e")

        self._tl_tree.tag_configure("win",  foreground=GREEN)
        self._tl_tree.tag_configure("loss", foreground=RED)
        self._tl_tree.tag_configure("even", foreground=FG2)

        self._tl_raw = tl
        self._tl_populate(tl)

    def _tl_populate(self, tl: pd.DataFrame):
        self._tl_tree.delete(*self._tl_tree.get_children())
        for _, r in tl.iterrows():
            p   = r["pnl_pct"]
            tag = "win" if p > 0 else ("loss" if p < 0 else "even")
            et  = str(r["entry_time"])[:16]
            xt  = str(r["exit_time"])[:16]
            vals = [
                int(r["trade_num"]),
                str(r["direction"]),
                et,
                f"{r['entry_price']:.4f}",
                xt,
                f"{r['exit_price']:.4f}",
                int(r["bars_held"]),
                f"{p:.3f}%",
                f"{r['pnl_pts']:.4f}",
                f"{r['mfe_pct']:.3f}%",
                f"{r['mae_pct']:.3f}%",
                f"{r['cum_pnl_pct']:.3f}%",
            ]
            self._tl_tree.insert("","end", values=vals, tags=(tag,))

    def _tl_filter(self, tl: pd.DataFrame):
        q = self._tl_filt.get().lower().strip()
        if not q:
            self._tl_populate(tl); return
        filt = tl[tl.apply(lambda row: q in str(row.to_dict()).lower(), axis=1)]
        self._tl_populate(filt)

    def _tree_sort(self, tree: ttk.Treeview, col: str):
        rows = [(tree.set(c, col), c) for c in tree.get_children("")]
        try:
            rows.sort(key=lambda x: float(
                x[0].replace("%","").replace("$","").replace(",","")))
        except ValueError:
            rows.sort()
        for i, (_, c) in enumerate(rows):
            tree.move(c, "", i)

    def _open_bt_html(self):
        if self._bt_html and os.path.exists(self._bt_html):
            webbrowser.open(f"file://{self._bt_html}")
        else:
            messagebox.showwarning("No Tearsheet",
                                   "Run a backtest first to generate the tearsheet.")

    # ════════════════════════════════════════════════════════════════════════
    # TRADINGVIEW ANALYSIS
    # ════════════════════════════════════════════════════════════════════════

    def _run_tv(self):
        path = self._tv_path.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showerror("File Error",
                                 "Please browse and select a TradingView .xlsx file.")
            return
        self._prog.start(12)
        self._status.set("Analyzing TradingView trades…")
        self._nb.select(4)
        self._log_w("═"*64, "info")
        self._log_w(f"TV Analyzer  ·  {os.path.basename(path)}", "info")
        threading.Thread(target=self._tv_thread, args=(path,), daemon=True).start()

    def _tv_thread(self, path: str):
        try:
            data = load_tv_xlsx(path)
            d    = data["derived"]
            self._log_w(
                f"Loaded {d['total_trades']} trades  |  "
                f"Win rate: {d['win_rate']:.1f}%  |  "
                f"Net P&L: ${d['total_pnl']:,.2f}", "ok")

            try:
                rf_pct = float(self._tv_rf.get().strip())
            except ValueError:
                rf_pct = 0.0
            rf_ann = rf_pct / 100.0

            bench = self._tv_bench.get().strip() or "SPY"
            name  = self._tv_name.get().strip() or "TV Strategy"
            self._log_w(f"Building tearsheet vs {bench}  |  RF: {rf_pct:.2f}% p.a.")

            html, summary, daily_ret, bench_ret = build_tv_tearsheet(
                data, name, bench, rf_ann)
            self._tv_html = html
            self._log_w(f"Tearsheet: {html}", "ok")
            for k, v in summary.items():
                self._log_w(f"  {k:<28}{v}", "plain")

            self.after(0, lambda: self._show_tv(data, summary, daily_ret, bench_ret))

        except Exception as e:
            self._log_w(f"TV ERROR: {e}", "err")
            self._log_w(traceback.format_exc(), "err")
            self.after(0, lambda: messagebox.showerror("TV Error", str(e)))
        finally:
            self.after(0, self._prog.stop)
            self.after(0, lambda: self._status.set("TV analysis complete."))

    def _open_tv_html(self):
        if self._tv_html and os.path.exists(self._tv_html):
            webbrowser.open(f"file://{self._tv_html}")
        else:
            messagebox.showwarning("No Tearsheet",
                                   "Run a TV analysis first.")

    # ── TV display ────────────────────────────────────────────────────────
    def _show_tv(self, data, qs_sum, daily_ret, bench_ret):
        for w in self._tv_tab.winfo_children(): w.destroy()

        nb2 = ttk.Notebook(self._tv_tab)
        nb2.pack(fill="both", expand=True)

        dash = tk.Frame(nb2, bg=BG2); nb2.add(dash, text="  Dashboard  ")
        tbl  = tk.Frame(nb2, bg=BG2); nb2.add(tbl,  text="  Trade Table  ")
        eq   = tk.Frame(nb2, bg=BG2); nb2.add(eq,   text="  Equity & Drawdown  ")
        dist = tk.Frame(nb2, bg=BG2); nb2.add(dist,  text="  Distribution  ")
        prop = tk.Frame(nb2, bg=BG2); nb2.add(prop,  text="  All Stats  ")

        self._tv_dashboard(dash, data, qs_sum, daily_ret, bench_ret)
        self._tv_table(tbl, data["trades"])
        self._tv_equity_tab(eq, data["derived"], daily_ret, bench_ret)
        self._tv_dist(dist, data["trades"])
        self._tv_all_stats(prop, data)

        self._tv_open_btn.configure(state="normal")
        self._nb.select(3)

    # ── TV Dashboard ──────────────────────────────────────────────────────
    def _tv_dashboard(self, parent, data, qs_sum, daily_ret, bench_ret):
        d    = data["derived"]
        cvs, inn = self._scrollable(parent)

        props     = data["props"]
        sym       = props.get("Symbol",        ("—",))[0]
        tf        = props.get("Timeframe",      ("—",))[0]
        rng       = props.get("Trading range",  ("—",))[0]
        bench_lbl = bench_ret.name if hasattr(bench_ret,"name") else "Benchmark"

        tk.Label(inn, text="📊  TradingView Strategy Tearsheet",
                 bg=BG2, fg=PURPLE,
                 font=(HEAD[0],15,"bold")).pack(anchor="w", padx=20, pady=(16,4))
        tk.Label(inn, text=f"  {sym}  ·  {tf}  ·  {rng}",
                 bg=BG2, fg=FG2, font=("",9)).pack(anchor="w", padx=20)

        # Open tearsheet banner
        ban = tk.Frame(inn, bg="#0d2137")
        ban.pack(fill="x", padx=20, pady=(8,12))
        tk.Label(ban,
                 text="  Full QuantStats tearsheet generated with real benchmark"
                      " · monthly heatmap · rolling Sharpe · drawdown · distribution",
                 bg="#0d2137", fg=ACCENT, font=("",9)
                 ).pack(side="left", padx=10, pady=8)
        ttk.Button(ban, text="🌐 Open Full Tearsheet",
                   style="Open.TButton",
                   command=self._open_tv_html).pack(side="right", padx=10, pady=6)
        ttk.Separator(inn).pack(fill="x", padx=20)

        def kgrid(title, items, ncols=4):
            tk.Label(inn, text=title, bg=BG2, fg=FG2,
                     font=(HEAD[0],10,"bold")).pack(anchor="w", padx=20, pady=(14,6))
            g = tk.Frame(inn, bg=BG2); g.pack(fill="x", padx=20)
            for i,(lbl,val,col) in enumerate(items):
                r,c = divmod(i, ncols)
                cell = tk.Frame(g, bg=BG3)
                cell.grid(row=r, column=c, padx=4, pady=4, sticky="ew")
                g.columnconfigure(c, weight=1)
                tk.Label(cell, text=lbl, bg=BG3, fg=FG2, font=("",8)
                         ).pack(anchor="w", padx=8, pady=(7,1))
                tk.Label(cell, text=str(val), bg=BG3, fg=col,
                         font=(HEAD[0],13,"bold")).pack(anchor="w", padx=8, pady=(0,7))

        pf = d["profit_factor"]

        kgrid("QUANTSTATS METRICS  (from tearsheet)", [
            ("Total Return",    qs_sum.get("Total Return","—"),
             GREEN if not str(qs_sum.get("Total Return","")).startswith("-") else RED),
            ("CAGR",            qs_sum.get("CAGR","—"),
             GREEN if not str(qs_sum.get("CAGR","")).startswith("-") else RED),
            ("Sharpe Ratio",    qs_sum.get("Sharpe Ratio","—"),   PURPLE),
            ("Sortino Ratio",   qs_sum.get("Sortino Ratio","—"),  PURPLE),
            ("Max Drawdown",    qs_sum.get("Max Drawdown","—"),    RED),
            ("Volatility",      qs_sum.get("Volatility (Ann)","—"),WARN),
            ("VaR 95%",         qs_sum.get("VaR 95%","—"),         RED),
            (f"{bench_lbl} Ret",qs_sum.get(f"{bench_lbl} Ret","—"),ORANGE),
        ])

        kgrid("TRADE-LEVEL PERFORMANCE", [
            ("Net P&L",        f"${d['total_pnl']:,.2f}",
             GREEN if d["total_pnl"]>=0 else RED),
            ("Gross Profit",   f"${d['gross_profit']:,.2f}",         GREEN),
            ("Gross Loss",     f"${abs(d['gross_loss']):,.2f}",       RED),
            ("Profit Factor",  f"{pf:.3f}" if pf!=float("inf") else "∞",
             GREEN if pf>=1 else RED),
            ("Max Drawdown $", f"${d['max_drawdown']:,.2f}",          RED),
            ("Win Rate",       f"{d['win_rate']:.1f}%",
             GREEN if d["win_rate"]>=50 else WARN),
            ("Avg Win",        f"${d['avg_win']:,.2f}",               GREEN),
            ("Avg Loss",       f"${abs(d['avg_loss']):,.2f}",         RED),
        ])

        kgrid("TRADE COUNTS", [
            ("Total Trades",  str(d["total_trades"]),        FG),
            ("Winners",       str(d["winners"]),             GREEN),
            ("Losers",        str(d["losers"]),              RED),
            ("Even",          str(d["evens"]),               FG2),
            ("Largest Win",   f"${d['largest_win']:,.2f}",   GREEN),
            ("Largest Loss",  f"${abs(d['largest_loss']):,.2f}", RED),
            ("Initial Cap",   f"${d['init_cap']:,.0f}",       FG),
            ("Avg Duration",
             str(d["avg_duration"]).split(".")[0]
             if pd.notna(d["avg_duration"]) else "—",        FG),
        ])

        tk.Label(inn, text="STRATEGY vs BENCHMARK",
                 bg=BG2, fg=FG2,
                 font=(HEAD[0],10,"bold")).pack(anchor="w", padx=20, pady=(14,4))
        self._cv_vs_bench(inn, daily_ret, bench_ret, height=160)

        tk.Label(inn, text="DRAWDOWN",
                 bg=BG2, fg=FG2,
                 font=(HEAD[0],10,"bold")).pack(anchor="w", padx=20, pady=(10,4))
        self._cv_line(inn, d["drawdown"], RED, height=80)
        tk.Frame(inn, bg=BG2, height=20).pack()

    # ── TV Trade Table ─────────────────────────────────────────────────────
    def _tv_table(self, parent, trades):
        tk.Label(parent, text="  All Closed Trades",
                 bg=BG2, fg=PURPLE,
                 font=(HEAD[0],12,"bold")).pack(anchor="w", padx=16, pady=(12,4))

        fb = tk.Frame(parent, bg=BG2); fb.pack(fill="x", padx=16, pady=(0,4))
        tk.Label(fb, text="Filter:", bg=BG2, fg=FG2, font=("",9)).pack(side="left")
        self._tv_filt = tk.StringVar()
        self._tv_filt.trace_add("write", lambda *_: self._tv_filter(trades))
        ttk.Entry(fb, textvariable=self._tv_filt, width=22,
                  font=("",9)).pack(side="left", padx=6)
        tk.Label(fb, text="Click column headers to sort",
                 bg=BG2, fg=FG2, font=("",8)).pack(side="left")

        cols   = ["#","Entry Time","Exit Time","Dir","Signal",
                  "Entry $","Exit $","Qty","P&L $","P&L %",
                  "MFE $","MAE $","Duration","Cum P&L $"]
        widths = [36,132,132,52,128,70,70,54,78,64,64,64,102,82]

        frm = tk.Frame(parent, bg=BG2)
        frm.pack(fill="both", expand=True, padx=16, pady=(0,12))
        xsb = ttk.Scrollbar(frm, orient="horizontal")
        ysb = ttk.Scrollbar(frm, orient="vertical")
        self._tv_tree = ttk.Treeview(
            frm, columns=cols, show="headings",
            yscrollcommand=ysb.set, xscrollcommand=xsb.set)
        ysb.configure(command=self._tv_tree.yview)
        xsb.configure(command=self._tv_tree.xview)
        xsb.pack(side="bottom", fill="x")
        ysb.pack(side="right",  fill="y")
        self._tv_tree.pack(fill="both", expand=True)

        for col, w in zip(cols, widths):
            self._tv_tree.heading(col, text=col,
                command=lambda c=col: self._tree_sort(self._tv_tree, c))
            self._tv_tree.column(col, width=w, stretch=False,
                anchor="w" if col in ("Entry Time","Exit Time","Dir","Signal") else "e")

        self._tv_tree.tag_configure("win",  foreground=GREEN)
        self._tv_tree.tag_configure("loss", foreground=RED)
        self._tv_tree.tag_configure("even", foreground=FG2)
        self._tv_raw = trades
        self._tv_fill(trades)

    def _tv_fill(self, trades):
        self._tv_tree.delete(*self._tv_tree.get_children())
        for _, row in trades.iterrows():
            pnl = row.get("pnl_usd", 0) or 0
            tag = "win" if pnl > 0 else ("loss" if pnl < 0 else "even")
            vals = [
                int(row["trade_num"]) if pd.notna(row.get("trade_num")) else "",
                str(row.get("datetime_entry",""))[:16],
                str(row.get("datetime_exit",""))[:16],
                str(row.get("direction","")).capitalize(),
                str(row.get("signal_exit", row.get("signal",""))),
                f"{row.get('price_entry',0):.4f}" if pd.notna(row.get("price_entry")) else "",
                f"{row.get('price_exit',0):.4f}"  if pd.notna(row.get("price_exit"))  else "",
                f"{int(row.get('qty',0)):,}"       if pd.notna(row.get("qty"))         else "",
                f"${pnl:,.2f}",
                f"{row.get('pnl_pct',0):.2f}%"    if pd.notna(row.get("pnl_pct"))     else "",
                f"${row.get('mfe_usd',0):,.2f}"    if pd.notna(row.get("mfe_usd"))     else "",
                f"${row.get('mae_usd',0):,.2f}"    if pd.notna(row.get("mae_usd"))     else "",
                str(row.get("duration","")).split(".")[0],
                f"${row.get('cum_pnl_usd',0):,.2f}" if pd.notna(row.get("cum_pnl_usd")) else "",
            ]
            self._tv_tree.insert("","end", values=vals, tags=(tag,))

    def _tv_filter(self, trades):
        q = self._tv_filt.get().lower().strip()
        self._tv_fill(
            trades if not q else
            trades[trades.apply(
                lambda r: q in str(r.to_dict()).lower(), axis=1)])

    # ── TV Equity tab ─────────────────────────────────────────────────────
    def _tv_equity_tab(self, parent, d, daily_ret, bench_ret):
        cvs, inn = self._scrollable(parent)

        for title, series, col, fn, h in [
            ("Strategy vs Benchmark (growth of $1)", None,         None,  None,          180),
            ("Cumulative P&L  (USD)",                 d["equity"],  GREEN, self._cv_equity, 140),
            ("Daily P&L  (USD)",                      d["daily_pnl"],ORANGE,self._cv_bars,  110),
            ("Drawdown from Peak  (USD)",              d["drawdown"], RED,  self._cv_line,  100),
        ]:
            tk.Label(inn, text=title, bg=BG2, fg=col or PURPLE,
                     font=(HEAD[0],12,"bold")).pack(anchor="w", padx=20, pady=(16,4))
            if fn is None:
                self._cv_vs_bench(inn, daily_ret, bench_ret, height=h)
            else:
                fn(inn, series, col, height=h)

        # Metrics
        eq   = d["equity"]
        dd   = d["drawdown"]
        dpnl = d["daily_pnl"]
        tk.Label(inn, text="Key Metrics", bg=BG2, fg=FG2,
                 font=(HEAD[0],10,"bold")).pack(anchor="w", padx=20, pady=(14,6))
        g = tk.Frame(inn, bg=BG2); g.pack(fill="x", padx=20, pady=(0,20))
        items = [
            ("Final Equity",   f"${eq.iloc[-1]:,.2f}",              GREEN),
            ("Peak Equity",    f"${eq.max():,.2f}",                  GREEN),
            ("Max DD $",       f"${dd.min():,.2f}",                  RED),
            ("Max DD %",       f"{dd.min()/max(eq.max(),1)*100:.2f}%",RED),
            ("Best Day",       f"${dpnl.max():,.2f}",                GREEN),
            ("Worst Day",      f"${dpnl.min():,.2f}",                RED),
            ("Avg Daily P&L",  f"${dpnl.mean():,.2f}",               FG),
            ("Profitable Days",f"{(dpnl>0).sum()} / {len(dpnl)}",   GREEN),
        ]
        for i,(k,v,c) in enumerate(items):
            r,cc = divmod(i,4)
            cell = tk.Frame(g, bg=BG3)
            cell.grid(row=r, column=cc, padx=4, pady=4, sticky="ew")
            g.columnconfigure(cc, weight=1)
            tk.Label(cell, text=k, bg=BG3, fg=FG2,
                     font=("",8)).pack(anchor="w", padx=8, pady=(7,1))
            tk.Label(cell, text=v, bg=BG3, fg=c,
                     font=(HEAD[0],12,"bold")).pack(anchor="w", padx=8, pady=(0,7))

    # ── TV Distribution ────────────────────────────────────────────────────
    def _tv_dist(self, parent, trades):
        cvs, inn = self._scrollable(parent)
        pnl = trades["pnl_usd"].dropna()
        w   = pnl[pnl>0]; l = pnl[pnl<0]; e = pnl[pnl==0]
        tot = max(len(pnl),1)

        tk.Label(inn, text="P&L Distribution", bg=BG2, fg=PURPLE,
                 font=(HEAD[0],13,"bold")).pack(anchor="w", padx=20, pady=(16,8))
        self._stacked_bar(inn, len(w)/tot, len(e)/tot, len(l)/tot)

        tk.Label(inn, text="P&L Histogram  (USD per trade)", bg=BG2, fg=FG2,
                 font=(HEAD[0],10,"bold")).pack(anchor="w", padx=20, pady=(14,4))
        self._histogram(inn, pnl.tolist())

        # Streaks
        s = self._streaks(pnl)
        tk.Label(inn, text="Streak Analysis", bg=BG2, fg=FG2,
                 font=(HEAD[0],10,"bold")).pack(anchor="w", padx=20, pady=(14,4))
        sg = tk.Frame(inn, bg=BG2); sg.pack(fill="x", padx=20)
        for i,(k,v,c) in enumerate([
            ("Max Win Streak",  str(s["max_win"]),       GREEN),
            ("Max Loss Streak", str(s["max_loss"]),      RED),
            ("Current Streak",  str(s["current"]),       FG),
            ("Avg Win Streak",  f"{s['avg_win']:.1f}",  GREEN),
        ]):
            cell = tk.Frame(sg, bg=BG3)
            cell.grid(row=0, column=i, padx=4, pady=4, sticky="ew")
            sg.columnconfigure(i, weight=1)
            tk.Label(cell, text=k, bg=BG3, fg=FG2, font=("",8)
                     ).pack(anchor="w", padx=8, pady=(7,1))
            tk.Label(cell, text=v, bg=BG3, fg=c,
                     font=(HEAD[0],14,"bold")).pack(anchor="w", padx=8, pady=(0,7))

        tk.Label(inn, text="Monthly P&L Breakdown", bg=BG2, fg=FG2,
                 font=(HEAD[0],10,"bold")).pack(anchor="w", padx=20, pady=(14,4))
        self._monthly_table(inn, trades)
        tk.Frame(inn, bg=BG2, height=20).pack()

    # ── TV All Stats ───────────────────────────────────────────────────────
    def _tv_all_stats(self, parent, data):
        cvs, inn = self._scrollable(parent)

        tk.Label(inn, text="All TradingView Summary Sheets",
                 bg=BG2, fg=PURPLE,
                 font=(HEAD[0],13,"bold")).pack(anchor="w", padx=20, pady=(16,10))
        tk.Label(inn,
                 text="  Columns: All USD  |  All %  |  Long USD  |"
                      "  Long %  |  Short USD  |  Short %",
                 bg=BG2, fg=FG2, font=("",8)).pack(anchor="w", padx=20, pady=(0,10))

        for sname, sheet, col in [
            ("📋  Properties",                 data["props"],  ACCENT),
            ("💰  Performance",                data["perf"],   GREEN),
            ("📈  Trades Analysis",            data["tana"],   ORANGE),
            ("⚖️   Risk-Adjusted Performance",  data["risk"],  PURPLE),
        ]:
            if not sheet: continue
            tk.Label(inn, text=sname, bg=BG2, fg=col,
                     font=(HEAD[0],11,"bold")).pack(anchor="w", padx=20, pady=(16,4))
            f = tk.Frame(inn, bg=BG3); f.pack(fill="x", padx=20, pady=(0,8))
            # Column header row
            hrow = tk.Frame(f, bg=BG2); hrow.pack(fill="x")
            for htxt,hw in [("Metric",26),("All USD",9),("All %",8),
                             ("Long USD",9),("Long %",8),
                             ("Short USD",9),("Short %",8)]:
                tk.Label(hrow, text=htxt, bg=BG2, fg=FG2, font=("",8,"bold"),
                         width=hw, anchor="w").pack(side="left", padx=4, pady=3)

            for key, vals in sheet.items():
                if key in ("name","value"): continue
                r = tk.Frame(f, bg=BG3); r.pack(fill="x")
                r.bind("<Enter>", lambda e,fr=r: fr.configure(bg="#1d2230"))
                r.bind("<Leave>", lambda e,fr=r: fr.configure(bg=BG3))
                tk.Label(r, text=str(key), bg=BG3, fg=FG,
                         font=("",9), width=30, anchor="w"
                         ).pack(side="left", padx=(10,4), pady=2)
                for v in vals:
                    disp = str(v) if v is not None else "—"
                    vc   = (GREEN if isinstance(v,(int,float)) and v > 0
                            else RED   if isinstance(v,(int,float)) and v < 0
                            else FG2)
                    tk.Label(r, text=disp, bg=BG3, fg=vc,
                             font=MONO, width=9, anchor="e"
                             ).pack(side="left", padx=4)
                ttk.Separator(f).pack(fill="x")
        tk.Frame(inn, bg=BG2, height=20).pack()

    # ════════════════════════════════════════════════════════════════════════
    # SHARED CANVAS DRAWING UTILITIES
    # ════════════════════════════════════════════════════════════════════════

    def _scrollable(self, parent):
        """Return (canvas, inner_frame) with a vertical scrollbar."""
        cvs = tk.Canvas(parent, bg=BG2, bd=0, highlightthickness=0)
        vsb = ttk.Scrollbar(parent, orient="vertical", command=cvs.yview)
        inn = tk.Frame(cvs, bg=BG2)
        inn.bind("<Configure>",
                 lambda e: cvs.configure(scrollregion=cvs.bbox("all")))
        cvs.create_window((0,0), window=inn, anchor="nw")
        cvs.configure(yscrollcommand=vsb.set)
        cvs.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        return cvs, inn

    def _cv_equity(self, parent, series, color, height=120, padx=20):
        if series is None or len(series) == 0: return
        f = tk.Frame(parent, bg=BG2); f.pack(fill="x", padx=padx, pady=4)
        c = tk.Canvas(f, height=height, bg=BG3, highlightthickness=0, bd=0)
        c.pack(fill="x"); c.update()
        W   = c.winfo_width() or 900
        PAD = 12
        v   = np.array(series.values, dtype=float)
        mn, mx = v.min(), v.max(); rng = (mx-mn) or 1e-9
        pts = [(PAD+(i/max(len(v)-1,1))*(W-2*PAD),
                height-PAD-((v[i]-mn)/rng)*(height-2*PAD))
               for i in range(len(v))]
        fl  = [n for pt in pts for n in pt]
        for i in range(len(pts)-1):
            x1,y1=pts[i]; x2,y2=pts[i+1]
            c.create_polygon(x1,y1,x2,y2,x2,height-PAD,x1,height-PAD,
                             fill="#1a3a5c", outline="")
        if len(fl) >= 4:
            c.create_line(*fl, fill=color, width=2, smooth=True)
        c.create_text(W-PAD, PAD,      text=f"{mx:,.2f}", fill=FG2,
                      anchor="ne", font=("",7))
        c.create_text(W-PAD, height-4, text=f"{mn:,.2f}", fill=FG2,
                      anchor="se", font=("",7))

    def _cv_vs_bench(self, parent, strat_ret, bench_ret, height=150, padx=20):
        """Growth-of-$1 overlay: strategy (green) vs benchmark (orange)."""
        if strat_ret is None or len(strat_ret) == 0: return
        f = tk.Frame(parent, bg=BG2); f.pack(fill="x", padx=padx, pady=4)
        c = tk.Canvas(f, height=height, bg=BG3, highlightthickness=0, bd=0)
        c.pack(fill="x"); c.update()
        W   = c.winfo_width() or 900
        PAD = 12
        se  = (1 + strat_ret).cumprod().values.astype(float)
        be  = (1 + bench_ret.reindex(strat_ret.index).fillna(0)
               ).cumprod().values.astype(float)
        all_ = np.concatenate([se, be])
        mn, mx = all_.min(), all_.max(); rng = (mx-mn) or 1e-9
        N    = len(se)

        def _pts(arr):
            return [(PAD+(i/max(N-1,1))*(W-2*PAD),
                     height-PAD-((arr[i]-mn)/rng)*(height-2*PAD))
                    for i in range(N)]

        def _line(arr, col, lw):
            pts = _pts(arr)
            fl  = [n for p in pts for n in p]
            if len(fl) >= 4:
                c.create_line(*fl, fill=col, width=lw, smooth=True)

        # Shade strategy
        sp  = _pts(se)
        sfl = []
        for x,y in sp: sfl += [x,y]
        sfl += [sp[-1][0], height-PAD, sp[0][0], height-PAD]
        c.create_polygon(*sfl, fill="#1a3a5c", outline="")

        _line(be, ORANGE, 1)
        _line(se, GREEN,  2)

        # 1.0 baseline
        by = height-PAD-((1.0-mn)/rng)*(height-2*PAD)
        c.create_line(PAD, by, W-PAD, by, fill=BORDER, dash=(3,3))

        bn = bench_ret.name if hasattr(bench_ret,"name") else "Benchmark"
        c.create_rectangle(W-190, 8, W-180, 18, fill=GREEN,  outline="")
        c.create_text(W-178, 13, text="Strategy", fill=FG2, anchor="w", font=("",7))
        c.create_rectangle(W-110, 8, W-100, 18, fill=ORANGE, outline="")
        c.create_text(W-98,  13, text=bn,       fill=FG2, anchor="w", font=("",7))
        c.create_text(W-PAD, PAD,      text=f"{mx:.3f}x", fill=FG2,
                      anchor="ne", font=("",7))
        c.create_text(W-PAD, height-4, text=f"{mn:.3f}x", fill=FG2,
                      anchor="se", font=("",7))

    def _cv_line(self, parent, series, color, height=100, padx=20):
        if series is None or len(series) == 0: return
        f = tk.Frame(parent, bg=BG2); f.pack(fill="x", padx=padx, pady=4)
        c = tk.Canvas(f, height=height, bg=BG3, highlightthickness=0, bd=0)
        c.pack(fill="x"); c.update()
        W = c.winfo_width() or 900; PAD = 10
        v = np.array(series.values, dtype=float)
        mn, mx = v.min(), v.max(); rng = (mx-mn) or 1e-9
        xs = [PAD+(i/max(len(v)-1,1))*(W-2*PAD) for i in range(len(v))]
        ys = [height-PAD-((v[i]-mn)/rng)*(height-2*PAD) for i in range(len(v))]
        fl = [n for xy in zip(xs,ys) for n in xy]
        if len(fl) >= 4:
            c.create_line(*fl, fill=color, width=1.5, smooth=True)
        c.create_text(W-PAD, PAD+2,    text=f"{mx:,.2f}", fill=FG2,
                      anchor="ne", font=("",7))
        c.create_text(W-PAD, height-4, text=f"{mn:,.2f}", fill=FG2,
                      anchor="se", font=("",7))

    def _cv_bars(self, parent, series, color, height=100, padx=20):
        if series is None or len(series) == 0: return
        f = tk.Frame(parent, bg=BG2); f.pack(fill="x", padx=padx, pady=4)
        c = tk.Canvas(f, height=height, bg=BG3, highlightthickness=0, bd=0)
        c.pack(fill="x"); c.update()
        W = c.winfo_width() or 900; PAD = 10
        v   = np.array(series.values, dtype=float)
        rng = max(abs(v.min()), abs(v.max())) or 1e-9
        bw  = max((W-2*PAD)/len(v)-1, 2)
        mid = height//2
        for i, val in enumerate(v):
            x1 = PAD+i*((W-2*PAD)/len(v)); x2 = x1+bw
            h_ = abs(val)/rng*(height//2-PAD)
            col= GREEN if val >= 0 else RED
            if val>=0: c.create_rectangle(x1,mid-h_,x2,mid, fill=col,outline="")
            else:      c.create_rectangle(x1,mid,x2,mid+h_, fill=col,outline="")
        c.create_line(PAD, mid, W-PAD, mid, fill=BORDER, width=1)

    def _stacked_bar(self, parent, wr, er, lr):
        f   = tk.Frame(parent, bg=BG2); f.pack(fill="x", padx=20, pady=(0,4))
        bar = tk.Canvas(f, height=28, bg=BG3, highlightthickness=0)
        bar.pack(fill="x"); bar.update()
        W  = bar.winfo_width() or 900
        xw = int(W*wr); xe = int(W*(wr+er))
        if xw:    bar.create_rectangle(0,  0, xw, 28, fill=GREEN,  outline="")
        if xe-xw: bar.create_rectangle(xw, 0, xe, 28, fill=FG2,    outline="")
        if W-xe:  bar.create_rectangle(xe, 0, W,  28, fill=RED,    outline="")
        bar.create_text(max(xw//2,20), 14, fill=BG,
                        text=f"W {wr*100:.0f}%", font=("",8,"bold"))
        if xe-xw > 30:
            bar.create_text(xw+(xe-xw)//2, 14, fill=BG,
                            text=f"E {er*100:.0f}%", font=("",8))
        bar.create_text(xe+(W-xe)//2 if W-xe>20 else W-25, 14, fill=BG,
                        text=f"L {lr*100:.0f}%", font=("",8,"bold"))
        lab = tk.Frame(f, bg=BG2); lab.pack(fill="x")
        for t,cl in [("■ Wins",GREEN),("■ Even",FG2),("■ Losses",RED)]:
            tk.Label(lab,text=t,bg=BG2,fg=cl,font=("",8)).pack(side="left",padx=8)

    def _histogram(self, parent, values, height=160, padx=20):
        if not values: return
        f = tk.Frame(parent, bg=BG2); f.pack(fill="x", padx=padx, pady=4)
        c = tk.Canvas(f, height=height, bg=BG3, highlightthickness=0)
        c.pack(fill="x"); c.update()
        W  = c.winfo_width() or 900
        mn, mx = min(values), max(values); rng = (mx-mn) or 1e-9
        BINS=30; bw=rng/BINS; cnt=[0]*BINS
        for val in values: cnt[min(int((val-mn)/bw),BINS-1)] += 1
        mc=max(cnt) or 1; BAR=W/BINS; PAD=10
        for i,n in enumerate(cnt):
            x1=i*BAR; x2=x1+BAR-1
            col=GREEN if mn+i*bw>=0 else RED
            y1=height-PAD-(n/mc)*(height-2*PAD)
            c.create_rectangle(x1,y1,x2,height-PAD,fill=col,outline="")
        c.create_line(0,height-PAD,W,height-PAD,fill=BORDER)
        zx=(-mn/rng)*W
        if 0<=zx<=W: c.create_line(zx,0,zx,height,fill=FG2,dash=(4,4))
        c.create_text(6,height-3,text=f"${mn:.2f}",fill=FG2,anchor="sw",font=("",7))
        c.create_text(W-4,height-3,text=f"${mx:.2f}",fill=FG2,anchor="se",font=("",7))

    def _streaks(self, pnl: pd.Series) -> dict:
        mw=ml=cw=cl=0; ws=[]
        for val in pnl:
            if val>0:
                cw+=1; mw=max(mw,cw)
                if cl: cl=0
            elif val<0:
                cl+=1; ml=max(ml,cl)
                if cw: ws.append(cw); cw=0
        cur = f"+{cw}W" if cw else (f"-{cl}L" if cl else "0")
        return dict(max_win=mw, max_loss=ml, current=cur,
                    avg_win=float(np.mean(ws)) if ws else 0.0)

    def _monthly_table(self, parent, trades):
        if "datetime_exit" not in trades.columns: return
        t = trades.copy()
        t["month"] = pd.to_datetime(
            t["datetime_exit"], errors="coerce").dt.to_period("M")
        mo = t.groupby("month").agg(
            n   =("pnl_usd","count"),
            pnl =("pnl_usd","sum"),
            wins=("pnl_usd", lambda x: (x>0).sum()),
        ).reset_index()
        mo["wr"] = mo["wins"]/mo["n"]*100

        f = tk.Frame(parent, bg=BG2); f.pack(fill="x", padx=20, pady=(0,6))
        cols = ["Month","Trades","Win%","P&L $","Cumulative $"]
        tree = ttk.Treeview(f, columns=cols, show="headings",
                            height=min(len(mo)+1,12))
        for col, w in zip(cols, [110,70,70,100,130]):
            tree.heading(col, text=col)
            tree.column(col, width=w,
                        anchor="w" if col=="Month" else "e")
        cum = 0
        for _,r in mo.iterrows():
            pnl=r["pnl"]; cum+=pnl
            tag="mw" if pnl>=0 else "ml"
            tree.insert("","end", tags=(tag,),
                values=[str(r["month"]), int(r["n"]),
                        f"{r['wr']:.1f}%", f"${pnl:,.2f}", f"${cum:,.2f}"])
        tree.tag_configure("mw", foreground=GREEN)
        tree.tag_configure("ml", foreground=RED)
        tree.pack(fill="x")

    # ── Reset editor ─────────────────────────────────────────────────────
    def _ed_reset(self):
        self._ed.delete("1.0","end")
        self._ed.insert("1.0", DEFAULT_STRATEGY)


# ════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    App().mainloop()

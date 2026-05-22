#!/usr/bin/env python3
"""
QuantStats Strategy Backtesting Report Generator  v2.0
+ TradingView XLSX Trade Analyzer Module

FIXES:
  - datetime64[us, UTC] vs datetime64[us] QuantStats error
    (strips tz before passing to qs.reports.html)
  - Robust CSV / Alpaca data loading with full timezone normalisation
NEW:
  - TradingView .xlsx analyzer tab (List of trades + all summary sheets)
"""

# ─── AUTO-INSTALL DEPENDENCIES ──────────────────────────────────────────────
import subprocess, sys

REQUIRED = {
    "quantstats": "quantstats",
    "pandas":     "pandas",
    "numpy":      "numpy",
    "pytz":       "pytz",
    "openpyxl":   "openpyxl",
}

def install_if_missing():
    for mod, pkg in REQUIRED.items():
        try:
            __import__(mod)
        except ImportError:
            print(f"Installing {pkg}…")
            subprocess.check_call(
                [sys.executable, "-m", "pip", "install", pkg, "-q",
                 "--break-system-packages"],
                stderr=subprocess.DEVNULL,
            )

install_if_missing()

# ─── IMPORTS ────────────────────────────────────────────────────────────────
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading, traceback, os, tempfile, webbrowser, re, math
import pandas as pd
import numpy as np
import pytz
import quantstats as qs
from datetime import datetime
import openpyxl

# ─── THEME ──────────────────────────────────────────────────────────────────
BG     = "#0d1117"
BG2    = "#161b22"
BG3    = "#21262d"
BORDER = "#30363d"
ACCENT = "#58a6ff"
GREEN  = "#3fb950"
WARN   = "#d29922"
RED    = "#f85149"
PURPLE = "#bc8cff"
ORANGE = "#ffa657"
FG     = "#e6edf3"
FG2    = "#8b949e"

_P  = sys.platform
MONO = ("Cascadia Code", 10) if _P == "win32" else ("Menlo", 10)
UI   = ("Segoe UI",      10) if _P == "win32" else ("SF Pro Display", 10)
HEAD = ("Segoe UI Semibold", 11) if _P == "win32" else ("SF Pro Display", 11)

TF_MAP = {
    "1 Min": "1Min", "5 Min": "5Min", "15 Min": "15Min",
    "30 Min": "30Min", "1 Hour": "1Hour", "4 Hour": "4Hour", "1 Day": "1Day",
}

DEFAULT_STRATEGY = '''\
# ══════════════════════════════════════════════════════════════════════════
# STRATEGY CODE  —  Required Function Signature:
#
#   def strategy(data: pd.DataFrame) -> pd.Series
#
# INPUT  `data` columns (all lowercase, index = DatetimeTZAware UTC):
#   open, high, low, close  – OHLC prices (float)
#   volume                  – volume (float)
#   vwap                    – volume-weighted avg price (float)
#
# OUTPUT  pd.Series of SIGNALS aligned to data.index:
#   +1  = Long   |   0  = Flat   |  -1  = Short
#
# Signals are auto-shifted by 1 bar → fills at next-bar open (no look-ahead).
# ══════════════════════════════════════════════════════════════════════════

import pandas as pd
import numpy as np

def strategy(data: pd.DataFrame) -> pd.Series:
    """Simple Golden Cross: SMA 20 / SMA 50"""
    close = data["close"]
    fast  = close.rolling(20).mean()
    slow  = close.rolling(50).mean()

    signal = pd.Series(0, index=data.index, dtype=float)
    signal[fast > slow]  = 1    # long
    signal[fast <= slow] = 0    # flat
    return signal
'''

# ════════════════════════════════════════════════════════════════════════════
# DATA LAYER
# ════════════════════════════════════════════════════════════════════════════

def detect_and_load_csv(filepath: str) -> pd.DataFrame:
    df = pd.read_csv(filepath)
    df.columns = [c.strip().lower() for c in df.columns]
    ts_candidates = ["timestamp","datetime","date","time","t",
                     "date_time","bar_time","open_time"]
    ts_col = next((c for c in ts_candidates if c in df.columns), None)
    if ts_col is None:
        for c in df.columns:
            if re.search(r"\d{4}-\d{2}-\d{2}", str(df[c].iloc[0])):
                ts_col = c; break
    if ts_col is None:
        raise ValueError(
            "Cannot detect timestamp column. "
            "Expected: 'timestamp', 'datetime', or 'date'.")
    df["timestamp"] = pd.to_datetime(df[ts_col], utc=True, errors="coerce")
    df = df.dropna(subset=["timestamp"]).set_index("timestamp").sort_index()
    aliases = {
        "open":   ["open","o","open_price"],
        "high":   ["high","h","high_price"],
        "low":    ["low","l","low_price"],
        "close":  ["close","c","last","adj_close","close_price"],
        "volume": ["volume","vol","v","qty","quantity"],
        "vwap":   ["vwap","vw","weighted_avg"],
    }
    for canon, alts in aliases.items():
        if canon not in df.columns:
            for a in alts:
                if a in df.columns:
                    df.rename(columns={a: canon}, inplace=True); break
    miss = [r for r in ["open","high","low","close"] if r not in df.columns]
    if miss:
        raise ValueError(f"CSV missing required columns: {miss}. Found: {list(df.columns)}")
    for col in ["volume","vwap"]:
        if col not in df.columns:
            df[col] = df["close"] if col == "vwap" else 0.0
    for col in ["open","high","low","close","volume","vwap"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    return df[["open","high","low","close","volume","vwap"]].dropna(
        subset=["open","high","low","close"])


def fetch_alpaca_data(ticker, start, end, tf_label, api_key, secret_key):
    try:
        from alpaca.data.historical import StockHistoricalDataClient
        from alpaca.data.requests   import StockBarsRequest
        from alpaca.data.timeframe  import TimeFrame, TimeFrameUnit
    except ImportError:
        raise RuntimeError("alpaca-py not installed. Run: pip install alpaca-py")
    if not api_key or not secret_key:
        raise ValueError("Alpaca API Key and Secret Key are required.")
    tf_map = {
        "1Min":  TimeFrame(1, TimeFrameUnit.Minute),
        "5Min":  TimeFrame(5, TimeFrameUnit.Minute),
        "15Min": TimeFrame(15,TimeFrameUnit.Minute),
        "30Min": TimeFrame(30,TimeFrameUnit.Minute),
        "1Hour": TimeFrame(1, TimeFrameUnit.Hour),
        "4Hour": TimeFrame(4, TimeFrameUnit.Hour),
        "1Day":  TimeFrame(1, TimeFrameUnit.Day),
    }
    tf_str  = TF_MAP.get(tf_label, "1Day")
    tz_ny   = pytz.timezone("America/New_York")
    s_dt    = datetime.strptime(start, "%Y-%m-%d").replace(tzinfo=tz_ny)
    e_dt    = datetime.strptime(end,   "%Y-%m-%d").replace(
                  hour=23, minute=59, second=59, tzinfo=tz_ny)
    client  = StockHistoricalDataClient(api_key=api_key, secret_key=secret_key)
    req     = StockBarsRequest(symbol_or_symbols=ticker.upper(),
                               timeframe=tf_map[tf_str],
                               start=s_dt, end=e_dt, adjustment="all")
    df = client.get_stock_bars(req).df
    if df.empty:
        raise ValueError(f"No data returned for {ticker} in the given date range.")
    if isinstance(df.index, pd.MultiIndex):
        df = df.reset_index(level=0, drop=True)
    # Normalise to UTC-aware
    df.index = pd.to_datetime(df.index, utc=True)
    df.index.name = "timestamp"
    df.columns = [c.lower() for c in df.columns]
    if "volume" not in df.columns: df["volume"] = 0.0
    if "vwap"   not in df.columns: df["vwap"]   = df["close"]
    return df[["open","high","low","close","volume","vwap"]].sort_index()


# ════════════════════════════════════════════════════════════════════════════
# BACKTEST ENGINE
# ════════════════════════════════════════════════════════════════════════════

def _strip_tz(s: pd.Series) -> pd.Series:
    """
    FIX: 'Cannot compare dtypes datetime64[us, UTC] and datetime64[us]'
    QuantStats / Matplotlib require a tz-naive index.
    We convert UTC → tz-naive UTC (values unchanged).
    """
    if hasattr(s.index, "tz") and s.index.tz is not None:
        s = s.copy()
        s.index = s.index.tz_convert("UTC").tz_localize(None)
    return s


def run_backtest(data: pd.DataFrame, code: str, name: str) -> dict:
    ns = {"pd": pd, "np": np, "__builtins__": __builtins__}
    try:
        exec(compile(code, "<strategy>", "exec"), ns)
    except Exception as e:
        raise RuntimeError(f"Strategy code compilation error:\n{e}")
    if "strategy" not in ns:
        raise ValueError("Strategy code must define a function named `strategy(data)`.")
    try:
        sigs = ns["strategy"](data.copy())
    except Exception:
        raise RuntimeError(f"Strategy execution error:\n{traceback.format_exc()}")
    if not isinstance(sigs, pd.Series):
        raise ValueError("`strategy()` must return a pandas Series of signals.")

    sigs  = sigs.reindex(data.index).fillna(0).shift(1).fillna(0)
    pret  = data["close"].pct_change().fillna(0)
    port  = (sigs * pret).rename(name)
    port.index = pd.to_datetime(port.index, utc=True)

    daily = port.resample("D").apply(lambda x: (1+x).prod()-1).dropna()
    daily = daily[daily.abs() < 1]
    if daily.empty:
        raise ValueError("Backtest produced no returns. Check date range / strategy.")

    bh = pret.resample("D").apply(lambda x: (1+x).prod()-1).dropna()
    bh = bh.reindex(daily.index).fillna(0)
    bh.name = "Buy & Hold"

    # ── KEY FIX ──────────────────────────────────────────────────────────
    daily_qs = _strip_tz(daily)
    bh_qs    = _strip_tz(bh)
    # ─────────────────────────────────────────────────────────────────────

    html = os.path.join(tempfile.gettempdir(),
        f"backtest_{re.sub(r'[^a-zA-Z0-9]','_',name)}"
        f"_{datetime.now():%Y%m%d_%H%M%S}.html")

    qs.extend_pandas()
    try:
        qs.reports.html(returns=daily_qs, benchmark=bh_qs,
                        title=name, output=html, download_filename=html)
    except Exception as e:
        raise RuntimeError(f"QuantStats report generation failed:\n{e}")

    def _s(fn, *a, **kw):
        try:    return fn(*a, **kw)
        except: return float("nan")

    summary = {
        "Total Return":     f"{_s(qs.stats.comp, daily_qs)*100:.2f}%",
        "CAGR":             f"{_s(qs.stats.cagr, daily_qs)*100:.2f}%",
        "Sharpe Ratio":     f"{_s(qs.stats.sharpe, daily_qs):.3f}",
        "Sortino Ratio":    f"{_s(qs.stats.sortino, daily_qs):.3f}",
        "Max Drawdown":     f"{_s(qs.stats.max_drawdown, daily_qs)*100:.2f}%",
        "Volatility (Ann)": f"{_s(qs.stats.volatility, daily_qs)*100:.2f}%",
        "Win Rate":         f"{_s(qs.stats.win_rate, daily_qs)*100:.2f}%",
        "Calmar Ratio":     f"{_s(qs.stats.calmar, daily_qs):.3f}",
        "Skew":             f"{_s(qs.stats.skew, daily_qs):.3f}",
        "Kurtosis":         f"{_s(qs.stats.kurtosis, daily_qs):.3f}",
        "VaR (95%)":        f"{_s(qs.stats.value_at_risk, daily_qs)*100:.2f}%",
        "CVaR (95%)":       f"{_s(qs.stats.cvar, daily_qs)*100:.2f}%",
        "Bars in Market":   f"{int((sigs.abs()>0).sum()):,}",
        "Total Bars":       f"{len(data):,}",
    }
    return dict(summary=summary, html_path=html,
                daily_rets=daily, data_rows=len(data))


# ════════════════════════════════════════════════════════════════════════════
# TRADINGVIEW XLSX ANALYZER
# ════════════════════════════════════════════════════════════════════════════

def load_tv_xlsx(filepath: str) -> dict:
    wb     = openpyxl.load_workbook(filepath, data_only=True)
    smap   = {s.lower(): s for s in wb.sheetnames}

    def _sheet_df(key):
        name = next((v for k,v in smap.items() if key in k), None)
        if name is None: return pd.DataFrame()
        ws   = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows: return pd.DataFrame()
        hdrs = [str(h) if h is not None else f"c{i}"
                for i,h in enumerate(rows[0])]
        return pd.DataFrame(rows[1:], columns=hdrs)

    def _kv_sheet(key):
        name = next((v for k,v in smap.items() if key in k), None)
        if name is None: return {}
        ws  = wb[name]
        out = {}
        for row in ws.iter_rows(values_only=True):
            if row[0] is not None:
                out[str(row[0])] = row[1:]
        return out

    # List of trades
    lot = _sheet_df("list of trade")
    lot.columns = [c.strip() for c in lot.columns]
    col_map = {
        "Trade #": "trade_num", "Type": "type",
        "Date and time": "datetime", "Signal": "signal",
        "Price USD": "price", "Position size (qty)": "qty",
        "Position size (value)": "value",
        "Net P&L USD": "pnl_usd", "Net P&L %": "pnl_pct",
        "Favorable excursion USD": "mfe_usd", "Favorable excursion %": "mfe_pct",
        "Adverse excursion USD":   "mae_usd", "Adverse excursion %":  "mae_pct",
        "Cumulative P&L USD": "cum_pnl_usd", "Cumulative P&L %": "cum_pnl_pct",
    }
    lot.rename(columns={k:v for k,v in col_map.items() if k in lot.columns}, inplace=True)
    lot["datetime"] = pd.to_datetime(lot["datetime"], errors="coerce")
    for col in ["price","qty","value","pnl_usd","pnl_pct",
                "mfe_usd","mfe_pct","mae_usd","mae_pct","cum_pnl_usd","cum_pnl_pct"]:
        if col in lot.columns:
            lot[col] = pd.to_numeric(lot[col], errors="coerce")

    # Split entry / exit and merge
    ent = lot[lot["type"].str.lower().str.contains("entry", na=False)].copy()
    ext = lot[lot["type"].str.lower().str.contains("exit",  na=False)].copy()
    ent = ent.sort_values("trade_num").drop_duplicates("trade_num", keep="last")
    ext = ext.sort_values("trade_num").drop_duplicates("trade_num", keep="last")
    trades = ent.merge(
        ext[["trade_num","datetime","price","pnl_usd","pnl_pct",
             "mfe_usd","mfe_pct","mae_usd","mae_pct","cum_pnl_usd","signal"]],
        on="trade_num", suffixes=("_entry","_exit"), how="inner")
    trades["duration"]  = trades["datetime_exit"] - trades["datetime_entry"]
    trades["direction"] = trades["type"].str.extract(
        r"(long|short)", flags=re.I)[0].str.lower()
    trades.rename(columns={"pnl_usd_exit":"pnl_usd",
                            "pnl_pct_exit":"pnl_pct"}, inplace=True)
    trades = trades.reset_index(drop=True)

    pnl     = trades["pnl_usd"].dropna()
    winners = pnl[pnl > 0]
    losers  = pnl[pnl < 0]
    evens   = pnl[pnl == 0]

    trades["date"]  = pd.to_datetime(trades["datetime_exit"], errors="coerce").dt.date
    daily_pnl       = trades.groupby("date")["pnl_usd"].sum()
    equity          = daily_pnl.cumsum()
    roll_max        = equity.cummax()
    drawdown        = equity - roll_max

    derived = {
        "total_trades":  len(trades),
        "winners":       len(winners),
        "losers":        len(losers),
        "evens":         len(evens),
        "win_rate":      len(winners) / max(len(trades)-len(evens), 1) * 100,
        "avg_win":       float(winners.mean())    if len(winners) else 0.0,
        "avg_loss":      float(losers.mean())     if len(losers)  else 0.0,
        "total_pnl":     float(pnl.sum()),
        "gross_profit":  float(winners.sum())     if len(winners) else 0.0,
        "gross_loss":    float(losers.sum())      if len(losers)  else 0.0,
        "profit_factor": abs(winners.sum()/losers.sum()) if losers.sum() != 0 else float("inf"),
        "largest_win":   float(winners.max())     if len(winners) else 0.0,
        "largest_loss":  float(losers.min())      if len(losers)  else 0.0,
        "max_drawdown":  float(drawdown.min()),
        "avg_duration":  trades["duration"].mean(),
        "equity":        equity,
        "drawdown":      drawdown,
        "daily_pnl":     daily_pnl,
    }
    return dict(trades=trades, lot_raw=lot,
                perf=_kv_sheet("performance"),
                tana=_kv_sheet("trades analysis"),
                risk=_kv_sheet("risk-adjusted"),
                props=_kv_sheet("properties"),
                derived=derived)


# ════════════════════════════════════════════════════════════════════════════
# GUI
# ════════════════════════════════════════════════════════════════════════════

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("QuantStats Backtest + TradingView Analyzer  v2.0")
        self.configure(bg=BG)
        self.geometry("1380x900")
        self.minsize(1150, 720)

        self._csv_path   = tk.StringVar()
        self._ticker     = tk.StringVar(value="AAPL")
        self._start      = tk.StringVar(value="2022-01-01")
        self._end        = tk.StringVar(value=datetime.today().strftime("%Y-%m-%d"))
        self._tf         = tk.StringVar(value="1 Day")
        self._src        = tk.StringVar(value="alpaca")
        self._api_key    = tk.StringVar()
        self._sec_key    = tk.StringVar()
        self._strat_name = tk.StringVar(value="My Strategy")
        self._status     = tk.StringVar(value="Ready.")
        self._html_path  = None
        self._tv_path    = tk.StringVar()

        self._sty()
        self._build()

    # ── styles ────────────────────────────────────────────────────────────
    def _sty(self):
        s = ttk.Style(self)
        s.theme_use("clam")
        base = dict(background=BG3, foreground=FG, fieldbackground=BG3,
                    insertbackground=FG, selectbackground=ACCENT,
                    selectforeground=BG, bordercolor=BORDER,
                    lightcolor=BORDER, darkcolor=BORDER,
                    relief="flat", padding=5)
        s.configure("TFrame",        background=BG)
        s.configure("TLabel",        background=BG,  foreground=FG,  font=UI)
        s.configure("Card.TLabel",   background=BG2, foreground=FG,  font=UI)
        s.configure("TEntry",        **base)
        s.configure("TCombobox",     **base)
        s.map("TCombobox", fieldbackground=[("readonly",BG3)],
              background=[("readonly",BG3)])
        s.configure("TButton",       background=BG3, foreground=FG,
                    bordercolor=BORDER, font=UI, padding=(10,5), relief="flat")
        s.map("TButton", background=[("active",BORDER)])
        s.configure("Run.TButton",   background=ACCENT, foreground=BG,
                    font=(UI[0],11,"bold"), padding=(16,8))
        s.map("Run.TButton",  background=[("active","#79c0ff")])
        s.configure("Open.TButton",  background=GREEN, foreground=BG,
                    font=(UI[0],10,"bold"), padding=(12,6))
        s.map("Open.TButton", background=[("active","#56d364")])
        s.configure("TV.TButton",    background=PURPLE, foreground=BG,
                    font=(UI[0],10,"bold"), padding=(12,6))
        s.map("TV.TButton",   background=[("active","#d2a8ff")])
        s.configure("TRadiobutton",  background=BG2, foreground=FG,
                    indicatorbackground=BG3, selectcolor=ACCENT)
        s.configure("TNotebook",     background=BG,  bordercolor=BORDER, tabmargins=0)
        s.configure("TNotebook.Tab", background=BG3, foreground=FG2,
                    padding=(14,7), bordercolor=BORDER)
        s.map("TNotebook.Tab",
              background=[("selected",BG2)], foreground=[("selected",FG)])
        s.configure("TProgressbar",  troughcolor=BG3, background=ACCENT,
                    bordercolor=BORDER, thickness=4)
        s.configure("Treeview",
                    background=BG3, foreground=FG, fieldbackground=BG3,
                    bordercolor=BORDER, relief="flat", rowheight=22,
                    font=("Cascadia Code",9) if _P=="win32" else ("Menlo",9))
        s.configure("Treeview.Heading",
                    background=BG2, foreground=FG2,
                    font=(UI[0],9,"bold"), relief="flat")
        s.map("Treeview",
              background=[("selected",ACCENT)],
              foreground=[("selected",BG)])

    # ── layout ───────────────────────────────────────────────────────────
    def _build(self):
        hdr = tk.Frame(self, bg=BG2, height=54)
        hdr.pack(fill="x"); hdr.pack_propagate(False)
        tk.Label(hdr, text="⚡  QuantStats Backtest Generator  +  TradingView Analyzer",
                 bg=BG2, fg=FG,
                 font=(HEAD[0],14,"bold")).pack(side="left", padx=20, pady=10)
        tk.Label(hdr, text="v2.0", bg=BG2, fg=FG2, font=("",9)
                 ).pack(side="right", padx=20)
        ttk.Separator(self).pack(fill="x")

        main = tk.Frame(self, bg=BG)
        main.pack(fill="both", expand=True)

        left = tk.Frame(main, bg=BG, width=390)
        left.pack(side="left", fill="y")
        left.pack_propagate(False)
        self._left_panel(left)

        tk.Frame(main, bg=BORDER, width=1).pack(side="left", fill="y")

        right = tk.Frame(main, bg=BG)
        right.pack(side="left", fill="both", expand=True)
        self._right_panel(right)

        sb = tk.Frame(self, bg=BG3, height=26)
        sb.pack(fill="x", side="bottom"); sb.pack_propagate(False)
        self._prog = ttk.Progressbar(sb, mode="indeterminate", length=120)
        self._prog.pack(side="right", padx=10, pady=5)
        tk.Label(sb, textvariable=self._status,
                 bg=BG3, fg=FG2, font=("",9), anchor="w"
                 ).pack(side="left", padx=10, pady=4)

    def _left_panel(self, p):
        cvs = tk.Canvas(p, bg=BG, bd=0, highlightthickness=0)
        vsb = ttk.Scrollbar(p, orient="vertical", command=cvs.yview)
        inner = tk.Frame(cvs, bg=BG)
        inner.bind("<Configure>",
                   lambda e: cvs.configure(scrollregion=cvs.bbox("all")))
        cvs.create_window((0,0), window=inner, anchor="nw")
        cvs.configure(yscrollcommand=vsb.set)
        cvs.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        M = 14

        def sec(title, col=ACCENT):
            f = tk.Frame(inner, bg=BG2)
            f.pack(fill="x", padx=M, pady=(10,0))
            tk.Label(f, text=title, bg=BG2, fg=col,
                     font=(HEAD[0],10,"bold")).pack(anchor="w", padx=10, pady=(8,4))
            return f

        def entry_row(par, lbl, var, show=None):
            r = tk.Frame(par, bg=BG2); r.pack(fill="x", padx=10, pady=3)
            tk.Label(r, text=lbl, bg=BG2, fg=FG2, font=("",9),
                     width=13, anchor="w").pack(side="left")
            ttk.Entry(r, textvariable=var, font=MONO, show=show
                      ).pack(side="left", fill="x", expand=True)

        # Strategy name
        f0 = sec("◈  STRATEGY")
        entry_row(f0, "Name", self._strat_name)

        # Data source
        f1 = sec("◈  DATA SOURCE")
        rb = tk.Frame(f1, bg=BG2); rb.pack(fill="x", padx=10, pady=4)
        for txt, val in [("Alpaca API","alpaca"),("CSV File","csv")]:
            ttk.Radiobutton(rb, text=txt, variable=self._src, value=val,
                            command=self._toggle_src).pack(side="left", padx=(0,14))

        self._alpaca_f = tk.Frame(f1, bg=BG2); self._alpaca_f.pack(fill="x")
        entry_row(self._alpaca_f, "API Key",    self._api_key)
        entry_row(self._alpaca_f, "Secret Key", self._sec_key, show="•")
        tk.Label(self._alpaca_f, text="  alpaca.markets → Paper → API Keys",
                 bg=BG2, fg=FG2, font=("",8)).pack(anchor="w", padx=10, pady=(0,6))

        self._csv_f = tk.Frame(f1, bg=BG2)
        cr = tk.Frame(self._csv_f, bg=BG2); cr.pack(fill="x", padx=10, pady=4)
        tk.Label(cr, text="CSV File", bg=BG2, fg=FG2, font=("",9),
                 width=13, anchor="w").pack(side="left")
        ttk.Entry(cr, textvariable=self._csv_path,
                  font=("",9)).pack(side="left", fill="x", expand=True, padx=(0,4))
        ttk.Button(cr, text="Browse",
                   command=lambda: self._browse(self._csv_path)).pack(side="left")
        tk.Label(self._csv_f,
                 text="  timestamp, open, high, low, close, volume",
                 bg=BG2, fg=FG2, font=("",8)).pack(anchor="w", padx=10, pady=(0,6))

        # Ticker
        f2 = sec("◈  TICKER & TIMEFRAME")
        entry_row(f2, "Ticker",     self._ticker)
        r = tk.Frame(f2, bg=BG2); r.pack(fill="x", padx=10, pady=3)
        tk.Label(r, text="Timeframe", bg=BG2, fg=FG2, font=("",9),
                 width=13, anchor="w").pack(side="left")
        ttk.Combobox(r, textvariable=self._tf, state="readonly",
                     values=list(TF_MAP.keys()), font=UI
                     ).pack(side="left", fill="x", expand=True)
        entry_row(f2, "Start Date", self._start)
        entry_row(f2, "End Date",   self._end)
        tk.Label(f2, text="  Format: YYYY-MM-DD",
                 bg=BG2, fg=FG2, font=("",8)).pack(anchor="w", padx=10, pady=(0,8))

        # Backtest buttons
        bf = tk.Frame(inner, bg=BG); bf.pack(fill="x", padx=M, pady=(12,0))
        ttk.Button(bf, text="▶  Run Backtest", style="Run.TButton",
                   command=self._run_bt).pack(fill="x")
        self._open_btn = ttk.Button(
            bf, text="🌐  Open QuantStats Tearsheet",
            style="Open.TButton", command=self._open_report, state="disabled")
        self._open_btn.pack(fill="x", pady=(8,0))

        # TradingView section
        f3 = sec("◈  TRADINGVIEW ANALYZER", col=PURPLE)
        tk.Label(f3,
                 text="  Load TradingView strategy backtest\n"
                      "  export (.xlsx)  — 'List of trades' sheet",
                 bg=BG2, fg=FG2, font=("",8), justify="left"
                 ).pack(anchor="w", padx=10, pady=(0,4))
        tvr = tk.Frame(f3, bg=BG2); tvr.pack(fill="x", padx=10, pady=(0,6))
        ttk.Entry(tvr, textvariable=self._tv_path,
                  font=("",9)).pack(side="left", fill="x", expand=True, padx=(0,4))
        ttk.Button(tvr, text="Browse",
                   command=lambda: self._browse(
                       self._tv_path,
                       [("Excel XLSX","*.xlsx"),("All","*.*")]
                   )).pack(side="left")

        tv_bf = tk.Frame(inner, bg=BG); tv_bf.pack(fill="x", padx=M, pady=(6,0))
        ttk.Button(tv_bf, text="📊  Analyze TradingView Trades",
                   style="TV.TButton",
                   command=self._run_tv).pack(fill="x")

        tk.Frame(inner, bg=BG, height=16).pack()
        self._toggle_src()

    def _right_panel(self, p):
        self._nb = ttk.Notebook(p)
        self._nb.pack(fill="both", expand=True)

        # Editor
        et = tk.Frame(self._nb, bg=BG2)
        self._nb.add(et, text="  Strategy Editor  ")
        tb = tk.Frame(et, bg=BG3, height=34); tb.pack(fill="x"); tb.pack_propagate(False)
        tk.Label(tb, text="strategy.py", bg=BG3, fg=FG2, font=MONO
                 ).pack(side="left", padx=12, pady=7)
        ttk.Button(tb, text="Reset", command=self._reset_ed
                   ).pack(side="right", padx=8, pady=5)
        self._ed = scrolledtext.ScrolledText(
            et, bg="#0d1117", fg="#e6edf3", font=MONO,
            insertbackground=FG, selectbackground=ACCENT, selectforeground=BG,
            relief="flat", bd=0, wrap="none", undo=True, tabs="1c")
        self._ed.pack(fill="both", expand=True)
        self._ed.insert("1.0", DEFAULT_STRATEGY)
        self._highlight()

        # Backtest results
        self._bt_tab = tk.Frame(self._nb, bg=BG2)
        self._nb.add(self._bt_tab, text="  Backtest Results  ")
        tk.Label(self._bt_tab, text="Run a backtest to see results here.",
                 bg=BG2, fg=FG2, font=(UI[0],12)).pack(expand=True)

        # TV Analysis
        self._tv_tab = tk.Frame(self._nb, bg=BG2)
        self._nb.add(self._tv_tab, text="  📊 TV Trade Analysis  ")
        tk.Label(self._tv_tab,
                 text="Load a TradingView XLSX export and click 'Analyze Trades'.",
                 bg=BG2, fg=FG2, font=(UI[0],12)).pack(expand=True)

        # Log
        log_t = tk.Frame(self._nb, bg=BG2)
        self._nb.add(log_t, text="  Log  ")
        self._log = scrolledtext.ScrolledText(
            log_t, bg="#0d1117", fg="#8b949e",
            font=("Cascadia Code",9) if _P=="win32" else ("Menlo",9),
            relief="flat", bd=0, state="disabled", wrap="word")
        self._log.pack(fill="both", expand=True)
        for tag, col in [("ok",GREEN),("err",RED),("warn",WARN),
                         ("info",ACCENT),("plain",FG2)]:
            self._log.tag_config(tag, foreground=col)

    # ── helpers ───────────────────────────────────────────────────────────
    def _toggle_src(self):
        if self._src.get() == "alpaca":
            self._alpaca_f.pack(fill="x"); self._csv_f.pack_forget()
        else:
            self._csv_f.pack(fill="x"); self._alpaca_f.pack_forget()

    def _browse(self, var, filetypes=None):
        ft = filetypes or [("CSV","*.csv"),("All","*.*")]
        p  = filedialog.askopenfilename(filetypes=ft)
        if p: var.set(p)

    def _reset_ed(self):
        self._ed.delete("1.0","end")
        self._ed.insert("1.0", DEFAULT_STRATEGY)

    def _highlight(self):
        self._ed.tag_config("kw",      foreground="#ff7b72")
        self._ed.tag_config("str",     foreground="#a5d6ff")
        self._ed.tag_config("comment", foreground="#8b949e")
        content = self._ed.get("1.0","end")
        for m in re.finditer(
                r"\b(def|class|import|from|return|if|else|elif|for|while|"
                r"in|not|and|or|True|False|None|try|except|raise|with|as|"
                r"lambda|pass|break|continue)\b", content):
            self._ed.tag_add("kw", f"1.0+{m.start()}c", f"1.0+{m.end()}c")
        for m in re.finditer(
                r'(""".*?"""|\'\'\'.*?\'\'\'|"[^"]*"|\'[^\']*\')',
                content, re.DOTALL):
            self._ed.tag_add("str", f"1.0+{m.start()}c", f"1.0+{m.end()}c")
        for m in re.finditer(r"#[^\n]*", content):
            self._ed.tag_add("comment", f"1.0+{m.start()}c", f"1.0+{m.end()}c")

    def _log_w(self, msg, tag="plain"):
        self._log.configure(state="normal")
        self._log.insert("end", f"[{datetime.now():%H:%M:%S}] {msg}\n", tag)
        self._log.see("end")
        self._log.configure(state="disabled")

    # ── backtest ──────────────────────────────────────────────────────────
    def _run_bt(self):
        self._prog.start(12)
        self._status.set("Running backtest…")
        self._log_w("═"*60, "info")
        self._log_w(f"Backtest: {self._strat_name.get()}", "info")
        threading.Thread(target=self._bt_task, daemon=True).start()

    def _bt_task(self):
        try:
            self._log_w("Loading data…", "plain")
            data = self._load_data()
            self._log_w(
                f"Data: {len(data):,} bars  "
                f"{data.index[0].date()} → {data.index[-1].date()}", "ok")
            result = run_backtest(data, self._ed.get("1.0","end"),
                                  self._strat_name.get())
            self._html_path = result["html_path"]
            self.after(0, lambda: self._show_bt(result))
            self._log_w("Complete!", "ok")
            for k,v in result["summary"].items():
                self._log_w(f"  {k:<24}{v}", "plain")
        except Exception as e:
            self._log_w(f"ERROR: {e}", "err")
            self._log_w(traceback.format_exc(), "err")
            self.after(0, lambda: messagebox.showerror("Backtest Error", str(e)))
            self.after(0, lambda: self._status.set(f"Error: {e}"))
        finally:
            self.after(0, self._prog.stop)

    def _load_data(self):
        src = self._src.get()
        if src == "alpaca":
            return fetch_alpaca_data(
                self._ticker.get().strip().upper(),
                self._start.get().strip(), self._end.get().strip(),
                self._tf.get(), self._api_key.get().strip(),
                self._sec_key.get().strip())
        path = self._csv_path.get().strip()
        if not path or not os.path.exists(path):
            raise ValueError("Please select a valid CSV file.")
        df = detect_and_load_csv(path)
        try:
            s = pd.Timestamp(self._start.get().strip(), tz="UTC")
            e = pd.Timestamp(self._end.get().strip(),   tz="UTC") + pd.Timedelta(days=1)
            df = df[(df.index >= s) & (df.index <= e)]
        except Exception: pass
        if df.empty: raise ValueError("CSV empty after date filtering.")
        return df

    def _show_bt(self, result):
        for w in self._bt_tab.winfo_children(): w.destroy()
        hdr = tk.Frame(self._bt_tab, bg=BG2); hdr.pack(fill="x", padx=20, pady=(16,0))
        tk.Label(hdr, text=f"✓  {self._strat_name.get()}",
                 bg=BG2, fg=GREEN,
                 font=(HEAD[0],14,"bold")).pack(side="left")
        ttk.Button(hdr, text="🌐 Open Full Tearsheet",
                   style="Open.TButton",
                   command=self._open_report).pack(side="right")
        ttk.Separator(self._bt_tab).pack(fill="x", padx=20, pady=10)
        grid = tk.Frame(self._bt_tab, bg=BG2); grid.pack(fill="x", padx=20)
        for i, (k,v) in enumerate(result["summary"].items()):
            r, c = divmod(i, 3)
            cell = tk.Frame(grid, bg=BG3)
            cell.grid(row=r, column=c, padx=5, pady=5, sticky="ew")
            grid.columnconfigure(c, weight=1)
            col = GREEN if ("%" in v and not v.startswith("-")) else \
                  RED   if v.startswith("-") else FG
            if "Drawdown" in k:
                col = RED if v.startswith("-") else WARN
            tk.Label(cell, text=k, bg=BG3, fg=FG2,
                     font=("",9)).pack(anchor="w", padx=10, pady=(8,2))
            tk.Label(cell, text=v, bg=BG3, fg=col,
                     font=(HEAD[0],14,"bold")).pack(anchor="w", padx=10, pady=(0,8))
        self._canvas_equity(self._bt_tab, result["daily_rets"], ACCENT,
                            "Equity Curve", height=120, padx=20)
        self._open_btn.configure(state="normal")
        self._nb.select(1)
        self._status.set("Backtest complete. Tearsheet ready.")

    def _open_report(self):
        if self._html_path and os.path.exists(self._html_path):
            webbrowser.open(f"file://{self._html_path}")
        else:
            messagebox.showwarning("No Report",
                                   "Run a backtest first to generate the tearsheet.")

    # ════════════════════════════════════════════════════════════════════════
    # TV ANALYZER UI
    # ════════════════════════════════════════════════════════════════════════

    def _run_tv(self):
        path = self._tv_path.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showerror("File Error",
                                 "Please select a TradingView .xlsx export file.")
            return
        self._prog.start(12)
        self._status.set("Analyzing TradingView trades…")
        self._log_w("═"*60, "info")
        self._log_w(f"TV Analyzer: {os.path.basename(path)}", "info")
        threading.Thread(target=self._tv_task, args=(path,), daemon=True).start()

    def _tv_task(self, path):
        try:
            data = load_tv_xlsx(path)
            d    = data["derived"]
            self._log_w(
                f"Loaded {d['total_trades']} trades | "
                f"Win rate: {d['win_rate']:.1f}% | "
                f"Net P&L: ${d['total_pnl']:,.2f}", "ok")
            self.after(0, lambda: self._build_tv_ui(data))
        except Exception as e:
            self._log_w(f"TV ERROR: {e}", "err")
            self._log_w(traceback.format_exc(), "err")
            self.after(0, lambda: messagebox.showerror("TV Analyzer Error", str(e)))
        finally:
            self.after(0, self._prog.stop)
            self.after(0, lambda: self._status.set("Analysis complete."))

    def _build_tv_ui(self, data):
        for w in self._tv_tab.winfo_children(): w.destroy()
        nb2 = ttk.Notebook(self._tv_tab)
        nb2.pack(fill="both", expand=True)

        dash = tk.Frame(nb2, bg=BG2); nb2.add(dash, text="  Dashboard  ")
        self._tv_dashboard(dash, data)

        tbl  = tk.Frame(nb2, bg=BG2); nb2.add(tbl,  text="  Trade Table  ")
        self._tv_table(tbl, data["trades"])

        eq   = tk.Frame(nb2, bg=BG2); nb2.add(eq,   text="  Equity & Drawdown  ")
        self._tv_equity(eq, data["derived"])

        dist = tk.Frame(nb2, bg=BG2); nb2.add(dist,  text="  Distribution  ")
        self._tv_dist(dist, data["trades"])

        prop = tk.Frame(nb2, bg=BG2); nb2.add(prop,  text="  Properties  ")
        self._tv_props(prop, data)

        self._nb.select(2)

    # ── Dashboard ─────────────────────────────────────────────────────────
    def _tv_dashboard(self, parent, data):
        d    = data["derived"]
        risk = data["risk"]
        props= data["props"]

        cvs  = tk.Canvas(parent, bg=BG2, bd=0, highlightthickness=0)
        vsb  = ttk.Scrollbar(parent, orient="vertical", command=cvs.yview)
        inn  = tk.Frame(cvs, bg=BG2)
        inn.bind("<Configure>",
                 lambda e: cvs.configure(scrollregion=cvs.bbox("all")))
        cvs.create_window((0,0), window=inn, anchor="nw")
        cvs.configure(yscrollcommand=vsb.set)
        cvs.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        sym = props.get("Symbol",("—",))[0]
        tf  = props.get("Timeframe",("—",))[0]
        rng = props.get("Trading range",("—",))[0]
        tk.Label(inn, text="📊  TradingView Strategy Analysis",
                 bg=BG2, fg=PURPLE,
                 font=(HEAD[0],15,"bold")).pack(anchor="w", padx=20, pady=(16,4))
        tk.Label(inn, text=f"  {sym}  ·  {tf}  ·  {rng}",
                 bg=BG2, fg=FG2, font=("",9)).pack(anchor="w", padx=20, pady=(0,12))
        ttk.Separator(inn).pack(fill="x", padx=20)

        def kpi_grid(frame, title, items, ncols=4):
            tk.Label(frame, text=title, bg=BG2, fg=FG2,
                     font=(HEAD[0],10,"bold")).pack(anchor="w", padx=20, pady=(14,6))
            g = tk.Frame(frame, bg=BG2); g.pack(fill="x", padx=20)
            for i,(lbl,val,col) in enumerate(items):
                r,c = divmod(i, ncols)
                cell = tk.Frame(g, bg=BG3)
                cell.grid(row=r, column=c, padx=5, pady=5, sticky="ew")
                g.columnconfigure(c, weight=1)
                tk.Label(cell, text=lbl, bg=BG3, fg=FG2,
                         font=("",8)).pack(anchor="w", padx=10, pady=(8,2))
                tk.Label(cell, text=str(val), bg=BG3, fg=col,
                         font=(HEAD[0],15,"bold")).pack(
                             anchor="w", padx=10, pady=(0,8))

        net   = d["total_pnl"]
        pf    = d["profit_factor"]
        wr    = d["win_rate"]
        mdd   = d["max_drawdown"]

        kpi_grid(inn, "PERFORMANCE", [
            ("Net P&L",       f"${net:,.2f}",            GREEN if net >= 0 else RED),
            ("Gross Profit",  f"${d['gross_profit']:,.2f}", GREEN),
            ("Gross Loss",    f"${abs(d['gross_loss']):,.2f}", RED),
            ("Profit Factor", f"{pf:.3f}" if pf != float('inf') else "∞",
             GREEN if pf >= 1 else RED),
            ("Max Drawdown",  f"${mdd:,.2f}",             RED if mdd < 0 else WARN),
            ("Win Rate",      f"{wr:.1f}%",               GREEN if wr >= 50 else WARN),
            ("Avg Win",       f"${d['avg_win']:,.2f}",    GREEN),
            ("Avg Loss",      f"${abs(d['avg_loss']):,.2f}", RED),
        ])

        kpi_grid(inn, "TRADE STATISTICS", [
            ("Total Trades",  str(d["total_trades"]),  FG),
            ("Winners",       str(d["winners"]),       GREEN),
            ("Losers",        str(d["losers"]),        RED),
            ("Even",          str(d["evens"]),         FG2),
            ("Largest Win",   f"${d['largest_win']:,.2f}",          GREEN),
            ("Largest Loss",  f"${abs(d['largest_loss']):,.2f}",    RED),
            ("Avg Duration",
             str(d["avg_duration"]).split(".")[0]
             if pd.notna(d["avg_duration"]) else "—", FG),
            ("",              "",                      BG3),
        ])

        def _r(key):
            v = risk.get(key, (None,))
            return f"{v[0]:.3f}" if v and v[0] is not None else "—"

        kpi_grid(inn, "RISK-ADJUSTED METRICS", [
            ("Sharpe Ratio",  _r("Sharpe ratio"),  ACCENT),
            ("Sortino Ratio", _r("Sortino ratio"), ACCENT),
            ("Profit Factor", _r("Profit factor"), ACCENT),
            ("",              "",                  BG3),
        ])

        tk.Label(inn, text="EQUITY CURVE", bg=BG2, fg=FG2,
                 font=(HEAD[0],10,"bold")).pack(anchor="w", padx=20, pady=(14,4))
        self._canvas_equity(inn, d["equity"], PURPLE, "", height=140, padx=20)

        tk.Label(inn, text="DRAWDOWN", bg=BG2, fg=FG2,
                 font=(HEAD[0],10,"bold")).pack(anchor="w", padx=20, pady=(12,4))
        self._canvas_line(inn, d["drawdown"], RED, height=80, padx=20)
        tk.Frame(inn, bg=BG2, height=20).pack()

    # ── Trade Table ───────────────────────────────────────────────────────
    def _tv_table(self, parent, trades):
        tk.Label(parent, text="  All Closed Trades",
                 bg=BG2, fg=PURPLE,
                 font=(HEAD[0],12,"bold")).pack(anchor="w", padx=16, pady=(12,4))

        fb = tk.Frame(parent, bg=BG2); fb.pack(fill="x", padx=16, pady=(0,6))
        tk.Label(fb, text="Filter:", bg=BG2, fg=FG2,
                 font=("",9)).pack(side="left")
        self._tv_filt = tk.StringVar()
        self._tv_filt.trace_add("write",
            lambda *a: self._tv_filter(trades))
        ttk.Entry(fb, textvariable=self._tv_filt,
                  width=20, font=("",9)).pack(side="left", padx=6)
        tk.Label(fb, text="(any column value — case-insensitive)",
                 bg=BG2, fg=FG2, font=("",8)).pack(side="left")

        cols   = ["#","Entry Time","Exit Time","Dir","Signal",
                  "Entry $","Exit $","Qty","P&L $","P&L %",
                  "MFE $","MAE $","Duration","Cum P&L $"]
        widths = [38,132,132,55,130,68,68,55,75,62,62,62,100,80]
        frm    = tk.Frame(parent, bg=BG2)
        frm.pack(fill="both", expand=True, padx=16, pady=(0,12))
        xsb = ttk.Scrollbar(frm, orient="horizontal")
        ysb = ttk.Scrollbar(frm, orient="vertical")
        self._tv_tree = ttk.Treeview(
            frm, columns=cols, show="headings",
            yscrollcommand=ysb.set, xscrollcommand=xsb.set)
        ysb.configure(command=self._tv_tree.yview)
        xsb.configure(command=self._tv_tree.xview)
        xsb.pack(side="bottom", fill="x")
        ysb.pack(side="right",  fill="y")
        self._tv_tree.pack(fill="both", expand=True)

        for col, w in zip(cols, widths):
            self._tv_tree.heading(col, text=col,
                command=lambda c=col: self._tv_sort(c))
            self._tv_tree.column(col, width=w, stretch=False,
                anchor="w" if col in ("Entry Time","Exit Time","Dir","Signal") else "e")

        self._tv_tree.tag_configure("win",  foreground=GREEN)
        self._tv_tree.tag_configure("loss", foreground=RED)
        self._tv_tree.tag_configure("even", foreground=FG2)
        self._tv_raw_trades = trades
        self._tv_fill(trades)

    def _tv_fill(self, trades):
        self._tv_tree.delete(*self._tv_tree.get_children())
        for _, row in trades.iterrows():
            pnl = row.get("pnl_usd", 0) or 0
            tag = "win" if pnl > 0 else ("loss" if pnl < 0 else "even")
            vals = [
                int(row["trade_num"]) if pd.notna(row.get("trade_num")) else "",
                str(row.get("datetime_entry",""))[:16],
                str(row.get("datetime_exit",""))[:16],
                str(row.get("direction","")).capitalize(),
                str(row.get("signal_exit", row.get("signal",""))),
                f"{row.get('price_entry',0):.4f}" if pd.notna(row.get("price_entry")) else "",
                f"{row.get('price_exit',0):.4f}"  if pd.notna(row.get("price_exit"))  else "",
                f"{int(row.get('qty',0)):,}"       if pd.notna(row.get("qty"))         else "",
                f"${pnl:,.2f}",
                f"{row.get('pnl_pct',0):.2f}%"    if pd.notna(row.get("pnl_pct"))     else "",
                f"${row.get('mfe_usd',0):,.2f}"    if pd.notna(row.get("mfe_usd"))     else "",
                f"${row.get('mae_usd',0):,.2f}"    if pd.notna(row.get("mae_usd"))     else "",
                str(row.get("duration","")).split(".")[0],
                f"${row.get('cum_pnl_usd',0):,.2f}" if pd.notna(row.get("cum_pnl_usd")) else "",
            ]
            self._tv_tree.insert("","end", values=vals, tags=(tag,))

    def _tv_filter(self, trades):
        q = self._tv_filt.get().lower().strip()
        if not q:
            self._tv_fill(trades); return
        filt = trades[trades.apply(
            lambda r: q in str(r.to_dict()).lower(), axis=1)]
        self._tv_fill(filt)

    def _tv_sort(self, col):
        data = [(self._tv_tree.set(c, col), c)
                for c in self._tv_tree.get_children("")]
        try:
            data.sort(key=lambda x: float(
                x[0].replace("$","").replace("%","").replace(",","")))
        except ValueError:
            data.sort()
        for i, (_,c) in enumerate(data):
            self._tv_tree.move(c,"",i)

    # ── Equity & Drawdown ─────────────────────────────────────────────────
    def _tv_equity(self, parent, d):
        cvs = tk.Canvas(parent, bg=BG2, bd=0, highlightthickness=0)
        vsb = ttk.Scrollbar(parent, orient="vertical", command=cvs.yview)
        inn = tk.Frame(cvs, bg=BG2)
        inn.bind("<Configure>",
                 lambda e: cvs.configure(scrollregion=cvs.bbox("all")))
        cvs.create_window((0,0), window=inn, anchor="nw")
        cvs.configure(yscrollcommand=vsb.set)
        cvs.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        for title, series, col, h in [
            ("Cumulative Equity (USD)", d["equity"],   PURPLE, 180),
            ("Daily P&L (USD)",         d["daily_pnl"],ORANGE, 120),
            ("Drawdown from Peak (USD)",d["drawdown"], RED,    110),
        ]:
            tk.Label(inn, text=title, bg=BG2, fg=col,
                     font=(HEAD[0],12,"bold")).pack(anchor="w", padx=20, pady=(16,4))
            if title.startswith("Daily"):
                self._canvas_bars(inn, series, height=h, padx=20)
            else:
                (self._canvas_equity if "Equity" in title else self._canvas_line)(
                    inn, series, col, "", height=h, padx=20)

        # Stats
        eq   = d["equity"]
        dd   = d["drawdown"]
        dpnl = d["daily_pnl"]
        tk.Label(inn, text="Key Metrics", bg=BG2, fg=FG2,
                 font=(HEAD[0],10,"bold")).pack(anchor="w", padx=20, pady=(14,6))
        g = tk.Frame(inn, bg=BG2); g.pack(fill="x", padx=20, pady=(0,20))
        items = [
            ("Final Equity",      f"${eq.iloc[-1]:,.2f}",              GREEN),
            ("Peak Equity",       f"${eq.max():,.2f}",                  GREEN),
            ("Max Drawdown $",    f"${dd.min():,.2f}",                  RED),
            ("Max DD %",          f"{dd.min()/max(eq.max(),1)*100:.2f}%", RED),
            ("Best Day",          f"${dpnl.max():,.2f}",                GREEN),
            ("Worst Day",         f"${dpnl.min():,.2f}",                RED),
            ("Avg Daily P&L",     f"${dpnl.mean():,.2f}",               FG),
            ("Profitable Days",   f"{(dpnl>0).sum()} / {len(dpnl)}",   GREEN),
        ]
        for i,(k,v,c) in enumerate(items):
            r,col_ = divmod(i,4)
            cell = tk.Frame(g, bg=BG3)
            cell.grid(row=r, column=col_, padx=5, pady=5, sticky="ew")
            g.columnconfigure(col_, weight=1)
            tk.Label(cell, text=k, bg=BG3, fg=FG2,
                     font=("",8)).pack(anchor="w", padx=8, pady=(7,1))
            tk.Label(cell, text=v, bg=BG3, fg=c,
                     font=(HEAD[0],12,"bold")).pack(anchor="w", padx=8, pady=(0,7))

    # ── Distribution ──────────────────────────────────────────────────────
    def _tv_dist(self, parent, trades):
        cvs = tk.Canvas(parent, bg=BG2, bd=0, highlightthickness=0)
        vsb = ttk.Scrollbar(parent, orient="vertical", command=cvs.yview)
        inn = tk.Frame(cvs, bg=BG2)
        inn.bind("<Configure>",
                 lambda e: cvs.configure(scrollregion=cvs.bbox("all")))
        cvs.create_window((0,0), window=inn, anchor="nw")
        cvs.configure(yscrollcommand=vsb.set)
        cvs.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        pnl    = trades["pnl_usd"].dropna()
        wins   = pnl[pnl > 0]
        losses = pnl[pnl < 0]
        evens  = pnl[pnl == 0]
        total  = max(len(pnl), 1)

        tk.Label(inn, text="P&L Distribution", bg=BG2, fg=PURPLE,
                 font=(HEAD[0],13,"bold")).pack(anchor="w", padx=20, pady=(16,8))

        self._stacked_bar(inn, len(wins)/total, len(evens)/total, len(losses)/total)

        tk.Label(inn, text="P&L Histogram (USD per trade)", bg=BG2, fg=FG2,
                 font=(HEAD[0],10,"bold")).pack(anchor="w", padx=20, pady=(14,4))
        self._histogram(inn, pnl.tolist(), height=180, padx=20)

        tk.Label(inn, text="Streak Analysis", bg=BG2, fg=FG2,
                 font=(HEAD[0],10,"bold")).pack(anchor="w", padx=20, pady=(14,4))
        s  = self._streaks(pnl)
        sg = tk.Frame(inn, bg=BG2); sg.pack(fill="x", padx=20)
        for i,(k,v,c) in enumerate([
            ("Max Win Streak",  str(s["max_win"]),        GREEN),
            ("Max Loss Streak", str(s["max_loss"]),       RED),
            ("Current Streak",  str(s["current"]),        FG),
            ("Avg Win Streak",  f"{s['avg_win']:.1f}",   GREEN),
        ]):
            cell = tk.Frame(sg, bg=BG3)
            cell.grid(row=0, column=i, padx=5, pady=5, sticky="ew")
            sg.columnconfigure(i, weight=1)
            tk.Label(cell, text=k, bg=BG3, fg=FG2,
                     font=("",8)).pack(anchor="w", padx=8, pady=(7,1))
            tk.Label(cell, text=v, bg=BG3, fg=c,
                     font=(HEAD[0],14,"bold")).pack(anchor="w", padx=8, pady=(0,7))

        tk.Label(inn, text="Monthly P&L Breakdown", bg=BG2, fg=FG2,
                 font=(HEAD[0],10,"bold")).pack(anchor="w", padx=20, pady=(14,4))
        self._monthly_table(inn, trades)
        tk.Frame(inn, bg=BG2, height=20).pack()

    def _stacked_bar(self, parent, wr, er, lr):
        f = tk.Frame(parent, bg=BG2); f.pack(fill="x", padx=20, pady=(0,4))
        bar = tk.Canvas(f, height=28, bg=BG3, highlightthickness=0)
        bar.pack(fill="x"); bar.update()
        W = bar.winfo_width() or 800
        xw = int(W * wr); xe = int(W * (wr+er))
        if xw:    bar.create_rectangle(0,  0, xw, 28, fill=GREEN,  outline="")
        if xe-xw: bar.create_rectangle(xw, 0, xe, 28, fill=FG2,    outline="")
        if W-xe:  bar.create_rectangle(xe, 0, W,  28, fill=RED,    outline="")
        bar.create_text(max(xw//2,20), 14, fill=BG,
                        text=f"W {wr*100:.0f}%", font=("",8,"bold"))
        if xe-xw > 30:
            bar.create_text(xw+(xe-xw)//2, 14, fill=BG,
                            text=f"E {er*100:.0f}%", font=("",8))
        bar.create_text(xe+(W-xe)//2 if W-xe>20 else W-25, 14, fill=BG,
                        text=f"L {lr*100:.0f}%", font=("",8,"bold"))
        lab = tk.Frame(f, bg=BG2); lab.pack(fill="x")
        for txt,col in [("■ Wins",GREEN),("■ Even",FG2),("■ Losses",RED)]:
            tk.Label(lab, text=txt, bg=BG2, fg=col, font=("",8)).pack(side="left", padx=8)

    def _histogram(self, parent, values, height=160, padx=20):
        if not values: return
        f = tk.Frame(parent, bg=BG2); f.pack(fill="x", padx=padx, pady=4)
        c = tk.Canvas(f, height=height, bg=BG3, highlightthickness=0)
        c.pack(fill="x"); c.update()
        W   = c.winfo_width() or 800
        mn, mx = min(values), max(values)
        bins = 30; rng = (mx - mn) or 1e-9; bw = rng / bins
        counts = [0]*bins
        for v in values:
            counts[min(int((v-mn)/bw), bins-1)] += 1
        mc = max(counts) or 1
        BAR = W/bins; PAD = 10
        for i,cnt in enumerate(counts):
            x1 = i*BAR; x2 = x1+BAR-1
            bs = mn+i*bw
            col = GREEN if bs >= 0 else RED
            y1 = height-PAD-(cnt/mc)*(height-2*PAD)
            c.create_rectangle(x1, y1, x2, height-PAD, fill=col, outline="")
        c.create_line(0, height-PAD, W, height-PAD, fill=BORDER)
        zx = (-mn/rng)*W
        if 0 <= zx <= W:
            c.create_line(zx, 0, zx, height, fill=FG2, dash=(4,4))
        c.create_text(6, height-3, text=f"${mn:.2f}",
                      fill=FG2, anchor="sw", font=("",7))
        c.create_text(W-4, height-3, text=f"${mx:.2f}",
                      fill=FG2, anchor="se", font=("",7))

    def _streaks(self, pnl):
        max_w = max_l = cw = cl = 0
        ws = []
        for v in pnl:
            if v > 0:
                cw += 1; max_w = max(max_w, cw)
                if cl: cl = 0
            elif v < 0:
                cl += 1; max_l = max(max_l, cl)
                if cw: ws.append(cw); cw = 0
        last = pnl.iloc[-1] if len(pnl) else 0
        cur  = f"+{cw}W" if cw else (f"-{cl}L" if cl else "0")
        return dict(max_win=max_w, max_loss=max_l,
                    current=cur, avg_win=float(np.mean(ws)) if ws else 0.0)

    def _monthly_table(self, parent, trades):
        if "datetime_exit" not in trades.columns: return
        t = trades.copy()
        t["month"] = pd.to_datetime(
            t["datetime_exit"], errors="coerce").dt.to_period("M")
        mo = t.groupby("month").agg(
            n=("pnl_usd","count"),
            pnl=("pnl_usd","sum"),
            wins=("pnl_usd", lambda x: (x>0).sum()),
        ).reset_index()
        mo["wr"] = mo["wins"]/mo["n"]*100
        f = tk.Frame(parent, bg=BG2); f.pack(fill="x", padx=20, pady=(0,6))
        cols = ["Month","Trades","Win%","P&L $"]
        tree = ttk.Treeview(f, columns=cols, show="headings", height=8)
        for col, w in zip(cols, [100,70,70,100]):
            tree.heading(col, text=col)
            tree.column(col, width=w, anchor="w" if col=="Month" else "e")
        for _,r in mo.iterrows():
            pnl = r["pnl"]
            tag = "mw" if pnl >= 0 else "ml"
            tree.insert("","end", tags=(tag,),
                values=[str(r["month"]), int(r["n"]),
                        f"{r['wr']:.1f}%", f"${pnl:,.2f}"])
        tree.tag_configure("mw", foreground=GREEN)
        tree.tag_configure("ml", foreground=RED)
        tree.pack(fill="x")

    # ── Properties ────────────────────────────────────────────────────────
    def _tv_props(self, parent, data):
        cvs = tk.Canvas(parent, bg=BG2, bd=0, highlightthickness=0)
        vsb = ttk.Scrollbar(parent, orient="vertical", command=cvs.yview)
        inn = tk.Frame(cvs, bg=BG2)
        inn.bind("<Configure>",
                 lambda e: cvs.configure(scrollregion=cvs.bbox("all")))
        cvs.create_window((0,0), window=inn, anchor="nw")
        cvs.configure(yscrollcommand=vsb.set)
        cvs.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        tk.Label(inn, text="Strategy Properties & All Summary Sheets",
                 bg=BG2, fg=PURPLE,
                 font=(HEAD[0],13,"bold")).pack(anchor="w", padx=20, pady=(16,10))

        for name, sheet, col in [
            ("Properties",               data["props"],  ACCENT),
            ("Performance",              data["perf"],   GREEN),
            ("Trades Analysis",          data["tana"],   ORANGE),
            ("Risk-Adjusted Performance",data["risk"],   PURPLE),
        ]:
            if not sheet: continue
            tk.Label(inn, text=name, bg=BG2, fg=col,
                     font=(HEAD[0],10,"bold")).pack(anchor="w", padx=20, pady=(14,4))
            f = tk.Frame(inn, bg=BG3)
            f.pack(fill="x", padx=20, pady=(0,6))
            for key, vals in sheet.items():
                if key in ("name","value"): continue
                r = tk.Frame(f, bg=BG3); r.pack(fill="x")
                tk.Label(r, text=str(key), bg=BG3, fg=FG2,
                         font=("",9), width=36, anchor="w"
                         ).pack(side="left", padx=(10,4), pady=2)
                v_str = "  |  ".join(
                    str(v) for v in vals if v is not None) or "—"
                tk.Label(r, text=v_str, bg=BG3, fg=FG,
                         font=MONO, anchor="w").pack(side="left", padx=4)
            ttk.Separator(f).pack(fill="x")
        tk.Frame(inn, bg=BG2, height=20).pack()

    # ════════════════════════════════════════════════════════════════════════
    # SHARED CANVAS DRAWING
    # ════════════════════════════════════════════════════════════════════════

    def _canvas_equity(self, parent, series, color, title="",
                        height=120, padx=16):
        if series is None or len(series) == 0: return
        f = tk.Frame(parent, bg=BG2); f.pack(fill="x", padx=padx, pady=4)
        c = tk.Canvas(f, height=height, bg=BG3, highlightthickness=0, bd=0)
        c.pack(fill="x"); c.update()
        W = c.winfo_width() or 800; PAD = 12
        vals = np.array(series.values, dtype=float)
        mn, mx = vals.min(), vals.max(); rng = (mx-mn) or 1e-9

        def xy(i,v):
            x = PAD+(i/max(len(vals)-1,1))*(W-2*PAD)
            y = height-PAD-((v-mn)/rng)*(height-2*PAD)
            return x, y

        pts = [xy(i,v) for i,v in enumerate(vals)]
        flat = [n for pt in pts for n in pt]
        for i in range(len(pts)-1):
            x1,y1=pts[i]; x2,y2=pts[i+1]
            c.create_polygon(x1,y1,x2,y2,x2,height-PAD,x1,height-PAD,
                             fill="#1a3a5c",outline="")
        if len(flat) >= 4:
            c.create_line(*flat, fill=color, width=2, smooth=True)
        c.create_text(W-PAD, PAD,       text=f"{mx:,.2f}", fill=FG2,
                      anchor="ne", font=("",7))
        c.create_text(W-PAD, height-4,  text=f"{mn:,.2f}", fill=FG2,
                      anchor="se", font=("",7))

    def _canvas_line(self, parent, series, color, title="",
                      height=100, padx=16):
        if series is None or len(series) == 0: return
        f = tk.Frame(parent, bg=BG2); f.pack(fill="x", padx=padx, pady=4)
        c = tk.Canvas(f, height=height, bg=BG3, highlightthickness=0, bd=0)
        c.pack(fill="x"); c.update()
        W = c.winfo_width() or 800; PAD = 10
        vals = np.array(series.values, dtype=float)
        mn, mx = vals.min(), vals.max(); rng = (mx-mn) or 1e-9
        pts = [PAD+(i/max(len(vals)-1,1))*(W-2*PAD) for i in range(len(vals))]
        ys  = [height-PAD-((v-mn)/rng)*(height-2*PAD) for v in vals]
        flat = [n for xy in zip(pts,ys) for n in xy]
        if len(flat) >= 4:
            c.create_line(*flat, fill=color, width=1.5, smooth=True)
        c.create_text(W-PAD, PAD+2,    text=f"{mx:,.2f}", fill=FG2,
                      anchor="ne", font=("",7))
        c.create_text(W-PAD, height-4, text=f"{mn:,.2f}", fill=FG2,
                      anchor="se", font=("",7))

    def _canvas_bars(self, parent, series, height=100, padx=16):
        if series is None or len(series) == 0: return
        f = tk.Frame(parent, bg=BG2); f.pack(fill="x", padx=padx, pady=4)
        c = tk.Canvas(f, height=height, bg=BG3, highlightthickness=0, bd=0)
        c.pack(fill="x"); c.update()
        W = c.winfo_width() or 800; PAD = 10
        vals = np.array(series.values, dtype=float)
        rng  = max(abs(vals.min()), abs(vals.max())) or 1e-9
        bw   = max((W-2*PAD)/len(vals)-1, 2)
        mid  = height//2
        for i,v in enumerate(vals):
            x1 = PAD+i*((W-2*PAD)/len(vals)); x2 = x1+bw
            h_ = abs(v)/rng*(height//2-PAD)
            col = GREEN if v >= 0 else RED
            if v >= 0: c.create_rectangle(x1, mid-h_, x2, mid, fill=col, outline="")
            else:      c.create_rectangle(x1, mid, x2, mid+h_, fill=col, outline="")
        c.create_line(PAD, mid, W-PAD, mid, fill=BORDER, width=1)


# ════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    App().mainloop()
